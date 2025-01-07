
import os
from openpyxl import Workbook, load_workbook
import arcpy
import logging
import traceback
import multiprocessing as mp
from mp_worker import process_job_mp
from aoi_utilities import build_aoi_from_shp
from aoi_utilities import build_aoi_from_kml


class BATCH_FACTORY:
    ''' Batch factory class reads a spreadsheet to gather parameters and then batch runs any given tool '''
    XLSX_SHEET_NAME = 'ast_config'
    BATCH_PARAMETERS = {}
    

    
    BATCH_CONDITION_COLUMN = 'batch_condition'
    # DONT_OVERWRITE_OUTPUTS = 'dont_overwrite_outputs'
    BATCH_SCRIPT = ''
    job_index = None  # Initialize job_index as a global variable
    
    def __init__(self, queuefile, db_user, db_pass, logger=None, current_path=None) -> None:
            self.user = db_user
            self.user_cred = db_pass
            self.queuefile = queuefile
            self.jobs = []
            self.logger = logger or logging.getLogger(__name__)
            self.current_path = current_path
            self.parameter_names = []   
#LOAD JOBS
    def load_jobs(self):
        '''
        load jobs will check for the existence of the queuefile, if it exists it will load the jobs from the queuefile. Checking if they 
        are Complete and if not, it will add them to the jobs  as Queued
        '''
        # NOTE pass job index into load jobs function
        #global job_index
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Loading Jobs...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")

        # Initialize the jobs list to store jobs
        self.jobs = []

        # Check if the queue file exists
        assert os.path.exists(self.queuefile), "Queue file does not exist"
        if os.path.exists(self.queuefile):

            try:
                print(f'Queue file exists....Queue file is {self.queuefile}')
                # Open the Excel workbook and select the correct sheet
                print("Loading workbook.....")
                wb = load_workbook(filename=self.queuefile)
                print(f'Workbook loaded is {wb}')
                
                print("Loading worksheet.....")
                ws = wb[self.XLSX_SHEET_NAME]
                
                # Check if the sheet exists in the workbook
            
                
                
                if ws not in wb.sheetnames:
                    raise ValueError(f"'{self.XLSX_SHEET_NAME}' sheet not found in the workbook.")
                
                # Get the header row (parameter names)
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if not header_row or not any(header_row):
                    raise ValueError("No header found in the spreadsheet.")

                # Convert None headers to empty strings if needed
                header = [h if h is not None else "" for h in header_row]
                
                # Store these header names in self.parameter_names so we can reference them later
                # but exclude any known special columns like self.BATCH_CONDITION_COLUMN:
                self.parameter_names = []
                for col_name in header:
                    if col_name and col_name != self.BATCH_CONDITION_COLUMN:
                        self.parameter_names.append(col_name)
                        
                self.logger.info(f"Identified parameter names: {self.parameter_names}")
                
                self.jobs = []  # clear existing jobs

                # Create the dictionary the holds the keys (header names/parameters) and values (job details) for each job
                # Read all the data rows (starting from the second row to skip the header)
                data = []
                
                for row_idx, row_data in enumerate(ws.iter_rows(min_row=2, max_col=None, values_only=True), start=2):
                          # If the row is entirely blank, skip it
                    if all((value is None or str(value).strip() == '') for value in row_data):
                        continue
                    print(f'Row index is {row_idx} and row is {row_data}')
                    # data.append(row)
                    
                    # Create a dictionary to store the row data
                    job = {}
                    for col_idx, cell_value in enumerate(row_data):
                        # Only process cells that match up to existing header columns
                        if col_idx < len(header):
                            param_name = header[col_idx]
                            if param_name:
                            # Store the cell_value in the dictionary
                                job[param_name] = cell_value
                                print(f'Job dictionary is {job}')
                
                
                                                    # # Iterate over each row of data; enumerate to keep track of the row number in Excel
                                                    # for job_index, row_data in enumerate(data):  

                                                    #     # Dictionary where key is index key is Job number dictionary is the dictionary of jobs
                                                    #     # Send job to processer and include status
                                                    #     self.logger.info(f"\n")
                                                    #     self.logger.info(f"-------------------------------------------------------------------------------")
                                                    #     self.logger.info(f"-                        Load Jobs: Start of Job {job_index}                               -")
                                                    #     self.logger.info(f"-------------------------------------------------------------------------------")
                                                    #     self.logger.info(f"\n")
                                                        
                                                    #     # Initialize a dictionary to store the job's parameters
                                                    #     job = {}
                                                    #     self.logger.info('Load Jobs - Creating empty dictionary')
                                                    #     batch_condition = None  # Initialize the batch_condition for the current row
                                                            
                                                        
                                                    #     # Skip any completely blank rows
                                                    #     if all((value is None or str(value).strip() == '') for value in row_data):
                                                    #         print(f"Load Jobs - Skipping blank row at job index ({job_index}) ")
                                                    #         self.logger.info(f"Load Jobs - Skipping blank row at index ({job_index}) ")
                                                    #         continue  # Skip this row entirely


                    # Loop through each column header and corresponding value in the current row
                    for key, value in zip(header, row_data):
                        # If the key isn't empty and equals "batch_condition", assign the value to batch_condition
                        if key is not None and key.lower() == self.BATCH_CONDITION_COLUMN.lower():
                            batch_condition = value if value is not None else ""

                        # Assign an empty string to any None values
                        value = "" if value is None else value

                        # Assign the value to the job dictionary if the key is not None
                        if key is not None:
                            job[key] = value

                    # Skip if marked as "COMPLETE"
                    batch_condition = job.get(self.BATCH_CONDITION_COLUMN, "")
                    if batch_condition.upper() == 'COMPLETE':
                        print(f"Skipping job {row_idx} as it is marked COMPLETE.")
                        self.logger.info(f"Load Jobs - Skipping job {row_idx} as it is marked COMPLETE.")
                        # continue  # Skip this job as it's already marked as COMPLETE

                    # Check if the batch_condition is None, empty, or not 'COMPLETE' assign it to queued
                    if batch_condition is None or batch_condition.strip() == '' or batch_condition.upper() != 'COMPLETE':
                        # Assign 'Queued' to the batch_condition and update the job dictionary
                        batch_condition = 'Queued'
                        
                        # Assign the updated batch_condition to the job dictionary (queued)
                        job[self.BATCH_CONDITION_COLUMN] = batch_condition
                        self.logger.info(f"Load Jobs - (Queued assigned to Job ({job_index}) is ({batch_condition})")

                        # Immediately update the Excel sheet with the new condition
                        #LOAD JOBS ADD_JOB_RESULT FUNCTION IS CALLED HERE
                        try:
                            self.add_job_result(row_idx, batch_condition)
                            self.logger.info(f"Load Jobs - Added job condition '{batch_condition}' for job {row_idx} to jobs list")
                        except Exception as e:
                            print(f"Error updating Excel sheet at row {row_idx}: {e}")
                            self.logger.error(f"Load Jobs - Error updating Excel sheet at row {row_idx}: {e}")
                            continue

                        # Classify the input type for the job
                        try:
                            self.logger.info(f"Classifying input type for job {row_idx}")
                            self.classify_input_type(job)
 
                        except Exception as e:
                            print(f"Error classifying input type for job {job}: {e}")
                            self.logger.error(f"Error classifying input type for job {job}: {e}")
                            
                    # Add the job to the jobs list after all checks and processing
                    self.jobs.append(job)
                    print(f"Load Jobs - Job Condition for job ({row_idx}) is not Complete: Writing ({batch_condition}) to batch_contion. Adding job: {row_idx} to jobs list")
                    self.logger.info(f"Load Jobs - Job Condition is not Complete ({batch_condition}), adding job: {row_idx} to jobs list")

                    # print(f"Load Jobs - Job dictionary is {job}")
                    # self.logger.info(f"Load Jobs - Job {job_index} dictionary is {job}")
                    self.logger.info(f"\n")
                    self.logger.info(f"-------------------------------------------------------------------------------")
                    self.logger.info(f"-                        End of Job {row_idx}                                -")
                    self.logger.info(f"-------------------------------------------------------------------------------")
                    self.logger.info(f"\n")
                    
                    
            except FileNotFoundError as e:
                print(f"Error: Queue file not found - {e}")
                self.logger.error(f"Error: Queue file not found - {e}")
            except Exception as e:
                print(f"Unexpected error loading jobs: {e}")
                self.logger.error(f"Unexpected error loading jobs: {e}")

            return self.jobs


    def classify_input_type(self, job):
        '''Classify the input type and process accordingly.'''

        if job.get('feature_layer'):
            print(f'Feature layer found: {job["feature_layer"]}')
            self.logger.info(f'Classifying Input Type - Feature layer found: {job["feature_layer"]}')
            feature_layer_path = job['feature_layer']
            print(f"Processing feature layer: {feature_layer_path}")
            self.logger.info(f"Classifying Input Type - Processing feature layer: {feature_layer_path}")

            if feature_layer_path.lower().endswith('.kml'):
                print('KML found, building AOI from KML')
                self.logger.info('Classifying Input Type - KML found, building AOI from KML')
                job['feature_layer'] = build_aoi_from_kml(job, feature_layer_path)

            elif feature_layer_path.lower().endswith('.shp'):
                if job.get('file_number'):
                    print(f"File number found, running FW setup on shapefile: {feature_layer_path}")
                    self.logger.info(f"Classifying Input Type - File number found, running FW setup on shapefile: {feature_layer_path}")
                    new_feature_layer_path = build_aoi_from_shp(job, feature_layer_path)
                    job['feature_layer'] = new_feature_layer_path
                else:
                    print('No FW File Number provided for the shapefile, using original shapefile path')
                    self.logger.info('Classifying Input Type - No FW File Number provided, using original shapefile path')
            else:
                print(f"Unsupported feature layer format: {feature_layer_path}")
                self.logger.warning(f"Classifying Input Type - Unsupported feature layer format: {feature_layer_path} - Marking job as Failed")
                self.add_job_result(job, 'Failed')
        else:
            print('No feature layer provided in job')
            self.logger.warning('Classifying Input Type - No feature layer provided in job')

#ADD JOB RESULT                        
    def add_job_result(self, job_index, condition):
        ''' 
        Function adds result information to the Excel spreadsheet. If the job is successful, it will update the batch_condition column to "COMPLETE",
        if the job failed, it will update the batch_condition column to "Failed".
        '''

        print("Running Add Job Results...")
        self.logger.info("\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Running Add Job Results from Load Jobs Function")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info("\n")
     

        try:
            # Load the workbook
            wb = load_workbook(filename=self.queuefile)
            self.logger.info(f"Add Job Result - Workbook loaded")
            
            # Load the correct worksheet
            ws = wb[self.XLSX_SHEET_NAME]

            # Read the header index for the batch_condition column
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            
            # Check if 'BATCH CONDITION COLUMN' exists in the header. If it is not found, raise a ValueError
            # if batch condition column IS found, log a message
            if self.BATCH_CONDITION_COLUMN not in header:
                raise ValueError(f"'{self.BATCH_CONDITION_COLUMN}' column not found in the spreadsheet.")
            
            if self.BATCH_CONDITION_COLUMN in header:
                self.logger.info(f"Add Job Result - '{self.BATCH_CONDITION_COLUMN}' column found in the spreadsheet.")
                

            
            # Find the job condition column and assign it to the correct index
            batch_condition_index = header.index(self.BATCH_CONDITION_COLUMN) + 1  # +1 because Excel columns are 1-indexed

    
            
            # Calculate the actual row index in Excel, +2 to account for header and 0-index
            excel_row_index = job_index + 2  # NOTE I changed this to +1 and it changes the batch_condition header row to Failed. So it must stay at +2
            self.logger.info(f"Add Job Result - Calculated Excel row index as {excel_row_index} for job index {job_index}")
            
            # Check if the row is blank before updating,  If all cell values in row_values are either None or empty strings, then all() will return True, indicating that the row is blank.
            row_values = []
            for col in range(1, len(header) + 1):
                cell_value = ws.cell(row=excel_row_index, column=col).value
                row_values.append(cell_value)
            if all(value is None or str(value).strip() == '' for value in row_values):
                print(f"Row {excel_row_index} is blank, not updating.")
                self.logger.info(f"Add_Job_Result -Job {job_index} / Row {excel_row_index} ")
                return  # Do not update if the row is blank

            # Update the ast condition for the specific job to the new condition (failed, queued, complete)
            ws.cell(row=excel_row_index, column=batch_condition_index, value=condition)

            
        except FileNotFoundError as e:
            print(f"Error: Queue file not found - {e}")
            self.logger.error(f"Error: Queue file not found - {e}")

        except ValueError as e:
            print(f"Error: {e}")
            self.logger.error(f"Error: {e}")

        except IndexError as e:
            print(f"Error: Index out of range when accessing row {excel_row_index} - {e}")
            self.logger.error(f"Error: Index out of range when accessing row {excel_row_index} - {e}")

        except PermissionError as e:
            print(f"Error: Permission denied when trying to access the Excel file - {e}")
            self.logger.error(f"Error: Permission denied when trying to access the Excel file - {e}")

        except Exception as e:
            print(f"Unexpected error while adding job result: {e}")
            self.logger.error(f"Unexpected error while adding job result: {e}")

#BATCH AST
    def batch_ast(self):
        '''
        Uses multiprocessing to run the NUMBER_OF_JOBS in parallel.
        '''
        self.logger.info(f"\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Batch AST: Batching Jobs with Multiprocessing...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info(f"\n")
        
        import time
        
        # Set job timeout to 6 hours
        JOB_TIMEOUT = 21600  # 6 hours in seconds
        self.logger.info(f"Batch Ast: Job Timeout set to {JOB_TIMEOUT} seconds")
        print(f"Batch Ast: Job Timeout set to {JOB_TIMEOUT} seconds")

        processes = []
        manager = mp.Manager()
        return_dict = manager.dict()

        for job_index, job in enumerate(self.jobs):
            self.logger.info(f"Batch Ast: Starting job {job_index}")
            print(f"Batch Ast: Starting job {job_index} Job ({job})")

            # if ast condition is queued or requeued, run the job
            if job.get(self.BATCH_CONDITION_COLUMN) in ['Queued', 'Requeued']: #NOTE Add QUEUED AFTER TESTING***
                
                # Start each job in a separate process
                p = mp.Process(target=process_job_mp, args=(self, job, job_index, self.current_path, return_dict))
                
                # Append the process object to the processes list and job_index. This list keeps track of all the processes and their corresponding job indices.
                processes.append((p, job_index))
                
                # Start method is called on the process object p. This begins the execution of the job in a separate process.
                p.start()
                self.logger.info(f"Batch Ast: {job.get(self.BATCH_CONDITION_COLUMN)} Job {job_index}.....Multiproccessing started......")
                print(f"Batch Ast: Queued Job...Multiproccessing started......")
                

        # Monitor and enforce timeouts
        timeout_failed_counter = 0
        success_counter = 0
        worker_failed_counter = 0
        other_exception_failed_counter = 0
        for process, job_index in processes:
            
            # Join the process to timeout which waits for the process to complete within the timeout
            process.join(JOB_TIMEOUT)
            
            # If the process exceeds the timeout, terminate the process and mark the job as failed
            if process.is_alive():
                
                print(f"Batch Ast: Job {job_index} exceeded timeout. Terminating process.")
                self.logger.warning(f"Batch Ast: Job {job_index} exceeded timeout. Terminating process.")
                
                # End the hung up job
                process.terminate()
                
                # Call the join method again to ensure the process is terminated
                process.join()
                
                # Call add job result and update the job as failed
                self.add_job_result(job_index, 'Failed') 
                
                # Increase the job timeout counter
                timeout_failed_counter+= 1
                self.logger.error(f"Batch Ast: Job {job_index} exceeded timeout. Marking as Failed. Failed counter is {timeout_failed_counter}")
                
            else:
                # Get the result of the job from return_dict. 
                # If the result is 'Success', increment the success_counter and call the add_job_result method to mark the job as 'COMPLETE'
                
                result = return_dict.get(job_index)
                if result == 'Success':
                    success_counter += 1
                    self.add_job_result(job_index, 'COMPLETE')
                    print(f"Batch Ast: Job {job_index} completed successfully.")
                    self.logger.info(f"Batch Ast: Job {job_index} completed successfully. Success counter is {success_counter}")
                
                elif result == 'Failed':
                    
                    # If the result is 'Failed', increment the other_failed_counter and mark the job as 'Failed' (Other failed counter means it failed due to something other than a timeout)
                    # Job failed due to an exception in the worker
                    self.add_job_result(job_index, 'Failed')
                    worker_failed_counter += 1
                    print(f"Batch Ast: Job {job_index} failed due to an exception.")
                    self.logger.error(f"Batch AST: Job {job_index} failed due to an exception in the Worker. Other exception failed counter is {worker_failed_counter}")
                
                else:
                    # Handle unexpected cases
                    self.add_job_result(job_index, 'Unknown Error')
                    other_exception_failed_counter += 1
                    print(f"Batch Ast: Job {job_index} failed with unknown status.")
                    self.logger.error(f"Batch AST: Job {job_index} failed with unknown status. Other Exception failed counter is {other_exception_failed_counter}")
         
        self.logger.info('\n')    
        self.logger.info("Batch Ast Complete - Check separate worker log file for more details")
    


# NOTE ** Reload failed jobs may be able to be incorporated into load failed jobs to tighten up the script
#RELOAD JOBS
    def re_load_failed_jobs_V2(self):
        '''
        re load failed jobs will check for the existence of the queuefile, if it exists it will load the jobs from the queuefile. Checking if they 
        are Failed and if they are, will change Dont Overwrite Outputs to True and add them to the jobs list as Queued
        '''
        self.logger.info("\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Re loading Failed Jobs V2.....")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info("\n")


        # Initialize the jobs list to store jobs
        self.jobs = []

        # Check if the queue file exists
        assert os.path.exists(self.queuefile), "Queue file does not exist"
        if os.path.exists(self.queuefile):

            try:
                # Open the Excel workbook and select the correct sheet
                wb = load_workbook(filename=self.queuefile)
                ws = wb[self.XLSX_SHEET_NAME]
                
                
                print(f'Workbook loaded is {wb}')   
                self.logger.info(f'Re load Failed Jobs: Workbook loaded is {wb}') 
                
                # Get the header (column names) from the first row of the sheet
                header = list([row for row in ws.iter_rows(min_row=1, max_col=None, values_only=True)][0])

                # Read all the data rows (starting from the second row to skip the header)
                data = []
                self.logger.info(f'Re load Failed Jobs: Reading all data rows and building data list')
                for row in ws.iter_rows(min_row=2, max_col=None, values_only=True):
                    print(f'Row is {row}')
                    
                    data.append(row)

                # Iterate over each row of data; enumerate to keep track of the row number in Excel
                self.logger.info(f'Re load Failed Jobs: Iterating over each row of data')
                for job_index, row_data in enumerate(data):  # Start from 2 to account for Excel header
                    
                    self.logger.info(f"\n")
                    self.logger.info(f"------------------------------------------------------------------------------------")
                    self.logger.info(f"-                        Re Load Failed Jobs: Start of Job {job_index}                               -")
                    self.logger.info(f"------------------------------------------------------------------------------------")
                    self.logger.info(f"\n")
                    
                                    
                    # Initialize a dictionary to store the job's parameters
                    job = {}
                    self.logger.info('Re load Jobs - Creating empty dictionary')
                    batch_condition = ''  # Initialize the batch_condition for the current row
                        
                    # Skip any completely blank rows
                    if all((value is None or str(value).strip() == '') for value in row_data):
                        print(f"Re Load Failed Jobs: Skipping blank row at index {job_index}")
                        self.logger.info(f"Re Load Failed Jobs: Skipping blank row at index {job_index}")
                        continue  # Skip this row entirely


                    # Loop through each column header and corresponding value in the current row
                    for key, value in zip(header, row_data):
                        # Check if the key corresponds to the batch_condition column
                        if key is not None and key.lower() == self.BATCH_CONDITION_COLUMN.lower():
                            batch_condition = value if value is not None else ""

                        # Assign an empty string to any None values
                        value = "" if value is None else value

                        # Assign the value to the job dictionary if the key is not None
                        if key is not None:
                            # logger.info(f"Re Load Failed Jobs: Assigning values to job dictionary")
                            print(f"Re Load Failed Jobs: Assigning values to job dictionary")
                            job[key] = value

                    # Skip if marked as "COMPLETE"
                    if batch_condition.upper() == 'COMPLETE':
                        print(f"Re Load Failed Jobs: Skipping job {job_index} as it is marked {batch_condition}.")
                        self.logger.info(f"Re Load Failed Jobs: Adding Complete to dictionary Skipping job {job_index} as it is marked COMPLETE.")
                        # continue  
                        batch_condition = 'COMPLETE'    
                    
                    # Change batch condition to requeued if the job is failed
                    elif batch_condition.upper() == 'FAILED':
                        self.logger.info(f"Re Load Failed Jobs: Requeuing {job_index} as it is marked Failed.")
                        batch_condition = 'Requeued'
                    
                    else:
                        self.logger.warning(f"Re Load Failed Jobs: Job {job_index} is not marked as Complete or Failed. Please check the workbook. Skipping this job.")
                        # continue
                        batch_condition = 'ERROR'
                    
                    # Assign updated condition to the job dictionary
                    job[self.BATCH_CONDITION_COLUMN] = batch_condition
                        
                    print(f"Re Load Failed Jobs: Job {job_index} is marked as Failed, re-assigning batch condition to {batch_condition}")
                    self.logger.info(f"Re Load Failed Jobs: Job {job_index}'s batch condition has been updated as '{batch_condition}'")


                    # Immediately update the Excel sheet with the new condition
                    try:
                        self.add_job_result(job_index, batch_condition)
                        self.logger.info(f"Re load Jobs - Added job condition '{batch_condition}' for job {job_index} to jobs list")
                    # Added Dec 12th
                        # Save the workbook with the updated condition
                        self.logger.info(f"Reload Failed Jobs - Saving Job Index ({job_index}) with new condition.")
                        wb.save(self.queuefile)
                    except Exception as e:
                        print(f"Error updating Excel sheet at row {job_index}: {e}")
                        self.logger.error(f"Re load Jobs - Error updating Excel sheet at row {job_index}: {e}")
                        self.logger.error(traceback.format_exc())
                        continue
                        
                        
                    # # Check the condition of DONT_OVERWRITE_OUTPUTS
                    # current_value = job.get(self.DONT_OVERWRITE_OUTPUTS, '')

                    # # If DONT_OVERWRITE_OUTPUTS is anything but True, change it to 'True'
                    # if current_value != 'True':
                    #     print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is '{current_value}', changing to True")
                    #     self.logger.warning(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is '{current_value}', changing to True")
                    #     job[self.DONT_OVERWRITE_OUTPUTS] = "True"
                        
                    #     # Added Dec 12th
                    #     # Save the workbook with the updated condition
                    #     self.logger.info(f"Reload Failed Jobs - Saving Job Index ({job_index}) with new condition.")
                    #     wb.save(self.queuefile)

                        
                        
                    #     # # Log the current state before changing
                    #     # if current_value == 'False':
                    #     #     print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is False, changing to True")
                    #     #     self.logger.info(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is False, changing to True")
                    #     # elif current_value == '':
                    #     #     print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is empty, changing to True")
                    #     #     self.logger.error(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is empty, changing to True")
                    #     # else:
                    #     #     print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is '{current_value}', changing to True")
                    #     #     self.logger.warning(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is '{current_value}', changing to True")
                        
                    #     # # Set the value to 'True'
                    #     # job[self.DONT_OVERWRITE_OUTPUTS] = "True"
                    
                    # # If DONT_OVERWRITE_OUTPUTS is already 'True, don't change it. 
                    # else:
                    #     # If it's already 'True', log that no change is needed
                    #     print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is already True, no change needed")
                    #     self.logger.info(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is already True, no change needed")

                    # Add the job to the jobs list after all checks and processing
                    self.jobs.append(job)
                    print(f"Re load Jobs - Job Condition is not Complete ({batch_condition}), adding job: {job_index} to jobs list")
                    self.logger.info(f"Re load Jobs - Job Condition is not Complete ({batch_condition}), adding job: {job_index} to jobs list")
                    self.logger.info(f"\n")
                    self.logger.info(f"------------------------------------------------------------------------------------")   
                    self.logger.info(f" Job list is {job}")
                    self.logger.info(f"------------------------------------------------------------------------------------")
                    self.logger.info(f"\n")
                    
                    
                    
                    print(f"Re Load Jobs - Job dictionary is {job}")
                    self.logger.info(f"Re load Jobs - Job {job_index} dictionary is {job}")
                            
            except FileNotFoundError as e:
                print(f"Error: Queue file not found - {e}")
                self.logger.error(f"Re Load Failed Jobs Error: Queue file not found - {e}")
                self.logger.error(traceback.format_exc())
            except Exception as e:
                print(f"Unexpected error re loading jobs: {e}")
                self.logger.error(f"Re Load Failed Jobs Unexpected error loading jobs: {e}")
                self.logger.error(traceback.format_exc())

        return self.jobs   

    def create_new_queuefile(self):
        '''write a new queuefile with preset header'''

        
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Creating New Queuefile...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")

        
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.XLSX_SHEET_NAME
        headers = list(self.BATCH_PARAMETERS
.values())
        headers.append(self.BATCH_CONDITION_COLUMN)
        for h in headers:
            c = headers.index(h) + 1
            ws.cell(row=1, column=c).value = h
        wb.save(self.queuefile)

    
        

    def capture_arcpy_messages(self):
        ''' Re assigns the arcpy messages  (0 for all messages, 1 for warnings, and 2 for errors) to variables and passes them to the logger'''
        
        arcpy_messages = arcpy.GetMessages(0) # Gets all messages
        arcpy_warnings = arcpy.GetMessages(1) # Gets all warnings only
        arcpy_errors = arcpy.GetMessages(2) # Gets all errors only
        
        if arcpy_messages:
            self.logger.info(f'batch_toobox arcpy messages: {arcpy_messages}')
        if arcpy_warnings:
            self.logger.warning(f'batch_toobox arcpy warnings: {arcpy_warnings}')
        if arcpy_errors:
            self.logger.error(f'batch_toobox arcpy errors: {arcpy_errors}')   