'''
    Author:        Mark McGirr
    Editor:        Wes Smith/Steve Richards

    Purpose:       This script is intended to be run from a tool, and to have
                   arguments passed into it from the tool.  This is the "CALL ROUTINE" for
                   the automated status tool.  It takes all the input arguments from the tool
                   interface and builds a list out of them.  It then invokes the
                   Universal_Overlap_Tool (revolt), if it is selected,  and passes in that 
                   list.  The universal overlap tool will create a spreadsheet of the 
                   overlapping conflicts.  This script will invoke the automated_status_tool, 
                   which will generate a spreadsheet with tabs 1 and 2 of the final output.  
                   Finally, this script merges the spreadsheets from the universal overlap 
                   and automated status tools, if the workbooks exist: merging them into 
                   automated_status_sheet spreadsheet.  It will include the hyperlinks to 
                   the maps, if they exist. 
                   
                   If you run it directly from eclipse it has a section 
                   that creates all the variables that would normally have been passed in 
                   from the tool.  This makes testing easier and faster.

    Date:           Original:     July 5, 2012
                    Edited:       May 2022

    Arguments:      region = sys.argv[1]
                    feature_layer = sys.argv[2]
                    dont_overwrite_outputs = sys.argv[3]
                    skip_conflicts_and_constraints = sys.argv[4]
                    suppress_map_creation = sys.argv[5]
                    debug_version = sys.argv[6]
                    dont_delete_mxds = sys.argv[7]
                    bcgw_username = sys.argv[8]
                    bcgw_password = sys.argv[9]
                    
    Logic Overview:
                    This script accepts a feature layer from an APRX's table of 
                    contents.  The analysis will be run on the selected features 
                    within the features layer, if there are features selected, 
                    otherwise the tool will run the analysis on the whole dataset.
                    
                    The script creates a working GDB file in the directory_to_store_output
                    that is one of the arguments.  It extracts the shape into that
                    GDB.

                    It then runs the Universal_Overlap_Tool
                    (revolt_universal_overlap_tool.py), if specified by the user.  
                    This creates a spreadsheet named one_status_common_datasets_aoi.xlsx

                    It then runs the automated_status_sheet_part2.py which creates
                    a spreadsheet one_status.xlsx.

                    It then merges both of the above spreadsheets into a final
                    workbook named automated_status_sheet.xlsx

                    If the dont_overwrite_outputs flag is set then all the
                    current data clips and maps will be preserved.  This saves time when
                    it crashes for some reason, or when debugging / testing the script.

    History:
    --------------------------------------------------------------
    Date:          2012
    Author:        Mark McGirr
    Modification:  Original ArcGIS 10.X
    
    Date:          March-April 2020
    Author:        Wes Smith
    Modification:  - Convert script's syntax from Python 2.7 to Python 3.
                   - Updated/condensed some of the old code.  The script now uses the openpyxl
                   library for working with XLSX files (instead of win32.com)
                   - The script receives a user's BCGW credentials and validates them before
                   progressing through the script.
                   - The script also reports back on if a selection exists in the incoming
                   feature layer.
    
    --------------------------------------------------------------
'''
#___________________________________________________________________________
## IMPORTS ##
'''
Imports the needed libraries
'''
import sys, os, openpyxl, arcpy, runpy, shutil, subprocess
from openpyxl.styles import Alignment, Font, PatternFill #,Border
from openpyxl.styles.borders import Border, Side

# import both the statusing tools which create tabs 1, 2, 3
sys.path.append(r'\\GISWHSE.ENV.GOV.BC.CA\WHSE_NP\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\beta')
#sys.path.append(r'\\GISWHSE.ENV.GOV.BC.CA\WHSE_NP\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\Scripts')
import universal_overlap_tool_arcpro as revolt #@UnresolvedImport
import one_status_tabs_one_and_two_arcpro as one_status_part2
import create_bcgw_sde_connection as connect_bcgw
import config

#___________________________________________________________________________

## Process ##
#Check to ensure Advanced licencing has been applied.
arcpy.AddMessage("======================================================================")
arcpy.AddMessage("Checking for ArcGIS Pro Advanced license")

#Check to ensure Advanced licencing has been applied.
advStatus = ["Available", "AlreadyInitialized"]
if arcpy.CheckProduct("ArcInfo") not in advStatus:
    msg = 'ArcGIS Pro Advanced license not available. Set license to "Advanced" and try again'
    arcpy.AddError(msg)
    sys.exit()

arcpy.AddMessage("======================================================================")


def main():
    '''
    Function that prepares variables and data, checks for errors in the data, and
    runs the universal overlap and automated status tools
    '''
    message = "Now Running " + str(sys.argv[0])
    arcpy.AddMessage(message)


    # Read arguments passed by tool
    '''
    These arguments are passed in from the tool
    If you are running this tool from Eclipse there is
    a section below that realizes there are no arguments being
    passed in, and it will set up testing / default arguments.
    '''
    
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("\nReading Arguments")
    
    # Create empty variables (filled below)
    region = ""                         #@UnusedVariable
    disposition_number = ""             #@UnusedVariable
    crown_file_number = ""              #@UnusedVariable
    parcel_number = ""                  #@UnusedVariable
    debug_version = ""                  #@UnusedVariable
    suppress_map_creation = ""          #@UnusedVariable
    output_dir_same_as_input = ""       #@UnusedVariable
    output_directory = ""               #@UnusedVariable
    add_maps_to_current = ""

    
    # Update variables with user inputs, if possible
    region = arcpy.GetParameterAsText(0).lower()
    feature_layer = arcpy.GetParameter(1)
    # ___________
    crown_file_number = arcpy.GetParameterAsText(2)
    disposition_number = arcpy.GetParameterAsText(3)
    parcel_number = arcpy.GetParameterAsText(4)
    # ___________
    output_directory = arcpy.GetParameterAsText(5)
    output_dir_same_as_input = arcpy.GetParameterAsText(6)
    dont_overwrite_outputs = arcpy.GetParameterAsText(7)
    skip_conflicts_and_constraints = arcpy.GetParameterAsText(8)
    suppress_map_creation = arcpy.GetParameterAsText(9)
    add_maps_to_current = arcpy.GetParameterAsText(10)
    run_as_fcbc = arcpy.GetParameterAsText(11)
    # ___________
    debug_version = arcpy.GetParameterAsText(12)


    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("Checking BCGW Credentials - may take a minute to process...")

    #set the key name that will be used for storing credentials in keyring
    key_name = config.CONNNAME
    try:
        oracleCreds = connect_bcgw.ManageCredentials(key_name, output_directory)
        #get sde path location
        if not oracleCreds.check_credentials():
            arcpy.AddError("BCGW credentials could not be established.")
            sys.exit()
        sde = os.getenv("SDE_FILE_PATH")

    except Exception as e:
        arcpy.AddError(f"Failure occurred when establishing BCGW connection - {e}. Please try again.")
        sys.exit()

    #Check RAAD connection
    raad = os.path.join(sde, "WHSE_ARCHAEOLOGY.RAAD_TFM_SITE")
    try:
        arcpy.MakeFeatureLayer_management(raad, "RAAD_lyr")
    except arcpy.ExecuteError as e:
        arcpy.AddWarning(f"Unable to connect to RAAD data")

    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("Running Validation")

    # Check for selection in feature layer
    # number of features passed through feature layer (selection, if one exisits)
    if arcpy.Exists(feature_layer):
        # Get feature layer path to determine the total number of features
        count1 = int(str(arcpy.GetCount_management(feature_layer)))
        desc = arcpy.Describe(feature_layer)
        feat_lyr2 = arcpy.MakeFeatureLayer_management(desc.catalogPath, "temp_aoi_fl")
        # total number of feature in the dataset
        count2 = int(str(arcpy.GetCount_management(feat_lyr2)))
        # Compare the counts and make custom message
        if count2 == count1:
            message = "The analysis will be run on all {} feature(s), since no features were selected.".format(count2)
        elif count1 < count2:
            message = "There was a selection.  The analysis will be run on {} of {} feature(s).".format(count1, count2)
        else:
            message = "There was a ERROR in the validation section -- counting features"
        # Report on selection
        del count1, count2
    else: 
        message = "    No feature layer was provided.  Checking for Crown File, Disposition, and Parcel Numbers..."
    arcpy.AddMessage(message)

    # Report back on the set variables
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("PASSED ARGUMENTS:")
    arcpy.AddMessage("{:35} {:255}".format("region", region))
    ##arcpy.AddMessage("{:35} {:255}".format("feature_layer", feature_layer))
    arcpy.AddMessage("{:35} {:255}".format("output_dir_same_as_input", output_dir_same_as_input))
    arcpy.AddMessage("{:35} {:255}".format("output_directory", output_directory))
    # ___________
    arcpy.AddMessage("{:35} {:255}".format("crown_file_number", crown_file_number))
    arcpy.AddMessage("{:35} {:255}".format("disposition_number", disposition_number))
    arcpy.AddMessage("{:35} {:255}".format("parcel_number", parcel_number))
    # ___________
    arcpy.AddMessage("{:35} {:255}".format("dont_overwrite_outputs", dont_overwrite_outputs))
    arcpy.AddMessage("{:35} {:255}".format("skip_conflicts_and_constraints", skip_conflicts_and_constraints))
    arcpy.AddMessage("{:35} {:255}".format("debug_version", debug_version))
    arcpy.AddMessage("{:35} {:255}".format("suppress_map_creation", suppress_map_creation))
    
    
    #___________
    # Hard-coded and derived variables
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("Setting hard-coded variables")
    spatial_reference = arcpy.SpatialReference(3005)
    if region == 'Cariboo Debug':
        xls_file_for_analysis_input = r"\\giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\statusing_input_spreadsheets\one_status_common_datasets_debug_version.xlsx"
        xls_file_for_analysis_input2 = r"\\giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\statusing_input_spreadsheets\one_status_cariboo_specific_debug_version.xlsx"
    else:
        xls_file_for_analysis_input = r"\\giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\statusing_input_spreadsheets\one_status_common_datasets.xlsx"
        xls_file_for_analysis_input2 = r"\\giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\statusing_input_spreadsheets\one_status_" + region + "_specific.xlsx"
    
 
    '''
    Sets up the geodatabase name, and deletes that
    geodatabase if the dont_overwrite_outputs != True
    '''
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("   Checking existence of GDB")
    #specify output folder and gdb to store the input spatial depending
    #on parameters set by user in tool parameters.
    gdb_name = "aoi_boundary.gdb"
    try:
        if output_directory != "#" and output_directory != "":
            directory_to_store_output = output_directory
            data_gdb = os.path.join(directory_to_store_output, gdb_name)
            arcpy.AddMessage("Full_Path: " + data_gdb)
        else:
            desc = arcpy.Describe(feature_layer)
            path = desc.catalogPath
            arcpy.AddMessage("Feature Layer Input: " + path)
            directory_to_store_output = revolt.get_fc_directory_name(path)
            data_gdb = os.path.join(directory_to_store_output, gdb_name)
            arcpy.AddMessage("Full_Path: " + data_gdb)
    except Exception as e:
        arcpy.AddWarning(e)
        sys.exit()
    
    if arcpy.Exists(data_gdb):
        arcpy.AddMessage("found: " + data_gdb)
        if dont_overwrite_outputs == "false" or dont_overwrite_outputs == "#": #delete gdb if exists and overwrite <> true
            arcpy.AddMessage("deleting " + data_gdb)
            arcpy.Delete_management(data_gdb)
            arcpy.CreateFileGDB_management(directory_to_store_output, gdb_name)
        elif dont_overwrite_outputs == "true":
            arcpy.AddWarning("GDB exists and dont_overwrite_outputs is set to true")
    else:
        arcpy.AddMessage("creating " + data_gdb)
        arcpy.CreateFileGDB_management(directory_to_store_output, gdb_name)


    '''
    Gets the aoi_boundary, from the feature layer.
    this boundary is put in the aoi_boundary.gdb as AOI.  This
    AOI will be used by the revolt overlap tool as it's area of interest,
    and by "part2" of the tool that creates the first 2 tabs of the final xlsx.
    '''
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("Checking AOI")

    the_clean_output = os.path.join(data_gdb , "aoi_clean")
    if not arcpy.Exists(the_clean_output):
        # create feature dataset
        feature_dataset_name = "input_of_raw_data"
        arcpy.CreateFeatureDataset_management(data_gdb, feature_dataset_name, spatial_reference)

        # copy aoi feature(s) to the raw feature dataset
        the_output = os.path.join(data_gdb, feature_dataset_name, "aoi_boundary_raw")
        # check if output exists         
        if not arcpy.Exists(the_output):
            # Get data from user supplied feature layer
            if arcpy.Exists(feature_layer):
                arcpy.AddMessage("   getting AOI boundary from feature layer -- copying to GDB")
            # get data frin BCGW with user supplied crown_file_number
            elif crown_file_number:
                arcpy.AddMessage("   getting parcel boundary from tantalis")
                arcpy.MakeFeatureLayer_management(os.path.join(sde, "WHSE_TANTALIS.TA_CROWN_TENURES_SVW"),
                                                "input_layer")
                selection_string = "\"CROWN_LANDS_FILE\" = " + "'" + crown_file_number + "'"
                selection_string = selection_string + " and "
                selection_string = selection_string + "\"DISPOSITION_TRANSACTION_SID\" = " + "'" + disposition_number + "'"
                
                if parcel_number != "" and  parcel_number != "#":
                    selection_string = selection_string + " and "
                    selection_string = selection_string + "\"INTRID_SID\" = " + parcel_number
                
                arcpy.AddMessage(f'Query: {selection_string}')

                # Select features
                feature_layer, count = arcpy.SelectLayerByAttribute_management("input_layer", "NEW_SELECTION", selection_string)
                input_layer_count = int(str(arcpy.GetCount_management("input_layer")))
                
                arcpy.AddMessage(f'* count: {count}, input_layer_count: {input_layer_count}')

                # Check
                if input_layer_count == 0:
                    arcpy.AddError("The crown file / disposition pair was not found in the BCGW Tantalis data. ")
                    sys.exit()
            
            # Copy to input_of_raw_data dataset
            arcpy.CopyFeatures_management(feature_layer, the_output)
            # Copy to root of GDB
            arcpy.CopyFeatures_management(the_output, the_clean_output)
            arcpy.RepairGeometry_management(the_clean_output)



    #___________
    # AOI Validation
    '''
    Checks the inputs AOI data for common issues
    '''
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("Checking for possible errors in AOI shape")

    # Variables
    message_dict = {}                           # holds possible error name and message
    possible_problems_with_input_shape = 'no'   # FLAG - No errors, so far

    # Catch MULTIPART
    multipart = 'no'
    with arcpy.da.SearchCursor(the_clean_output, ["SHAPE@"]) as cursor:  #@UndefinedVariable
        for row in cursor:
            geometry = row[0]
            if geometry.isMultipart == True:
                multipart = 'yes'
                possible_problems_with_input_shape = 'yes'
    # Report
    message_string = "is this shape multipart ?  " + "   " +   multipart
    message_dict["multipart"] = [multipart, message_string]

    # Catch > 1 Features
    multiple_polygons = 'no'
    result = arcpy.GetCount_management(the_clean_output)
    count = int(result.getOutput(0))
    if count > 1:
        multiple_polygons = 'yes'
        possible_problems_with_input_shape = 'yes'
    # Report
    message_string = "are there multiple polygons ?  " + "   " +   multiple_polygons
    message_dict["multiple_polygons"] = [multiple_polygons, message_string]

    # Catch too many Vertices
    lots_of_vertices = 'no'
    features = [feature[0] for feature in arcpy.da.SearchCursor(the_clean_output,"SHAPE@")] #@UndefinedVariable
    count_vertices = sum([f.pointCount-f.partCount for f in features])
    #print "count_vertices " , count_vertices
    if count_vertices > 5000:
        lots_of_vertices = 'yes'
        possible_problems_with_input_shape = 'yes'
    # Report
    message_string = "are there lots of vertices ?  " + "   " +   lots_of_vertices
    message_dict["lots_of_vertices"] = [lots_of_vertices, message_string]
    
    message_string = "are there possible problems with you input FC ?  " + "   " +   possible_problems_with_input_shape
    message_dict["possible_problems_with_input_shape"] = [possible_problems_with_input_shape, message_string]
    
    for k, v in message_dict.items(): #@UnusedVariable
        flag, message = v
        if flag == "yes":
            pass
            #arcpy.AddWarning(message)
        else:
            pass
            #arcpy.AddMessage(message)
        #print(k)
        arcpy.AddMessage("   {}".format(message))
        
    the_aoi = os.path.join(data_gdb , "aoi") #create additional feature class as input to the UOT without the fields.


    #___________
    # Prepare Variables to pass into tool
    '''
    Sets up all the Variables that need to be passed into the
    Universal Overlap Tool.  Most of them will be blank because
    they are only accessed from the Generic Tool Interface.
    If the debug_version flag is set to True, than the input
    spreadsheets are set to small version so the testing runs
    quickly.

    Also sets up the variables to be passed into the the second
    part of the tool that creates tab1 and tab2 of the final spreadsheet
    '''
    
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("Preparing to pass variables")
    revolt_criteria_to_pass =  []
    revolt_criteria_to_pass.append(the_aoi)            # 0 TEXT - analyize_this_featureclass
    revolt_criteria_to_pass.append(r"")                         # 1 TEXT - create_subreports_on_this_field
    revolt_criteria_to_pass.append(r"")                         # 2 TEXT - what_type_of_overlap_to_run
    revolt_criteria_to_pass.append(xls_file_for_analysis_input) # 3 TEXT - xls_file_for_analysis_input
    revolt_criteria_to_pass.append(r"")                         # 4 TEXT - text_header1
    revolt_criteria_to_pass.append(r"")                         # 5 TEXT - text_header2
    revolt_criteria_to_pass.append(r"")                         # 6 TEXT - text_header3
    revolt_criteria_to_pass.append(r"")                         # 7 TEXT - text_header4
    revolt_criteria_to_pass.append(r"")                         # 8 TEXT - True - subreports_on_seperate_sheets
    revolt_criteria_to_pass.append(directory_to_store_output)   # 9 TEXT - directory_to_store_output
    revolt_criteria_to_pass.append(r"")                         # 10 TEXT - True - summary_fields_on_seperate_lines
    revolt_criteria_to_pass.append(dont_overwrite_outputs)      # 11 TEXT - True - test_dont_overwrite_data_and_maps
    revolt_criteria_to_pass.append(xls_file_for_analysis_input2)# 12 TEXT - xls_file_for_analysis_input2 reads a second input spreadsheet into the list
    revolt_criteria_to_pass.append("INTERNAL PROVINCIAL GOVERNMENT USE ONLY")   # 13 TEXT - xls_file_for_analysis_input2 reads a second input spreadsheet into the list
    revolt_criteria_to_pass.append(suppress_map_creation)       # 14 TEXT - Dont try to create the maps on tab 3
    revolt_criteria_to_pass.append(region)            #19 TEXT - Natural Resource Region
    revolt_criteria_to_pass.append(crown_file_number) #20 TEXT - Crown Lands File
    revolt_criteria_to_pass.append(disposition_number)#21 TEXT - Disposition Number
    revolt_criteria_to_pass.append(parcel_number)     #22 TEXT - Parcel Number
    revolt_criteria_to_pass.append(run_as_fcbc)       #23 TEXT - run spreadsheets in FCBC format
    revolt_criteria_to_pass.append(add_maps_to_current)


    one_status_part2_criteria_to_pass = []
    one_status_part2_criteria_to_pass.append(directory_to_store_output)
    one_status_part2_criteria_to_pass.append(dont_overwrite_outputs) # don't overwrite data if it already exists
    one_status_part2_criteria_to_pass.append(region)            # TEXT - Natural Resource Region
    one_status_part2_criteria_to_pass.append(crown_file_number) # TYPE - Crown Lands File
    one_status_part2_criteria_to_pass.append(disposition_number)# TYPE - Disposition Number
    one_status_part2_criteria_to_pass.append(parcel_number)     # TYPE - Parcel Number
    #one_status_part2_criteria_to_pass.append(r"")
    #one_status_part2_criteria_to_pass.append(r"")
    #one_status_part2_criteria_to_pass.append(r"")
    one_status_part2_criteria_to_pass.append(the_aoi)


    #___________
    # RUN TOOLS

    arcpy.AddMessage("Passing variables")
    # PART 1 - Run Conflicts and Constraints
    '''
    Runs the universal_overlap_tool on PARCEL_BOUNDARY.GDB\AOI
    if the flag to skip it has not been set to True
    '''
    if skip_conflicts_and_constraints == "false":
        arcpy.AddMessage("")
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("{0} Running Universal Overlap Tool {0}".format("*"*3))
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("")
        # Remove all non-required fields from the AOI feature class. When the AOI contains
        # fields that are common with the input constraints, it causes an issue with the output
        # XLS file where the reporting of the data in the common field does not get written.
        # (added May 3, 2021)
        arcpy.CopyFeatures_management(the_clean_output, the_aoi)
        arcpy.AddMessage("Deleting fields from AOI to avoid conflicts...")
        for field in arcpy.ListFields(the_aoi):
            if not field.required:
                try:
                    print(f"deleting {field.name} from AOI dataset")
                    arcpy.DeleteField_management(the_aoi, field.name)
                except:
                    arcpy.AddMessage(arcpy.GetMessages())
        revolt_obj = revolt.revolt_tool()
        revolt_obj.run_revolt_tool(revolt_criteria_to_pass)

    # PART 2 - Automated Status
    '''
    Runs Part2 of the Automated Status Tool.
    This creates the one_status_tabs_1_and_2.xlsx
    '''
    arcpy.AddMessage("")
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("{0} Running Automated Status Tool {0}".format("*"*3))
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage("")
    arcpy.Delete_management(the_aoi)
    arcpy.CopyFeatures_management(the_clean_output, the_aoi)
    arcpy.Delete_management(the_clean_output)
    onestatus_obj = one_status_part2.one_status_part2_tool()
    onestatus_obj.run_tool(one_status_part2_criteria_to_pass)

    #___________
    # Merge Sheets into Final Workbook
    '''
    Merges the 2 spreadsheets into the final Automated_Status_Sheet.xlsx
    Both input spreadsheets must exist for the final one to be created, else
    the user will just have to use the individual spreadsheets.
    '''
    arcpy.AddMessage("Starting the merge of two XLS files")
    # XLSX Variables
    part1_xlsx = os.path.join(directory_to_store_output,"one_status_common_datasets_aoi.xlsx")
    part2_xlsx = os.path.join(directory_to_store_output,"one_status_tabs_1_and_2.xlsx")
    part3_xlsx = os.path.join(directory_to_store_output,"automated_status_sheet.xlsx")
    # Catch Debug Flag
    if debug_version == 'true'  :
        part1_xlsx = os.path.join(directory_to_store_output,"one_status_common_datasets_debug_version_aoi.xlsx")

    # Prepare files (some as VARS and some deleted, if conditions are met)
    if os.path.isfile(part1_xlsx) and os.path.isfile(part2_xlsx):
        if os.path.isfile(part3_xlsx):
            arcpy.AddMessage(".  Overwriting the existing working XLSX:\n   {}".format(part3_xlsx))
            arcpy.Delete_management(part3_xlsx)
        # Make copy 
        shutil.copy(part1_xlsx, part3_xlsx)
        #shutil.copy(part2_xlsx, part3_xlsx)
        # workbook objects
        arcpy.AddMessage(".  loading the workbooks")
        dst_wb = openpyxl.load_workbook(part3_xlsx)
        src_wb = openpyxl.load_workbook(part2_xlsx)
        # names of sheets of interest
        src_ws_list = ["Status of Conflict", "Crown Land Status"]
        # worksheet names from src workbook
        src_wss = sorted(src_wb.sheetnames, reverse=True)
        # loop through sheet names (from src)
        for ws in src_wss:
            #print(ws)
            #loop through sheet names of interest 
            for name in sorted(src_ws_list, reverse=True):    
                # find match
                if ws == name:
                    # worksheet object
                    worksheet = src_wb[name]
                    # call the copy function
                    copySheet_toNewWB(worksheet, name, dst_wb, part3_xlsx)
        
    elif os.path.isfile(part2_xlsx):
        print("Only 1 file found:\n   {}".format(part2_xlsx))
        # Make copy
        shutil.copy(part2_xlsx, part3_xlsx)
    
    if arcpy.Exists(part3_xlsx):
        arcpy.AddMessage(".")
        arcpy.AddMessage(".")
        arcpy.AddMessage("Automated_status_sheet.xlsx is ready for you to use")
        arcpy.AddMessage("{}".format(part3_xlsx))
        arcpy.AddMessage(".")
        arcpy.AddMessage(".")

    #cleanup temporary sde file
    try:
        shutil.rmtree(os.path.dirname(os.path.abspath(os.getenv("SDE_FILE_PATH"))))
        del os.environ["SDE_FILE_PATH"]
    except Exception as e:
        pass
    

#___________________________________________________________________________

## FUNCTIONS ##

def apply_border2(ws, start_row, end_row, start_column, end_column):
        '''
        Applies a thick black border to a range of cells
        '''
        
        cell_range = start_column + str(start_row+1) + ":" + end_column + str(end_row)
        rows = ws[cell_range]
        for row in rows:
            if row == rows[0][0] or row == rows[0][-1] or row == rows[-1][0] or row == rows[-1][-1]:
                pass
            else:
                row[0].border = Border(left=Side(style='thick'))
                row[-1].border = Border(right=Side(style='thick'))
            for c in rows[0]:
                c.border = Border(top=Side(style='thick'))
            for c in rows[-1]:
                c.border = Border(bottom=Side(style='thick'))
        rows[0][0].border = Border(left=Side(style='thick'), top=Side(style='thick'))
        rows[0][-1].border = Border(right=Side(style='thick'), top=Side(style='thick'))
        rows[-1][0].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
        rows[-1][-1].border = Border(right=Side(style='thick'), bottom=Side(style='thick'))

def copySheet_toNewWB(src_ws, ws_name, dst_wb, dst_wb_path):
    '''
    Copies cell values and styles from a Source Worksheet to a Destination Worksheet.
    The script iterates through row and columns, writing cells one-by-one.  
    '''
    # range of cells to be copied    
    max_row = src_ws.max_row
    max_col = src_ws.max_column
    # Column widths from src
    col_width_dict = {}
    for col in range(1, max_col+1):
        ltr = openpyxl.utils.cell.get_column_letter(col) #@UndefinedVariable
        col_width_dict[col] = [ltr, src_ws.column_dimensions[ltr].width]
    # Get dst WS ready
    dst_ws_name = ws_name
    # New Workbook and Worksheet
    dst_wb.create_sheet(dst_ws_name, 0)
    dst_ws = dst_wb[dst_ws_name]
    # Update column width in dst WS
    for k, v in col_width_dict.items():
        dst_ws.column_dimensions[v[0]].width = v[1]
    del col_width_dict, k, v
    # Loop over rows
    for row in range(1, max_row+1):
        # set dst row height f rom src
        src_row_height = src_ws.row_dimensions[row].height
        if src_row_height != None:
            dst_ws.row_dimensions[row].height = src_row_height
        # loop over columns
        for col in range(1, max_col+1):
            # Copy src cell contents to dst
            src_cell = src_ws.cell(row=row, column=col)
            dst_cell = dst_ws.cell(row=row, column=col)
            # Cells not merged
            if type(dst_cell).__name__ != 'MergedCell':
                # value
                dst_cell.value = src_cell.value                
                # get source style
                src_font = src_cell.font
                src_fill = src_cell.fill
                src_alignment = src_cell.alignment
                src_border = src_cell.border
                # apply style
                if src_cell.has_style:  
                    # Applies only styles named below (Font, PatternFill, Alignment, and Border)
                    dst_cell.font = Font(name=src_font.name,
                                         size=src_font.size,
                                         color=src_font.color)
                    dst_cell.fill = PatternFill(fill_type=src_fill.fill_type,
                                         fgColor=src_fill.fgColor)
                    dst_cell.alignment = Alignment(horizontal=src_alignment.horizontal,
                                         wrap_text=src_alignment.wrap_text)
                    dst_cell.border = Border(left=src_border.left,
                                         right=src_border.right,
                                         top=src_border.top,
                                         bottom=src_border.bottom)
                
                # Catch cells that need merging
                # Variables
                col_ltr = openpyxl.utils.cell.get_column_letter(col) #@UndefinedVariable
                col_ltr_2 = openpyxl.utils.cell.get_column_letter(col+1) #@UndefinedVariable
                # Do the merging
                if dst_cell.value is not None and \
                        (dst_cell.value in ["Additional Comments", "Status Summary"] or \
                         "Purpose: " in dst_cell.value[:9]):
                    if "Purpose: " in dst_cell.value[:9] or "Status Summary" in dst_cell.value:
                        merge_range = col_ltr + str(row) + ":" + col_ltr_2 + str(row)
                    elif dst_cell.value == "Additional Comments":
                        row = row + 1
                        merge_range = col_ltr + str(row) + ":" + col_ltr_2 + str(row)
                        
                    dst_ws.merge_cells(merge_range)
                    apply_border2(dst_ws, row-1, row+1, "A", "B")
                    dst_cell.alignment = Alignment(horizontal=src_alignment.horizontal,
                                             wrap_text=src_alignment.wrap_text)
            #print("{}:{} --> {}".format(col_ltr, row, dst_cell.value))
    dst_wb.save(dst_wb_path)      

#___________________________________________________________________________

main()
