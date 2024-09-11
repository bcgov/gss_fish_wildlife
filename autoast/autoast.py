# autoast is a script for batch processing the automated status tool
# author: wburt
# copyrite Governent of British Columbia
# Copyright 2019 Province of British Columbia

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at 

#    http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import shutil
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import geopandas
import arcpy
import datetime
import logging


## *** INPUT YOUR EXCEL FILE NAME HERE ***
excel_file = '2_wmus.xlsx'

# Define the job timeout in seconds (6 hours)
JOB_TIMEOUT = 6 * 60 * 60

###############################################################################################################################################################################

###############################################################################################################################################################################




def import_ast():
    # Get the toolbox path from environment variables
    ast_toolbox = os.getenv('TOOLBOX') # File path 

    if ast_toolbox is None:
        print("Unable to find the toolbox. Check the path in .env file")
        self.logger.error("Unable to find the toolbox. Check the path in .env file")
        exit() 

    # Import the toolbox
    try:
        arcpy.ImportToolbox(ast_toolbox)
        print(f"AST Toolbox imported successfully.")
        self.logger.info(f"AST Toolbox imported successfully.")
    except Exception as e:
        print(f"Error importing toolbox: {e}")
        self.logger.error(f"Error importing toolbox: {e}")
        exit()

    # Assign the shapefile template for FW Setup to a variable
    template = os.getenv('TEMPLATE') # File path in .env
    if template is None:
        print("Unable to find the template. Check the path in .env file")
        self.logger.error("Unable to find the template. Check the path in .env file")
        
    return template
    

###############################################################################################################################################################################
#
# Set up the database connection
#
###############################################################################################################################################################################
def setup_bcgw():
    # Get the secret file containing the database credentials
    SECRET_FILE = os.getenv('SECRET_FILE')

    # If secret file found, load the secret file and display a print message, if not found display an error message
    if SECRET_FILE:
        load_dotenv(SECRET_FILE)
        print(f"Secret file {SECRET_FILE} found")
        self.logger.info(f"Secret file {SECRET_FILE} found")
    else:
        print("Secret file not found")
        self.logger.error("Secret file not found")

    # Assign secret file data to variables    
    DB_USER = os.getenv('BCGW_USER')
    DB_PASS = os.getenv('BCGW_PASS')

    # If DB_USER and DB_PASS found display a print message, if not found display an error message
    if DB_USER and DB_PASS:
        print(f"Database user {DB_USER} and password found")
        self.logger.info(f"Database user {DB_USER} and password found")
    else:
        print("Database user and password not found")
        self.logger.error("Database user and password not found")

    # Define current path of the executing script
    current_path = os.path.dirname(os.path.realpath(__file__))

    # Create the connection folder
    connection_folder = 'connection'
    connection_folder = os.path.join(current_path, connection_folder)

    # Check for the existance of the connection folder and if it doesn't exist, print an error and create a new connection folder
    if not os.path.exists(connection_folder):
        print("Connection folder not found, creating new connection folder")
        self.logger.info("Connection folder not found, creating new connection folder")
        os.mkdir(connection_folder)

    # Check for an existing bcgw connection, if there is one, remove it
    if os.path.exists(os.path.join(connection_folder, 'bcgw.sde')):
        os.remove(os.path.join(connection_folder, 'bcgw.sde'))

    # Create a bcgw connection
    bcgw_con = arcpy.management.CreateDatabaseConnection(connection_folder,
                                                        'bcgw.sde',
                                                        'ORACLE',
                                                        'bcgw.bcgov/idwprod1.bcgov',
                                                        'DATABASE_AUTH',
                                                        DB_USER,
                                                        DB_PASS,
                                                        'DO_NOT_SAVE_USERNAME')

    print("new db connection created")
    self.logger.info("new db connection created")


    arcpy.env.workspace = bcgw_con.getOutput(0)

    print("workspace set to bcgw connection")
    self.logger.info("workspace set to bcgw connection")
    
    secrets = [DB_USER, DB_PASS]
    
    return secrets
###############################################################################################################################################################################

class AST_FACTORY:
    ''' AST_FACTORY creates and manages status tool runs '''
    XLSX_SHEET_NAME = 'ast_config'
    AST_PARAMETERS = {
        0: 'region',
        1: 'feature_layer',
        2: 'crown_file_number',
        3: 'disposition_number',
        4: 'parcel_number',
        5: 'output_directory',
        6: 'output_directory_same_as_input',
        7: 'dont_overwrite_outputs',
        8: 'skip_conflicts_and_constraints',
        9: 'suppress_map_creation',
        10: 'add_maps_to_current',
        11: 'run_as_fcbc',

    }
    
    ADDITIONAL_PARAMETERS = {
        12: 'ast_condition',
        13: 'file_number'
    }
    
    AST_CONDITION_COLUMN = 'ast_condition'
    AST_SCRIPT = ''
    job_index = None  # Initialize job_index as a global variable
    
    def __init__(self, queuefile, db_user, db_pass, logger) -> None:
        self.user = db_user
        self.user_cred = db_pass
        self.queuefile = queuefile
        self.jobs = []
        self.logger = logger
        
    # Set up logging

    def setup_logging():
        ''' Set up logging for the script '''
        # Create the log folder filename
        log_folder = f'autoast_logs_{datetime.datetime.now().strftime("%Y%m%d")}'

        # Create the log folder in the current directory if it doesn't exits
        if not os.path.exists(log_folder):
            os.mkdir(log_folder)
        
        # Check if the log folder was created successfully
        assert os.path.exists(log_folder), "Error creating log folder, check permissions and path"

        # Create the log file path with the date and time appended
        log_file = os.path.join(log_folder, f'ast_log_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.log')



        # Set up logging config to DEBUG level
        logging.basicConfig(filename=log_file, 
                            level=logging.DEBUG, 
                            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

        # Create the self.logger object and set to the current file name
        logger = logging.getself.logger(__name__)

        print("Logging set up")
        logger.info("Logging set up")

        print("Starting Script")
        logger.info("Starting Script")
        
        return logger

    def load_jobs(self):
        '''
        load jobs will check for the exisitance of the queuefile, if it exists it will load the jobs from the queuefile. Checking if they 
        are Complete and if not, it will add them to the jobs list as Queued
        '''
        
        global job_index

        print("Loading jobs")
        self.logger.info("Loading jobs")

        # Initialize the jobs list to store jobs
        self.jobs = []

        # Check if the queue file exists
        assert os.path.exists(self.queuefile), "Queue file does not exist"
        if os.path.exists(self.queuefile):

            try:
                # Open the Excel workbook and select the correct sheet
                wb = load_workbook(filename=self.queuefile)
                ws = wb[self.XLSX_SHEET_NAME]

                # Get the header (column names) from the first row of the sheet
                header = list([row for row in ws.iter_rows(min_row=1, max_col=None, values_only=True)][0])

                # Read all the data rows (starting from the second row to skip the header)
                data = [row for row in ws.iter_rows(min_row=2, max_col=None, values_only=True)]

                # Iterate over each row of data; enumerate to keep track of the row number in Excel
                for row_index, row_data in enumerate(data, start=2):  # Start from 2 to account for Excel header

                    # Skip any completely blank rows
                    if all((value is None or str(value).strip() == '') for value in row_data):
                        print(f"Skipping blank row at index {row_index}")
                        self.logger.info(f"Skipping blank row at index {row_index}")
                        continue  # Skip this row entirely

                    # Initialize a dictionary to store the job's parameters
                    job = {}
                    ast_condition = None  # Initialize the ast_condition for the current row

                    # Loop through each column header and corresponding value in the current row
                    for key, value in zip(header, row_data):
                        # Check if the key corresponds to the ast_condition column
                        if key is not None and key.lower() == self.AST_CONDITION_COLUMN.lower():
                            ast_condition = value if value is not None else ""

                        # Assign an empty string to any None values
                        value = "" if value is None else value

                        # Assign the value to the job dictionary if the key is not None
                        if key is not None:
                            job[key] = value

                    # Skip if marked as "COMPLETE"
                    if ast_condition.upper() == 'COMPLETE':
                        print(f"Skipping job {row_index - 1} as it is marked COMPLETE.")
                        self.logger.info(f"Skipping job {row_index - 1} as it is marked COMPLETE.")
                        continue  # Skip this job as it's already marked as COMPLETE

                    # Check if the ast_condition is None, empty, or not 'COMPLETE'
                    if ast_condition is None or ast_condition.strip() == '' or ast_condition.upper() != 'COMPLETE':
                        # Assign 'Queued' to the ast_condition and update the job dictionary
                        ast_condition = 'Queued'
                        job[self.AST_CONDITION_COLUMN] = ast_condition
                        self.logger.info(f"Job {row_index - 1} is {ast_condition}")

                        # Immediately update the Excel sheet with the new condition
                        try:
                            self.add_job_result(row_index - 1, ast_condition)
                            self.logger.info(f"Added job condition '{ast_condition}' for job {row_index - 1} to jobs list")
                        except Exception as e:
                            print(f"Error updating Excel sheet at row {row_index}: {e}")
                            self.logger.error(f"Error updating Excel sheet at row {row_index}: {e}")
                            continue

                    # Classify the input type for the job
                    try:
                        self.classify_input_type(job)
                        print(f"Classifying input type for job {row_index - 1}")
                        self.logger.info(f"Classifying input type for job {row_index - 1}")
                    except Exception as e:
                        print(f"Error classifying input type for job {job}: {e}")
                        self.logger.error(f"Error classifying input type for job {job}: {e}")

                    # Add the job to the jobs list after all checks and processing
                    self.jobs.append(job)
                    print(f"Job Condition is not Complete ({ast_condition}), adding job: {row_index - 1} to jobs list")
                    self.logger.info(f"Job Condition is not Complete ({ast_condition}), adding job: {row_index - 1} to jobs list")

                    print(f"Job dictionary is {job}")
                    self.logger.info(f"Job dictionary is {job}")

            except FileNotFoundError as e:
                print(f"Error: Queue file not found - {e}")
                self.logger.error(f"Error: Queue file not found - {e}")
            except Exception as e:
                print(f"Unexpected error loading jobs: {e}")
                self.logger.error(f"Unexpected error loading jobs: {e}")

            return self.jobs



    def classify_input_type(self, input):
        ''' If the input file is a .kml it will build the aoi from the kml.
        If it is a .shp it will build the aoi based on the shapefile.
        If it is a shapefile AND a filenumber is present, it will run the FW setup script on the shapefile, writing the appended 
        shapefile to an output directory based on the file numer. This step of writing the appended shapefile to an output directory
        might be able to be removed'''
        

        # input_type = None
        # file_name, extension = os.path.basename(input).split()

        
        for job in self.jobs:                # Check if there is a file path in Feature Layer
            if job.get('feature_layer'):
                print(f'Feature layer found: {job["feature_layer"]}')
                self.logger.info(f'Feature layer found: {job["feature_layer"]}')
                feature_layer_path = job['feature_layer']
                print(f"Processing feature layer: {feature_layer_path}")
                self.logger.info(f"Processing feature layer: {feature_layer_path}")

                if feature_layer_path.lower().endswith('.kml'):
                    print(f'Kml found, building AOI from KML')
                    self.logger.info(f'Kml found, building AOI from KML')
                    job['feature_layer'] = self.build_aoi_from_kml(feature_layer_path)
                elif feature_layer_path.lower().endswith('.shp'):
                    if job.get('file_number'):
                        print(f"File number found for job {job_index}, running FW setup on shapefile: {feature_layer_path}")
                        self.logger.info(f"File number found for job {job_index}, running FW setup on shapefile: {feature_layer_path}")
                        new_feature_layer_path = self.build_aoi_from_shp(job, feature_layer_path)
                        job['feature_layer'] = new_feature_layer_path
                    else:
                        print(f'No FW File Number provided for the shapefile, loading original shapefile path')
                        self.logger.info(f'No FW File Number provided for the shapefile, loading original shapefile path')
                else:
                    print(f"Unsupported feature layer format: {feature_layer_path}")
                    self.logger.warning(f"Job number {job_index} Unsupported feature layer format: {feature_layer_path}")
                    self.add_job_result(job_index, 'Failed')

            
    
    
    def start_ast_tb(self, jobs):
        global job_index
        '''Starts an AST toolbox from job params. It will check the capitalization of the True or False inputs and 
        change them to appropriate booleans as the script was failing before implementing this.
        It will also create the output directory for the FW Setup Shapefiles, if it does not exist based on the job number. Currently this is being created in the T: drive.
        but should be updated once on the server. It checks to make a sure a region has been input on the excel sheet as this is a required parameter.
        It will also catch any errors that are thrown and print them to the console.'''
        try:
            print("Starting AST Toolbox")
            logging.info("Starting AST Toolbox")

            #DELETE the batch function should loop the spreadsheet and run the start_ast_tb function on each row of the excel sheet
            # Loop over the jobs in the spreadsheet#
            for job in jobs:
                params = []
                
                # Apply a separator line between each job in the log file
                
                self.logger.info(f"===================================================================")
                self.logger.info(f"======================= Starting Job #: {job} ======================")
                self.logger.info(f"====================================================================")
                try:
                    # Convert 'true'/'false' strings to booleans (For some reason the script was reading them all as lowercase strings)
                    for param in self.AST_PARAMETERS.values():
                        value = job[param]
                        if isinstance(value, str) and value.lower() in ['true', 'false']:
                            value = True if value.lower() == 'true' else False
                        params.append(value)

                    # Ensure that region has been entered otherwise job will fail
                    if not job.get('region'):
                        raise ValueError("Region is required and was not provided.")

                    # Run the ast tool 
                    print(f"Job Parameters are: {params}")
                    self.logger.info(f"Job Parameters are: {params}")
                    arcpy.MakeAutomatedStatusSpreadsheet_ast(*params)
                    
                    self.capture_arcpy_messages()
                    if arcpy.GetMessages(0) == "Automated_status_sheet.xlsx is ready for you to use":
                        self.logger.info(f"AST SUCCESS MESSAGE RECEIVED. Job {counter} COMPLETE")

                    print(f"AST SUCCESS MESSAGE RECEIVED. Job {counter} COMPLETE")
                    self.logger.info(f"AST SUCCESS MESSAGE RECEIVED. Job {counter} COMPLET ****Automated_status_sheet.xlsx is ready for you to useE")
                    self.logger.info("***************************************************************************************************************************")

                    job_index = self.jobs.index(job)

                    self.add_job_result(job_index, 'COMPLETE')
                    #self.add_job_result(job)
                    
                    return job_index
                    
                except KeyError as e:
                    print(f"Error: Missing parameter in the excel queuefile: {e}")
                    self.logger.error(f"Error: Missing parameter in the excel queuefile: {e}")
                except ValueError as e:
                    print(f"Error: {e}")
                    self.logger.error(f"Error: {e}")
                except arcpy.ExecuteError as e:
                    print(f"Arcpy error: {arcpy.GetMessages(2)}")
                    self.logger.error(f"Arcpy error: {arcpy.GetMessages(2)}")
                except Exception as e:
                    print(f"Unexpected error processing job: {e}")
                    self.logger.error(f"Unexpected error processing job: {e}")

        except ImportError as e:
            print(f"Error importing arcpy toolbox. Check file path in .env file: {e}")
            self.logger.error(f"Error importing arcpy toolbox. Check file path in .env file: {e}")
        except arcpy.ExecuteError as e:
            print(f"Arcpy error: {arcpy.GetMessages(2)}")
            self.logger.error(f"Arcpy error: {arcpy.GetMessages(2)}")
        except Exception as e:
            print(f"Unexpected error: {e}")
            self.logger.error(f"Unexpected error: {e}")
            
    def add_job_result(self, job_index, condition):
        ''' 
        Function adds result information to the Excel spreadsheet. If the job is successful, it will update the ast_condition column to "COMPLETE",
        if the job failed, it will update the ast_condition column to "Failed".
        '''

        print("Adding job result...")
        self.logger.info("Adding job result...")

        try:
            # Load the workbook
            wb = load_workbook(filename=self.queuefile)
            # Load the correct worksheet
            ws = wb[self.XLSX_SHEET_NAME]

            # Read the header index for the ast_condition column
            header = list([row for row in ws.iter_rows(min_row=1, max_col=None, values_only=True)][0])
            ast_condition_index = header.index(self.AST_CONDITION_COLUMN) + 1  # +1 because Excel columns are 1-indexed

            # Calculate the actual row index in Excel, +2 to account for header and 0-index
            excel_row_index = job_index + 2  

            # Check if the row is blank before updating
            row_values = [ws.cell(row=excel_row_index, column=col).value for col in range(1, len(header) + 1)]
            if all(value is None or str(value).strip() == '' for value in row_values):
                print(f"Row {excel_row_index} is blank, not updating.")
                self.logger.info(f"Row {excel_row_index} is blank, not updating.")
                return  # Do not update if the row is blank

            # Update the condition for the specific job
            ws.cell(row=excel_row_index, column=ast_condition_index, value=condition)

            # Save the workbook with the updated condition
            wb.save(self.queuefile)
            print(f"Updated row {excel_row_index} with condition '{condition}'.")
            self.logger.info(f"Updated row {excel_row_index} with condition '{condition}'.")

        except FileNotFoundError as e:
            print(f"Error: Queue file not found - {e}")
            self.logger.error(f"Error: Queue file not found - {e}")

        except IndexError as e:
            print(f"Error: Index out of range when accessing row {excel_row_index} - {e}")
            self.logger.error(f"Error: Index out of range when accessing row {excel_row_index} - {e}")

        except PermissionError as e:
            print(f"Error: Permission denied when trying to access the Excel file - {e}")
            self.logger.error(f"Error: Permission denied when trying to access the Excel file - {e}")

        except Exception as e:
            print(f"Unexpected error while adding job result: {e}")
            self.logger.error(f"Unexpected error while adding job result: {e}")


        
 


    def batch_ast(self):
        global counter, job_index 
        ''' Executes the loaded jobs'''

        counter = 1
        print("Batching AST")
        
        self.logger.info("***************************************************************************************************************************")
        self.logger.info("Batching AST")        
        self.logger.info("***************************************************************************************************************************")
        number_of_jobs = len(self.jobs)
        print(f"Number of jobs: {number_of_jobs}")
        self.logger.info(f"Number of jobs: {number_of_jobs}")
        
        # iterate through the jobs and run the start_ast_tb function on each row of the excel sheet
        for job in self.jobs:
            try:
                print(f"Starting job {counter} of {number_of_jobs}")
                self.logger.info(f"Starting job {counter} of {number_of_jobs}")
                
                # Start the Ast Tool
                self.start_ast_tb([job])
                #TODO insert time out function here
                # If the arcpymessage that is returned contains the text "Automated_status_sheet.xlsx is ready for you to use" then the job is complete
                if arcpy.GetMessages(0) == "Automated_status_sheet.xlsx is ready for you to use":
                    self.add_job_result(job_index, 'COMPLETE')

                    print(f"AST SUCCESS MESSAGE RECEIVED. Job {counter} COMPLETE")
                    self.logger.info(f"AST SUCCESS MESSAGE RECEIVED. Job {counter} COMPLETE")

            except Exception as e:
                # Log the exception and the job that caused it
                print(f"Error encountered with job {counter}: {e}")
                self.logger.error(f"Error encountered with job {counter}: {e}")
                self.add_job_result(job_index, 'Failed')
            finally:
                counter += 1

    
         
            
    def re_batch_failed_ast(self):
        global counter
        ''' Executes the loaded failed jobs'''

        counter = 1
        print("Re Batching AST")
        
        self.logger.info("***************************************************************************************************************************")
        self.logger.info("Re Batching Failed AST")        
        self.logger.info("***************************************************************************************************************************")
        print(f"Number of failed jobs: {len(self.jobs)}")
        self.logger.info(f"Number of failed jobs: {len(self.jobs)}")
        
        # iterate through the jobs and run the start_ast_tb function on each row of the excel sheet
        for job in self.jobs:
            try:
                print(f"Starting job {counter}")
                self.logger.info(f"Starting job {counter}")
                
                # Start the Ast Tool
                self.start_ast_tb([job])
                print(f"Job {counter} COMPLETE")
                self.logger.info(f"Job {counter} COMPLETE")
                self.add_job_result(job_index, 'COMPLETE')

            except Exception as e:
                # Log the exception and the job that caused it
                print(f"Error encountered with job {counter}: {e}")
                self.logger.error(f"Error encountered with job {counter}: {e}")
                self.add_job_result(job_index, 'Failed')
            finally:
                counter += 1



    def re_load_failed_jobs(self):
        global job_index

        print("Loading Failed jobs")
        self.logger.info("Loading Failed jobs")

        # Initialize the jobs list to store jobs
        self.jobs = []
    
        try:
            # Open the Excel workbook and select the correct sheet
            wb = load_workbook(filename=self.queuefile)
            ws = wb[self.XLSX_SHEET_NAME]
            
            # Get the header (column names) from the first row of the sheet
            header = list([row for row in ws.iter_rows(min_row=1, max_col=None, values_only=True)][0])
            
            # Read all the data rows (starting from the second row to skip the header)
            data = [row for row in ws.iter_rows(min_row=2, max_col=None, values_only=True)]

            # Iterate over each row of data; enumerate to keep track of the row number in Excel
            for row_index, row_data in enumerate(data, start=1):  # Start from 2 to account for Excel header

                # Skip any completely blank rows
                # Skip any completely blank rows
                if all((value is None or str(value).strip() == '') for value in row_data):
                    print(f"Skipping blank row at index {row_index}")
                    self.logger.info(f"Skipping blank row at index {row_index}")
                    continue


                # Initialize a dictionary to store the job's parameters
                job = {}
                ast_condition = None  # Initialize the ast_condition for the current row

                # Loop through each column header and corresponding value in the current row
                for key, value in zip(header, row_data):
                    # Check if the key corresponds to the ast_condition column
                    if key is not None and key.lower() == self.AST_CONDITION_COLUMN.lower():
                        if value == 'Failed':
                            # If the value exists, assign it to ast_condition
                            ast_condition = value
                            job[key] = value


                    # Assign the value to the job dictionary if the key is not None
                    # if key is not None:
                    #     print(f"Re running job value is {value}")
                    #     self.logger.info(f"Re running job value is {value}")
                    #     job[key] = value

                # Check if the ast_condition is None, empty, or not 'COMPLETE'
                if ast_condition is value:
                    # Assign 'Queued' to the ast_condition and update the job dictionary
                    ast_condition = 'Queued'
                    job[self.AST_CONDITION_COLUMN] = ast_condition

                    # Immediately update the Excel sheet with the new condition
                    try:
                        self.add_job_result(row_index - 2, ast_condition)  
                    except Exception as e:
                        print(f"Error updating Excel sheet at row {row_index}: {e}")
                        self.logger.error(f"Error updating Excel sheet at row {row_index}: {e}")
                        continue
                    
                # Immediately update the Excel sheet with the new condition
                    try:
                        self.add_job_result(row_index - 2, ast_condition)  
                    except Exception as e:
                        print(f"Error updating Excel sheet at row {row_index}: {e}")
                        self.logger.error(f"Error updating Excel sheet at row {row_index}: {e}")
                        continue  

                # Add the job to the jobs list
                self.jobs.append(job)
                print(f"Job Condition is Queued', adding job: {row_index} to jobs list")
                self.logger.info(f"Job Condition is 'Queued', adding job: {row_index} to jobs list")

                print(f"Job dictionary is {job}")
                self.logger.info(f"Job dictionary is {job}")

            # Classify input types for all loaded jobs
            print("Classifying input types for failed jobs")
            self.logger.info("Classifying input types for failed jobs")
            for job in self.jobs:
                try:
                    self.classify_input_type(job)
                except Exception as e:
                    print(f"Error classifying input type for failed job {job}: {e}")
                    self.logger.error(f"Error classifying input type for failed job {job}: {e}")

        except FileNotFoundError as e:
            print(f"Error: Queue file not found - {e}")
            self.logger.error(f"Error: Queue file not found - {e}")
        except Exception as e:
            print(f"Unexpected error loading jobs: {e}")
            self.logger.error(f"Unexpected error loading jobs: {e}")

        return self.jobs

        
    def create_new_queuefile(self):
        '''write a new queuefile with preset header'''
        print("Creating new queuefile")
        self.logger.info("Creating new queuefile")
        wb = Workbook()
        ws = wb.active
        ws.title = self.XLSX_SHEET_NAME
        headers = list(self.AST_PARAMETERS.values())
        headers.append(self.AST_CONDITION_COLUMN)
        for h in headers:
            c = headers.index(h) + 1
            ws.cell(row=1, column=c).value = h
        wb.save(self.queuefile)

    def build_aoi_from_kml(self, aoi):
        "Write shp file for temporary use"

        # Ensure the KML file exists
        if not os.path.exists(aoi):
            raise FileNotFoundError(f"The KML file '{aoi}' does not exist.")

        print("Building AOI from KML")
        self.logger.info("Building AOI from KML")
        from fiona.drvsupport import supported_drivers
        supported_drivers['LIBKML'] = 'rw'
        tmp = os.getenv('TEMP')
        if not tmp:
            raise EnvironmentError("TEMP environment variable is not set.")
        bname = os.path.basename(aoi).split('.')[0]
        fc = bname.replace(' ', '_')
        out_name = os.path.join(tmp, bname + '.gdb')
        if os.path.exists(out_name):
            shutil.rmtree(out_name, ignore_errors=True)
        df = geopandas.read_file(aoi)
        df.to_file(out_name, layer=fc, driver='OpenFileGDB')

        #DELETE
        print(f' kml ouput is {out_name} / {fc}')
        self.logger.info(f' kml ouput is {out_name} / {fc}')
        return out_name + '/' + fc

    def build_aoi_from_shp(self, job, feature_layer_path):
        """This is snippets of Mike Eastwoods FW Setup Script, if run FW Setup is set to true **Not sure if we need this
        as an option or just make it standard.
        This function will take the raw un-appended shapefile and run it through the FW Setup Script"""

        # Mike Eastwoods FW Setup Script
        print("Processing shapefile using FW Setup Script")
        self.logger.info("Processing shapefile using FW Setup Script")
        
        fsj_workspace = os.getenv('FSJ_WORKSPACE')
        arcpy.env.workspace = fsj_workspace
        arcpy.env.overwriteOutput = False

        # Check if there is a file path in Feature Layer
        if feature_layer_path:
            print(f"Processing feature layer: {feature_layer_path}")
            self.logger.info(f"Processing feature layer: {feature_layer_path}")

        # Check to see if a file number was entered in the excel sheet, if so, use it to name the output directory and authorize the build_aoi_from_shp function to run
        file_number = job.get('file_number')

        if not file_number:
            raise ValueError("Error: File Number is required if you are putting in a shapefile that has not been processed in the FW Setup Tool.")
        else:
            print(f"Running FW Setup on File Number: {file_number}")
            self.logger.info(f"Running FW Setup on File Number: {file_number}")

        # Convert file_number to string and make it uppercase
        file_number_str = str(file_number).upper()

        # Calculate date variables
        date = datetime.date.today()
        year = str(date.year)

        # Set variables
        base = arcpy.env.workspace
        baseYear = os.path.join(base, year)
        outName = file_number_str
        geometry = "POLYGON"
    
        m = "SAME_AS_TEMPLATE"
        z = "SAME_AS_TEMPLATE"
        spatialReference = arcpy.Describe(template).spatialReference

        # ===========================================================================
        # Create Folders
        # ===========================================================================

        print("Creating FW Setup folders . . .")
        self.logger.info("Creating FW Setup folders . . .")
        outName = file_number_str

        # Create path to folder location
        fileFolder = os.path.join(baseYear, outName)
        shapeFolder = fileFolder
        outPath = shapeFolder
        if os.path.exists(fileFolder):
            print(outName + " folder already exists.")
            self.logger.info(outName + " folder already exists.")
        else:
            os.mkdir(fileFolder)

        # ===========================================================================
        # Create Shapefile(s) and add them to the current map
        # ===========================================================================

        print("Creating Shapefiles using FW Setup . . .")
        self.logger.info("Creating Shapefiles using FW Setup . . .")
        if os.path.isfile(os.path.join(outPath, outName + ".shp")):
            print(os.path.join(outPath, outName + ".shp") + " already exists")
            self.logger.info(os.path.join(outPath, outName + ".shp") + " already exists")
            print("Exiting without creating files")
            self.logger.info("Exiting without creating files")
            return os.path.join(outPath, outName + ".shp")
        else:
            # Creating template shapefile
            create_shp = arcpy.management.CreateFeatureclass(outPath, outName, geometry, template, m, z, spatialReference)
            # Append the newly created shapefile with area of interest
            append_shp = arcpy.management.Append(feature_layer_path, create_shp, "NO_TEST")
            print("Append Successful")
            self.logger.info("Append Successful")
            # Making filename for kml
            create_kml = os.path.join(outPath, outName + ".kml")
            # Make layer for kml to be converted from 
            layer_shp = arcpy.management.MakeFeatureLayer(append_shp, outName)
            # Populate the shapefile                          
            arcpy.conversion.LayerToKML(layer_shp, create_kml)
            # Send message to user that kml has been created
            print("kml created: " + create_kml)
            self.logger.info("kml created: " + create_kml)

            print(f"FW Setup complete, returned shapefile is {os.path.join(outPath, outName + '.shp')}")
            self.logger.info(f"FW Setup complete, returned shapefile is {os.path.join(outPath, outName + '.shp')}")

            return os.path.join(outPath, outName + ".shp")
        

    def capture_arcpy_messages(self):
        ''' Re assigns the arcpy messages  (0 for all messages, 1 for warnings, and 2 for errors) to variables and passes them to the self.logger'''
        
        arcpy_messages = arcpy.GetMessages(0) # Gets all messages
        arcpy_warnings = arcpy.GetMessages(1) # Gets all warnings only
        arcpy_errors = arcpy.GetMessages(2) # Gets all errors only
        
        if arcpy_messages:
            self.logger.info(f'ast_toobox arcpy messages: {arcpy_messages}')
        if arcpy_warnings:
            self.logger.warning(f'ast_toobox arcpy warnings: {arcpy_warnings}')
        if arcpy_errors:
            self.logger.error(f'ast_toobox arcpy errors: {arcpy_errors}')   

if __name__ == '__main__':
    current_path = os.path.dirname(os.path.realpath(__file__))
    
    # Call the setup_logging function to log the messages
    self.logger = setup_logging()
    
    # Load the default environment
    load_dotenv()
    
    # Call the import_ast function to import the AST toolbox
    template = import_ast()
    
    # Call the setup_bcgw function to set up the database connection
    secrets = setup_bcgw()
    
    # Create the path for the queuefile
    qf = os.path.join(current_path, excel_file)

    # Create and instance of the Ast Factory class, assign the quefile path and the bcgw username and passwords to the instance
    ast = AST_FACTORY(qf, secrets[0], secrets[1])



    if not os.path.exists(qf):
        print("Queuefile not found, creating new queuefile")
        self.logger.info("Queuefile not found, creating new queuefile")
        ast.create_new_queuefile()
        
    # load the jobs using the load jobs method. This will scan the excel sheet and assign to "jobs"    
    jobs = ast.load_jobs()
    

    ast.batch_ast()
    
    # ast.re_load_failed_jobs()
    
    # ast.re_batch_failed_ast()


    print("AST Factory COMPLETE")
    self.logger.info("AST Factory COMPLETE")