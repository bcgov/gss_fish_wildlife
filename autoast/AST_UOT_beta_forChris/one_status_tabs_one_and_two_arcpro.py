'''

    Author:         Mark McGirr
    Edited by:      Wes Smith
    
    Purpose:        This module is intended to be called from another script.
                    The final output created by this code is one_status_tabs_1_and_2.xlsx
                   
    Date:           November, 2012
    Revised Date:   March, 2020
    
    Arguments:      self.data_path = passed_input_list[0]        
                    self.dont_overwrite_outputs = passed_input_list[1]
                    self.region = passed_input_list[2]
                    self.crown_lands_file = passed_input_list[3]
                    self.disposition_transaction = passed_input_list[4]
 
    Logic Overview: 

                  
    History:      
    --------------------------------------------------------------
    Date: March 2020
    Author: Wes Smith
    Modification: Updated Syntax from Python 2.1 to Python 3 for use in ArcGIS Pro.  
                Some updates to methods (use of libraries and lists for looping).
                Use of Openpyxl for the creation, population, and formating of the 
                Spreadsheet.
    --------------------------------------------------------------
    
'''

import sys, string, os, time, datetime, arcpy, csv, runpy    #@UnusedImport
import openpyxl
import keyring
from openpyxl import Workbook   #@UnusedImport
from openpyxl.styles import Alignment, NamedStyle, Font, Fill, PatternFill, colors, Color #Border, Side #@UnusedImport
from openpyxl.styles.borders import Border, Side

sys.path.append(r'\\GISWHSE.ENV.GOV.BC.CA\WHSE_NP\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\beta')
#sys.path.append(r'\\GISWHSE.ENV.GOV.BC.CA\WHSE_NP\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\Scripts')
import universal_overlap_tool_arcpro as revolt
import inactive_dispositions as inactives
import config

#------------------------------------------------------------------------------ 
arcpy.env.overwriteOutput = True

class one_status_part2_tool(object):
    # this is the object that creates the one_status_tabs_1_and_2.xlsx 
    
    def __init__(self):
        self.confirmation = "Tool Object Created"
        #arcpy.AddMessage(self.confirmation)

    def run_tool(self, passed_input_list):
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("running the one status part 2 tool")
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("Passed Inputs:")
        i = 0
        for item in passed_input_list:
            message = ("{}: {}".format(i, str(item)))
            arcpy.AddMessage(message)
            i += 1
        arcpy.AddMessage("======================================================================")
        
        self.sde_connection = os.getenv("SDE_FILE_PATH")

        self.split_passed_in_argument_into_individual_variables(passed_input_list)
        self.create_working_directories_geodatabases_and_variables()
        
        self.xls_name_to_save_as = os.path.join(self.work_directory,'one_status_tabs_1_and_2.xlsx')
        
        self.copy_aoi_into_analysis_gdb()
        self.get_all_needed_datasets()
        self.create_excel_cell_variables()
        self.inactive_features = self.get_inactives()
        self.create_spreadsheet_openpyxl()
        self.create_xls_files_section()
        self.create_xls_applicant_section2()
        self.create_xls_lands_section2()
        self.create_xls_adjudication_section2()
        self.create_xls_mines_section2()
        self.create_xls_forests_section2()
        self.create_xls_water_section2()
        self.create_xls_comments_section2()
        #        stop
        if self.region == 'omineca':
            self.create_xls_omineca_section2()

        self.warn_csv_files()
        self.create_spreadsheet_status_of_conflict2()

        # self.read_csv_files()
        # self.create_spreadsheet_status_of_conflict2_csv()


        arcpy.AddMessage("\nSpreadsheet Created:\n   {}".format(self.xls_name_to_save_as))

    def split_passed_in_argument_into_individual_variables(self, passed_input_list):
        '''
        this splits the passed in argument into individual variables
        '''
        
        # User Inputs
        self.data_path = passed_input_list[0]
        self.dont_overwrite_outputs = passed_input_list[1]
        self.region = passed_input_list[2]
        self.crown_lands_file = str(passed_input_list[3])
        self.disposition_transaction = str(passed_input_list[4])
        self.parcel_number = str(passed_input_list[5])
        self.analyize_this_featureclass = passed_input_list[6]
        
        # Derived
        self.directory_to_store_output = self.data_path
        
        # Hard-coded
        self.what_type_of_overlap_to_run = "one_status"
        self.dont_overwrite_data_and_maps = 'true'
        
        # Report
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("The input criteria are:")
        arcpy.AddMessage("data_path = " +                   self.data_path)
        arcpy.AddMessage("analyize_this_featureclass = " +  self.analyize_this_featureclass)
        arcpy.AddMessage("working_directory = " +           self.directory_to_store_output)
        arcpy.AddMessage("dont_overwrite_outputs = " +      self.dont_overwrite_outputs)
        arcpy.AddMessage("region = " +                      self.region)
        
        arcpy.AddMessage("crown_lands_file = " +            self.crown_lands_file)
        arcpy.AddMessage("disposition_transaction = " +     self.disposition_transaction)
        arcpy.AddMessage("parcel_number = " +               self.parcel_number)
        
        
    def create_working_directories_geodatabases_and_variables(self):
        '''
        Creates the working directory.  Deletes the existing working GDB and MAP directory 
        if the 'dont_overwrite_flag' is not set to true.  It then creates the MAP directory.
        
        
        Also sets the following variables for universal use. 
        
        self.work_directory     -    the directory where the work will be performed
                                ie.  \\granite\work\srm\wml\workarea\arcproj\!williams_lake_toolbox_development\5407649
        
        self.work_gdb           -    the complete path to the working GDB file, with a .GDB extension
                                ie.  \\granite\work\srm\wml\workarea\arcproj\!williams\5407649\one_status_common_datasets_aoi.gdb
        
        xls_for_analysis        -    the name of the spreadsheet to be read without the XLS extension.
                                ie.  one_status_common_datasets_debug_version

        featureclass_to_analyize  -  aoi (should always be this because the parcel boundary is copied into the GDB with this name
        
        self.map_directory      -    the complete path to the directory where the maps will be created.
                                ie.  \\granite\work\srm\wml\workarea\arcproj\!williams\5407649\maps_for_one_status_common_datasets_aoi
        
        '''
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Creating working directories, geodatabases, variables')


        #this is the default directory for where the input XLS files are stored (p:)
        #self.input_criteria_xls_directory = r"\\giswhse\whse_np\corp\script_whse\python\Utility_Misc\Ready\williams_lake_tools"
        # OCT 2013 new statusing_tools directory structure
        self.input_criteria_xls_directory = r"\\giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\statusing_input_spreadsheets"
        
        # if the type of overlap to run is selected in the dropdown box in the tool, and the xls not specified 
        if self.what_type_of_overlap_to_run != "#"  and self.what_type_of_overlap_to_run != "":
            self.xls_file_for_analysis_input = os.path.join(self.input_criteria_xls_directory, self.what_type_of_overlap_to_run + ".xlsx")
    
        # this just returns the name of the xls file without the .xls extension to print on the top of output report
        garbage_variable, self.xls_for_analysis = os.path.split(self.xls_file_for_analysis_input)   #@UnusedVariable
        self.xls_for_analysis, extension_type = os.path.splitext(self.xls_for_analysis)             #@UnusedVariable
        
        
        #print arcpy.AddMessage("using this file for analysis criteria   " + xls_file_for_analysis_input )
        self.excel_input = self.xls_file_for_analysis_input

        # determine what the output (working) directory is.  If it is specified by the user in the tool use that directory instead
        if self.directory_to_store_output != "#" and self.directory_to_store_output != "":
            self.work_directory = self.directory_to_store_output  # choose the user defined directory to store the output xls and maps
        else:
            #dataObj = wml_library.dataUtil_mcgirr()
            #self.work_directory = wml_library.get_fc_directory_name(self.analyize_this_featureclass)
            self.work_directory = revolt.get_fc_directory_name(self.analyize_this_featureclass)  #@UndefinedVariable
            #del dataObj


        #just the name of the fc to analyze with no path prefix, and no .shp extension
        garbage_variable, self.featureclass_to_analyize = os.path.split(self.analyize_this_featureclass)    #@UnusedVariable
        prefix, extension_type = os.path.splitext(self.featureclass_to_analyize)
        if extension_type == ".shp":
            self.featureclass_to_analyize = prefix

        self.work_gdb = os.path.join(self.work_directory, self.xls_for_analysis + "_tabs_1_and_2_datasets.gdb")
#         self.work_gdb = os.path.join(self.work_directory, self.xls_for_analysis + "_" + self.featureclass_to_analyize + ".gdb")
#         self.map_directory = os.path.join(self.work_directory, "maps_for_" + self.xls_for_analysis + "_" + self.featureclass_to_analyize)
        
        
        #print("excel_input  " , self.excel_input)
        #print("self.work_directory  " , self.work_directory)
        #print("work_directory  " , self.work_directory)
        #print("self.work_gdb  " , self.work_gdb)
        #print("xls_for_analysis " , self.xls_for_analysis)
        #print("featureclass_to_analyize  " , self.featureclass_to_analyize)
        #print("self.map_directory  " , self.map_directory)
        
        
        # delete the maps directory, and output GDB if flag is not set to true.    
        if  arcpy.Exists(self.work_gdb) and self.dont_overwrite_data_and_maps != 'true' :
            arcpy.AddMessage("deleting the existing working gdb")
            arcpy.Delete_management(self.work_gdb)
#         if  arcpy.Exists(self.map_directory) and  self.dont_overwrite_data_and_maps != 'true':
#             arcpy.AddMessage("deleting the existing map directory")
#             arcpy.Delete_management(self.map_directory)
#         if not os.path.exists(self.map_directory):
#             os.makedirs(self.map_directory)  # make the directory to store the maps

    
        # create the GDB and feature dataset to do the work in
        self.input_dataset = os.path.join(self.work_gdb, "input_of_raw_data")
        self.create_feature_dataset_if_needed(self.input_dataset) #@UndefinedVariable
    
    
    def copy_aoi_into_analysis_gdb(self):
        '''
        Copies the parcel boundary into the working gdb.  Buffers it at 1 meter if it a point or line
        because identities need a polygon.
        
        Adds a label field to the created AOI to be used in writing labels on the maps.
        '''
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Copying the AOI into the working .GDB')

        # copy the input data shape to be the AOI.  If it's a point or line,  buffer it to be the AOI
        arcpy.AddMessage(".   Creating the raw AOI ")
        desc = arcpy.Describe(self.analyize_this_featureclass)
        input_data_type = desc.ShapeType
        raw_output = os.path.join(self.input_dataset, "raw_aoi")
        if not arcpy.Exists(raw_output) :
            arcpy.CopyFeatures_management(self.analyize_this_featureclass, raw_output)
        
        # buffer the shape at 1 meter if it's not a polygon because identities need polygons.  This is now the AOI
        the_output = os.path.join(self.work_gdb, "aoi")
        if not arcpy.Exists(the_output) and input_data_type == "Polygon":
            arcpy.AddMessage(".   Creating the final AOI ")
            arcpy.CopyFeatures_management(raw_output, the_output)
            arcpy.AddField_management(the_output, "label_field", "TEXT", "", "", "40", "", "NULLABLE", "NON_REQUIRED", "")
    
        if not arcpy.Exists(the_output) and input_data_type != "Polygon":
            arcpy.AddMessage(".   Buffering the line or point feature class into final AOI ")
            arcpy.Buffer_analysis(raw_output, the_output, 1) 
            arcpy.AddField_management(the_output, "label_field", "TEXT", "", "", "40", "", "NULLABLE", "NON_REQUIRED", "")
    
    
    def create_feature_dataset_if_needed(self, fds_to_create):
        '''
        This creates a feature dataset in the provided path.  If the file geodatabase
        does not exist it will be created too
        
        @param fds_to_create: the complete path where the FDS will be created
        @type fds_to_create: string
        '''
        gdb, feature_dataset = os.path.split(fds_to_create)
        self.create_gdb_if_needed(gdb) #@UndefinedVariable
        if not arcpy.Exists(fds_to_create):
            # tempEnvironment0 = arcpy.env.XYResolution
            # arcpy.env.XYResolution = "0.1 Meters"
            # tempEnvironment1 = arcpy.env.XYTolerance
            # arcpy.env.XYTolerance = "0.1 Meters"
            arcpy.CreateFeatureDataset_management(gdb, feature_dataset, "PROJCS['NAD_1983_BC_Environment_Albers',GEOGCS['GCS_North_American_1983',DATUM['D_North_American_1983',SPHEROID['GRS_1980',6378137.0,298.2572221]],PRIMEM['Greenwich',0.0],UNIT['Degree',0.0174532925199433]],PROJECTION['Albers'],PARAMETER['False_Easting',1000000.0],PARAMETER['False_Northing',0.0],PARAMETER['Central_Meridian',-126.0],PARAMETER['Standard_Parallel_1',50.0],PARAMETER['Standard_Parallel_2',58.5],PARAMETER['Latitude_Of_Origin',45.0],UNIT['Meter',1.0]];-13239300 -8610100 10000;-100000 10000;-100000 10000;0.001;0.001;0.001;IsHighPrecision")
            # arcpy.env.XYResolution = tempEnvironment0
            # arcpy.env.XYTolerance = tempEnvironment1
    
    def create_gdb_if_needed(self, gdb_to_create):
        '''
        This creates a geodatabase in the provided path
        
        @param gdb_to_create: the complete path where the GDB will be created
        @type gdb_to_create: string
        '''
        if not arcpy.Exists(gdb_to_create):
            folder, file = os.path.split(gdb_to_create)
            #if not os.path.exists(folder):
            #    os.makedirs(folder)
            arcpy.CreateFileGDB_management(folder, file) 


    def get_inactives(self):
        '''
        retrieves a list of inactive tenures to be used in reporting
        '''
        
        print ('Retrieving the parcels list')
        parcel_fc = os.path.join(self.sde_connection, r'WHSE_TANTALIS.TA_INTEREST_PARCEL_SHAPES')
        clip_parcel = arcpy.Clip_analysis(parcel_fc, self.analyize_this_featureclass, r"memory\parcel_clip")
        result = int(arcpy.GetCount_management(clip_parcel).getOutput(0))
        print('{} has {} records'.format("Tantalis Parcels", result))
        if result > 0:
            parcel_list = [row[0] for row in arcpy.da.SearchCursor(clip_parcel,['INTRID_SID'])]
            print(len(parcel_list))

            #get credentials from keyring
            key_name = config.CONNNAME
            try:
                credentials = keyring.get_credential(key_name, "")
                username = credentials.username
                password = credentials.password
            except Exception as e:
                print(e)
                arcpy.AddWarning("Unable to generate TAB2: Credentials not available in keyring.")
                return
            #arcpy.AddMessage(f"username: {username} password: {password}")
            #pass credentials to get Oracle driver and then retrieve the list of inactive crown tenures.

            oracle_driver = inactives.get_oracle_driver()
            if oracle_driver:
                inactive_list = inactives.execute_process(parcel_list,username,password,oracle_driver)
                self.interest_status = inactive_list['interest_status']
                self.interest_type = inactive_list['interest_type']
                self.dpr_registry_name = inactive_list['dpr_registry_name']
                self.business_identifier = inactive_list['business_identifier']
                self.responsible_agency = inactive_list['responsible_agency']
                self.summary_holders_ilrr_identifier = inactive_list['summary_holders_ilrr_identifier']

                if self.business_identifier:
                    return True
                else:
                    return False
            else:
                arcpy.AddMessage("Oracle driver could not be returned!")
                return

        else:
            arcpy.AddMessage("No interest parcels returned!")
            return False
        

    def warn_csv_files(self):
        """
        Check if any file in the list `files` exists within the specified `folder_path`.
        
        If any file exists, print out a warning. If none exists, do nothing.

        Args:
            folder_path (str): The path to the folder.
            files (list): A list of file names to check for existence.
        """

        folder_path = self.work_directory
        csv_list = ['ilrrbusinesskeys', 'ilrrinterestholders', 'summary report', 'ilrrinterests', 'ilrrlocations', 'ilrrbusinesskeys']

        # Check if any CSV file exists in the folder
        files_exist = any(os.path.exists(os.path.join(folder_path, f"{file_name}.csv")) for file_name in csv_list)

        # If any file exists, print a warning
        if files_exist:
            arcpy.AddWarning("ILRR CSV files are no longer used to generate TAB2. TAB2 only contains inactive crown land tenure information.")


    
    def read_csv_files(self):
        '''
        If the ILRR CSV files have been created, they are read into a series of lists.
        These lists are written to various fields on the spreadsheets second worksheet.
        '''
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("Reading CSV files")
        #arcpy.AddWarning("In read_csv_files()")
        
        self.business_identifier = []
        self.ilrr_interest_identifier = ""
        #self.ilrr_interest_person_name = ""
        all_csv_files_found = 'yes'
    
        if not os.path.exists(os.path.join(self.work_directory,'ilrrbusinesskeys.csv')):
            all_csv_files_found = 'no'
            arcpy.AddMessage (" can't find ilrrbusinesskeys.csv " )
            
        if not os.path.exists(os.path.join(self.work_directory,'ilrrinterestholders.csv')):
            all_csv_files_found = 'no'
            arcpy.AddMessage (" can't find ilrrinterestholders.csv " )
            
        if not os.path.exists(os.path.join(self.work_directory,'summary report.csv')):
            all_csv_files_found = 'no'
            arcpy.AddMessage (" can't find summary report.csv " )
            
        if not os.path.exists(os.path.join(self.work_directory,'ilrrinterests.csv')):
            all_csv_files_found = 'no'
            arcpy.AddMessage (" can't find ilrrinterests.csv " )
    
        if not os.path.exists(os.path.join(self.work_directory,'ilrrlocations.csv')):
            all_csv_files_found = 'no'
            arcpy.AddMessage (" can't find ilrrlocations.csv " )
        
        if all_csv_files_found == 'yes':
            this_csv_file = os.path.join(self.work_directory,'ilrrbusinesskeys.csv')
            f = open(this_csv_file , 'r')
            #print("f" , f)
            for row in f:
                array = row.split(',')
                first_item = array[0] # ilrr interest number
                third_item = array[2]
                third_item = third_item.replace('"', '').strip()
                try:
                    third_item = int(third_item) # disposition number
                    if int(third_item) == int(self.disposition_transaction):
                        self.ilrr_interest_identifier = first_item
                except:
                    print("error in the 'read_csv_files' method")
                    
    
            this_csv_file = os.path.join(self.work_directory,'ilrrinterestholders.csv')
            #f = open(this_csv_file , 'r')
            f = csv.reader(open(this_csv_file, 'r'), delimiter=',', quotechar='"')
            self.interest_holders_ilrr_identifier = []
            for row in f:
                new_row = ""
                for field in row:
                    #print("field " , field)
                    field = field.replace(',', ' ')#.strip()
                    new_row = new_row + field + ","
                self.interest_holders_ilrr_identifier.append(new_row)
                #array = new_row.split(',')
                #first_item = array[0] # ilrr interest number
                #second_item = array[1] # interest persons name
                #third_item = array[2] # interest persons name
                #print("first_item   ",first_item)
                #print("   second_item   ",second_item)
                #print("   third_item   ",third_item)
                #print("   self.ilrr_interest_identifier  " , self.ilrr_interest_identifier)
                
                
                #if first_item == self.ilrr_interest_identifier :
                #    self.ilrr_interest_person_name = second_item 
    
    
             
            this_csv_file = os.path.join(self.work_directory,'ilrrlocations.csv')
            location_legal_description = []
            count = 1

            with open(this_csv_file, 'r') as f:
                reader = csv.reader(f)
                for row in reader:
                    if count > 1:
                        try:
                            if row[0].isdigit():
                                location_legal_description.append(row[6])
                            else:
                                arcpy.AddWarning(f"ILRR ID is not a number - row {count}")
                                continue

                        except IndexError:
                            arcpy.AddWarning(f"ILRR ID is 'None' Value - row {count}")
                            continue

                        finally:
                            count +=1
    
            this_csv_file = os.path.join(self.work_directory,'ilrrinterests.csv')
            self.interest_status = []
            self.interest_type = []
            self.dpr_registry_name = []
            with open(this_csv_file, 'r') as f:
                reader = csv.reader(f)
                for row in reader:
                    self.interest_status.append(row[4])
                    self.interest_type.append(row[5])
                    self.dpr_registry_name.append(row[6])
    
    
            this_csv_file = os.path.join(self.work_directory,'summary report.csv')
            self.business_identifier = []
            self.responsible_agency = []
            self.summary_holders_ilrr_identifier = []
            with open(this_csv_file, 'r') as f:
                reader = csv.reader(f)
                for row in reader:
                    self.business_identifier.append(row[2])
                    self.responsible_agency.append(row[5])
                    self.summary_holders_ilrr_identifier.append(row[0])
        else:
            arcpy.AddWarning("Tab 2 NOT created")
    
    def get_all_needed_datasets(self):
        def update_fields(input_layer, distField, newName, source_fc, source_path):
            #updates the reporting fields with the appropriate values
            arcpy.AddMessage(f".     Calculating necessary fields")
            arcpy.AddMessage(".      Adding nears fields to parcel dataset")

            #create the dictionary that cotnains the informatuon for adding the fields
            updateFld = {distField:"DOUBLE", newName:"TEXT"}
            #Iterate through the dictionary and create the fields
            for key, value in updateFld.items():
                try:
                    arcpy.AddField_management(input_layer, key, value)
                except Exception:
                    continue
            
            #List the fields and set up the expressions based on the available fields from the input feature class
            try:
                #test if source exists
                arcpy.MakeFeatureLayer_management(source_path, "check_source")
                flds = []
                field_list = arcpy.ListFields(source_path)
                for field in field_list:
                    flds.append(field.name)
                if 'BORDENNUMBER' in flds:
                    name_expr = f"!{source_fc}.BORDENNUMBER!"
                elif 'ENGLISH_NAME' in flds and 'BAND_NAME' in flds:
                    name_expr = f"!{source_fc}.ENGLISH_NAME! + ' - ' + !{source_fc}.BAND_NAME!"
                
                #set the expression needed for updating the distance fields
                dist_expr = "round(!NEAR_DIST!, 1)"
            except:
                #set the values if the dataset is not available
                arcpy.AddMessage(f".  Source dataset is not valid: {source_fc}")
                dist_expr = "-9999"
                name_expr = """'FAILED'"""
            
            #Calculate the distance field
            arcpy.CalculateField_management(input_layer, 
                                    field=distField,
                                    expression=dist_expr, 
                                    expression_type="PYTHON3")
            
            #calculate the name field
            arcpy.CalculateField_management(input_layer, 
                                    field=newName, 
                                    expression=name_expr, 
                                    expression_type="PYTHON3")

            return


        def get_first_nations_data():
            #set the work environment
            arcpy.env.workspace = self.work_gdb

            #create dictionary of first nations data for reporting on tab 1
            dict = {
                "bc_arch_potential":["WHSE_ARCHAEOLOGY.RAAD_AOA_PROVINCIAL", "aoi", "", True, "", "", ""],
                "bc_arch_sites":["WHSE_ARCHAEOLOGY.RAAD_TFM_SITES_SVW", None, "25 kilometers", False, "nearest_arch", "nearest_arch_name"],
                "bc_indian_res":["WHSE_ADMIN_BOUNDARIES.ADM_INDIAN_RESERVES_BANDS_SP", None, "", False, "nearest_ir", "nearest_ir_name"]
                }
            
            #set the datasets
            aoi = os.path.join(self.work_gdb , "aoi")
            parcel = os.path.join(self.work_gdb , "parcel_boundary_for_clips")

            #check if the dataset for storing the FN information exists. If not, create the datasets
            #and continue the analysis
            if not arcpy.Exists(parcel):
                arcpy.CopyFeatures_management(aoi, parcel)
                arcpy.MakeFeatureLayer_management(parcel , "output_layer")

                #iterate through the FN dictionary to create reporting objects
                for key, inputList in dict.items():
                    output = os.path.join("input_of_raw_data", key)

                    source_fc = inputList[0] #data source feature class
                    clip_fc = inputList[1] #feature class that will be used to clip the input data, if required
                    radius = inputList[2] #radius for near analysis
                    clip_only = inputList[3] #boolean to indicate if only a clip is required
                    distField = inputList[4] #new field to capture and report distance to nearest features
                    newName = inputList[5] #new name field of reporting nearest features

                    source_path = os.path.join(self.sde_connection, source_fc)

                    #clip the data if clip_only variable is True
                    #if it does work, continue through the next dataset
                    arcpy.AddMessage(f".  Getting {output}")
                    if clip_only is True:
                        try:
                            arcpy.Clip_analysis(source_path, clip_fc, output)
                        except:
                            arcpy.AddWarning(f".  could not clip {output}")
                            
                    else:
                        try:
                            #create the near tables needed for reporting distances to features
                            arcpy.MakeFeatureLayer_management(source_path , "source_layer")
                            arcpy.AddMessage(f".     Calculating nearest {output}")
                            arcpy.Near_analysis("output_layer", "source_layer", search_radius=radius, distance_unit='Kilometers')
                            arcpy.AddMessage(f".     performing join")
                            arcpy.AddJoin_management("output_layer", 
                                                    "NEAR_FID", 
                                                    "source_layer",
                                                    "OBJECTID", 
                                                    "KEEP_ALL")
                        except:
                            #if the above fails, simply run the update_fields function
                            #Errors will be caught in that function for reporting
                            arcpy.AddWarning(f".  could not create distances to {output}")
                            update_fields("output_layer", distField, newName, source_fc, source_path)
                            continue
                        else:
                            #update the reporting fields and remove the join and delete unneeded fields
                            update_fields("output_layer", distField, newName, source_fc, source_path)
                            arcpy.RemoveJoin_management("output_layer")
                            arcpy.DeleteField_management(parcel, ['NEAR_FID', 'NEAR_DIST'])
            else:
                arcpy.AddMessage(".  * Skipped adding and calculating IR and ARCH fields ")


        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("Buffering parcel boundary " )

        the_old_output = os.path.join(self.work_gdb , "aoi")
        the_output = os.path.join(self.work_gdb , "aoi_for_adjacent")
        
        if not arcpy.Exists(the_output):
            arcpy.AddMessage(".  5m" )
            arcpy.Buffer_analysis(the_old_output, the_output, 5, "OUTSIDE_ONLY")

        arcpy.AddMessage(".  25km" )
        the_old_output = os.path.join(self.work_gdb , "aoi")
        the_output = os.path.join(self.work_gdb , "aoi_25km")
        if not arcpy.Exists(the_output):
            arcpy.Buffer_analysis(the_old_output, the_output, 25000, "","","ALL")
        
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("Gathering data needed to run tool" )
        arcpy.AddMessage(".  Parcel centroid")
        the_input  = os.path.join(self.work_gdb , "aoi_25km")
        the_output = os.path.join(self.work_gdb , "aoi_centroid")
        if not arcpy.Exists(the_output):
            arcpy.FeatureToPoint_management(the_input, the_output, "INSIDE")

        disposition_transaction= '#'
        if disposition_transaction == '#' or disposition_transaction == "":
            disposition_transaction = "0" 

        the_input = os.path.join(self.sde_connection , "WHSE_TANTALIS.TA_DISPOSITION_TRANSACTIONS")
        the_output = os.path.join(self.work_gdb , "ta_disposition_transactions")
        
        arcpy.AddMessage(".  TA_DISPOSITION_TRANSACTIONS ")
        select_string = "DISPOSITION_TRANSACTION_SID  = " + str(int(disposition_transaction))
        if not arcpy.Exists(the_output):
            arcpy.MakeQueryTable_management(the_input, "table_layer", "USE_KEY_FIELDS", "", "", select_string)
            arcpy.CopyRows_management("table_layer", the_output, "")

        get_first_nations_data()

        poly_dict = {"parcel_mapsheet_polys": ["aoi", "WHSE_BASEMAPPING.BCGS_20K_GRID", False],
                    "parcel_for_district": ["aoi", "WHSE_ADMIN_BOUNDARIES.ADM_NR_DISTRICTS_SP", False],
                    "parcel_municipality": ["aoi", "WHSE_LEGAL_ADMIN_BOUNDARIES.ABMS_MUNICIPALITIES_SP", False],
                    "parcel_prov_forest": ["aoi", "WHSE_ADMIN_BOUNDARIES.FADM_PROV_FOREST", False],
                    "parcel_prov_forest_addition": ["aoi", "WHSE_ADMIN_BOUNDARIES.FADM_PROV_FOREST_ADDITION", False],
                    "parcel_prov_forest_deletion": ["aoi", "WHSE_ADMIN_BOUNDARIES.FADM_PROV_FOREST_DELETION", False],
                    "parcel_prov_forest_exclusion": ["aoi", "WHSE_ADMIN_BOUNDARIES.FADM_PROV_FOREST_EXCLUSION", False],
                    "parcel_assessment_area": ["aoi", "WHSE_TANTALIS.TA_ASSESSMENT_AREAS_SVW", False],
                    "parcel_land_title_district": ["aoi", "WHSE_TANTALIS.TA_LAND_TITLE_DISTRICTS_SVW", False],
                    "parcel_electoral_district": ["aoi", "WHSE_ADMIN_BOUNDARIES.PED_PROV_ELECTORAL_DIST_POLY", False],
                    "parcel_regional_district": ["aoi", "WHSE_LEGAL_ADMIN_BOUNDARIES.ABMS_REGIONAL_DISTRICTS_SP", False],
                    "parcel_pins": ["aoi", "WHSE_TANTALIS.TA_SURVEY_PARCELS_SVW", True],
                    "parcel_pins_adjacent": ["aoi_for_adjacent", "WHSE_TANTALIS.TA_SURVEY_PARCELS_SVW", True],
                    "parcel_pins_right_of_way": ["aoi", "WHSE_TANTALIS.TA_SURVEYED_ROW_PARCELS_SVW", True],
                    "parcel_alr": ["aoi", "WHSE_LEGAL_ADMIN_BOUNDARIES.OATS_ALR_POLYS", True]
                    }

        for name, inputs in poly_dict.items():
            rename = name.replace("_", " ")
            arcpy.AddMessage(f".  Getting {rename}")
            the_output = os.path.join (self.work_gdb , name)
            the_input  = os.path.join (self.work_gdb , inputs[0])
            clip_identity = os.path.join(self.sde_connection , inputs[1]) 
            if not arcpy.Exists(the_output):
                if inputs[2] == True:
                    arcpy.AddMessage(f".  Clipping for {name}")
                    arcpy.Clip_analysis(clip_identity, the_input, the_output)
                else:
                    arcpy.AddMessage(f".  Identitying for {name}")
                    arcpy.Identity_analysis(the_input, clip_identity, the_output)
    

    def get_values_from_field(self, fc_name, field_name, field_name2=''):
        '''
        Returns a unique list of all the values is a specific field (or two fields)
        in the specified featureclass. 
        '''

        field_values_list = []

        if arcpy.Exists(fc_name):
            fieldList = arcpy.ListFields(fc_name)
            for field in fieldList:
                fld_name = field.name
                if fld_name == field_name:
                    all_rows = arcpy.SearchCursor(fc_name)
                    for row in all_rows:
                        field1_value = (str(row.getValue(field_name)))
                        if field1_value[-2:] == '.0': # trim of .0 from strings
                            field1_value = field1_value[:-2]
                        #print(mark[:-2])
                        #print(mark[-2:])
                        if field_name2 != '':
                            field2_value = (str(row.getValue(field_name2)))
                            field1_value += " - " + field2_value
                        field_value_bytes = field1_value.encode("utf-8")
                        field_values_list.append(field_value_bytes)
        field_values_list = list(set(field_values_list))  # makes the list unique
        field_values_list.sort() # sort the list
        
        return  field_values_list 
    
    def get_values_from_field2(self, fc_name, field_name, field_name2=''):
        '''
        Returns a unique list of all the values is a specific field (or two fields)
        in the specified featureclass. 
        '''
        
        fields = [field_name, field_name2]

        field_values_list = []

        if arcpy.Exists(fc_name):
            with arcpy.da.SearchCursor(fc_name, fields) as cursor: #@UndefinedVariable    
                for row in cursor:
                    field1_value = str(row[0]).encode("utf-8")
                    if field1_value[-2:] == '.0': # trim of .0 from strings
                        field1_value = field1_value[:-2]
                    if field_name2 != '':
                            field2_value = str(row[1]).encode("utf-8")
                            field1_value += " - " + field2_value
                    field_values_list.append(field1_value)
        field_values_list = list(set(field_values_list))  # makes the list unique
        field_values_list.sort() # sort the list
        
        return  field_values_list

    
    def create_excel_cell_variables(self):
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("Creating the excel variables")

        #disposition_number = '902281'
        #disposition_number = int(self.disposition_transaction)
        #print("disposition_number " , disposition_number)
        
        applicant_table = os.path.join(self.sde_connection,"WHSE_TANTALIS.TA_INTEREST_HOLDER_VW")
        #print("applicant_table " , applicant_table)
        
        self.xl_applicant_name = ""
        if self.disposition_transaction != '' and self.disposition_transaction != '#':
            
            sel_string = "DISPOSITION_TRANSACTION_SID = " + self.disposition_transaction
            
            
            #arcpy.MakeTableView_management(applicant_table, "applicant_view", sel_string)
            arcpy.MakeQueryTable_management(applicant_table, "applicant_view", "USE_KEY_FIELDS", "", "", sel_string)

            
            #arcpy.MakeTableView_management(applicant_table, "applicant_view", "DISPOSITION_TRANSACTION_SID = 902281")
            #self.disposition_transaction
    
            if int(str(arcpy.GetCount_management("applicant_view"))) > 0:
                print("   its greater than zero")
                self.xl_applicant_name = ""
                all_rows = arcpy.SearchCursor("applicant_view")
                for row in all_rows:
                    field1_value = (str(row.getValue("ORGANIZATIONS_LEGAL_NAME")))
                    field2_value = (str(row.getValue("INDIVIDUALS_FIRST_NAME")))
                    field3_value = (str(row.getValue("INDIVIDUALS_LAST_NAME")))
                    if field1_value != 'None':
                        self.xl_applicant_name += field1_value
                        self.xl_applicant_name += " "
                    if field2_value != 'None':
                        self.xl_applicant_name += field2_value
                        self.xl_applicant_name += " "
                    if field3_value != 'None':
                        self.xl_applicant_name += field3_value
            #------------------------------------------------------------------------------ 

        fc_name = os.path.join(self.work_gdb, "ta_disposition_transactions")
        self.xl_received_date = self.get_values_from_field(fc_name,"RECEIVED_DAT")
        print("   self.xl_received_date " , self.xl_received_date)

        if self.disposition_transaction == 0:
            self.xl_received_date = ""
        
        print("   self.xl_received_date " , self.xl_received_date)

        fc_name = os.path.join(self.work_gdb, "parcel_regional_district")
        self.xl_regional_district = self.get_values_from_field(fc_name,"REGIONAL_DISTRICT_NAME")
        print("   self.xl_regional_district " , self.xl_regional_district)

        fc_name = os.path.join(self.work_gdb, "parcel_electoral_district")
        self.xl_electoral_district = self.get_values_from_field(fc_name,"ELECTORAL_DISTRICT_NAME")
        print("   self.xl_electoral_district " , self.xl_electoral_district)

        fc_name = os.path.join(self.work_gdb, "parcel_land_title_district")
        self.xl_land_title = self.get_values_from_field(fc_name,"LAND_TITLE_DISTRICT_NAME")
        print("   self.xl_land_title " , self.xl_land_title)

        fc_name = os.path.join(self.work_gdb, "input_of_raw_data\\bc_arch_potential")
        if arcpy.Exists(fc_name):
            self.xl_arch_potential = self.get_values_from_field(fc_name,"POTENTIAL")
        else:
            self.xl_arch_potential = 'FAILED'
        print("   self.xl_arch_potential " , self.xl_arch_potential)

        fc_name = os.path.join(self.work_gdb, "parcel_municipality")
        self.xl_municipality = self.get_values_from_field(fc_name,"MUNICIPALITY_NAME")
        print("   self.xl_municipality " , self.xl_municipality)

        fc_name = os.path.join(self.work_gdb, "parcel_assessment_area")
        self.xl_parcel_assessment_area = self.get_values_from_field(fc_name,"ASSESSMENT_AREA_NAME")
        print("   self.xl_parcel_assessment_area " , self.xl_parcel_assessment_area)

        fc_name = os.path.join(self.work_gdb, "parcel_prov_forest")
        self.xl_provincial_forest = self.get_values_from_field(fc_name,"PROV_FOREST_CD_DESCRIPTION","MAP_BLOCK_ID")
        print("   self.xl_provincial_forest " , self.xl_provincial_forest)

        fc_name = os.path.join(self.work_gdb, "parcel_prov_forest_addition")
        self.xl_provincial_forest_addition = self.get_values_from_field(fc_name,"PROV_FOREST_CD_DESCRIPTION","OIC_YEAR")
        print("   self.parcel_prov_forest_addition " , self.xl_provincial_forest_addition)

        fc_name = os.path.join(self.work_gdb, "parcel_prov_forest_addition")
        self.xl_provincial_forest_addition2 = self.get_values_from_field(fc_name,"OIC_NUMBER","MAP_BLOCK_ID")
        print("   self.parcel_prov_forest_addition2 " , self.xl_provincial_forest_addition2)

        fc_name = os.path.join(self.work_gdb, "parcel_prov_forest_deletion")
        self.xl_provincial_forest_deletion = self.get_values_from_field(fc_name,"PROV_FOREST_CD_DESCRIPTION","DOCUMENT_TYPE")
        print("   self.parcel_prov_forest_deletion " , self.xl_provincial_forest_deletion)

        fc_name = os.path.join(self.work_gdb, "parcel_prov_forest_deletion")
        self.xl_provincial_forest_deletion2 = self.get_values_from_field(fc_name,"DELETION_YEAR")
        print("   self.parcel_prov_forest_deletion2 " , self.xl_provincial_forest_deletion2)

        fc_name = os.path.join(self.work_gdb, "parcel_prov_forest_exclusion")
        self.xl_provincial_forest_exclusion = self.get_values_from_field(fc_name,"PROV_FOREST_CD_DESCRIPTION","AREA_TYPE")
        print("   self.parcel_prov_forest_exclusion " , self.xl_provincial_forest_exclusion)

        fc_name = os.path.join(self.work_gdb, "parcel_prov_forest_exclusion")
        self.xl_provincial_forest_exclusion2 = self.get_values_from_field(fc_name,"EXCLUSION_NUMBER")
        print("   self.parcel_prov_forest_exclusion2 " , self.xl_provincial_forest_exclusion2)

        fc_name = os.path.join(self.work_gdb, "parcel_prov_forest")
        self.xl_map_block_id = self.get_values_from_field(fc_name,"MAP_BLOCK_ID")
        print("   self.xl_map_block_id " , self.xl_map_block_id)

        #fc_name = os.path.join(self.work_gdb + "\\parcel_mapsheet")
        fc_name = os.path.join(self.work_gdb, "parcel_mapsheet_polys")
        self.xl_mapsheet = self.get_values_from_field(fc_name,"MAP_TILE_DISPLAY_NAME")
        print("   self.xl_mapsheet " , self.xl_mapsheet)
 
        fc_name = os.path.join(self.work_gdb, "parcel_for_district")
        self.xl_forest_district = self.get_values_from_field(fc_name,"DISTRICT_NAME")
        print("   self.xl_forest_district " , self.xl_forest_district)

        fc_name = os.path.join(self.work_gdb, "parcel_pins")
        self.xl_pins = self.get_values_from_field(fc_name,"PIN_SID","PARCEL_LEGAL_DESCRIPTION")
        print("   self.xl_pins " , self.xl_pins)

        fc_name = os.path.join(self.work_gdb, "parcel_pins_adjacent")
        self.xl_pins_adjacent = self.get_values_from_field(fc_name,"PIN_SID","PARCEL_LEGAL_DESCRIPTION")
        print("   self.xl_pins_adjacent " , self.xl_pins_adjacent)

        fc_name = os.path.join(self.work_gdb, "parcel_pins_right_of_way")
        self.xl_pins_right_of_way = self.get_values_from_field(fc_name,"PIN_SID","PARCEL_LEGAL_DESCRIPTION")
        print("   self.xl_pins_right_of_way " , self.xl_pins_right_of_way)

        #------------------------------------------------------------

        fc_name = os.path.join(self.work_gdb, "parcel_boundary_for_clips")

        self.xl_tenure_area_ha = self.get_values_from_field(fc_name,"TENURE_AREA_IN_HECTARES")
        print("   self.xl_tenure_area_ha " , self.xl_tenure_area_ha)

        self.xl_tenure_type = self.get_values_from_field(fc_name,"TENURE_TYPE")
        print("   self.xl_tenure_type " , self.xl_tenure_type)
        
        self.xl_intrid_sid = self.get_values_from_field(fc_name,"INTRID_SID")
        print("   self.xl_intrid_sid " , self.xl_intrid_sid)
        
        self.xl_tenure_subtype = self.get_values_from_field(fc_name,"TENURE_SUBTYPE")
        print("   self.xl_tenure_subtype " , self.xl_tenure_subtype)
        
        self.xl_tenure_purpose = self.get_values_from_field(fc_name,"TENURE_PURPOSE")
        print("   self.xl_tenure_purpose " , self.xl_tenure_purpose)
        
        self.xl_tenure_subpurpose = self.get_values_from_field(fc_name,"TENURE_SUBPURPOSE")
        print("   self.xl_tenure_subpurpose " , self.xl_tenure_subpurpose)
        
        self.xl_tenure_location = self.get_values_from_field(fc_name,"TENURE_LOCATION")
        print("   self.xl_tenure_location " , self.xl_tenure_location)

        self.xl_tenure_legal_description = self.get_values_from_field(fc_name,"TENURE_LEGAL_DESCRIPTION")
        print("   self.xl_tenure_legal_description " , self.xl_tenure_legal_description)
        
        #self.xl_tenure_area_ha = self.get_values_from_field(fc_name,"TENURE_AREA_IN_HECTARES")
        #print("   self.xl_tenure_area_ha " , self.xl_tenure_area_ha)
        
        
        #________________________________________________________________________________
        
        #self.xl_nearest_ir = self.get_values_from_field(fc_name,"nearest_ir","nearest_ir_name")
        self.xl_nearest_ir = self.get_values_from_field(fc_name,"nearest_ir")
        arcpy.AddMessage("   self.xl_nearest_ir:")
        for ir in self.xl_nearest_ir:
            arcpy.AddMessage("      {}, {}".format(ir, ir.decode('utf-8')))
        #print("   self.xl_nearest_ir " , self.xl_nearest_ir, self.xl_nearest_ir.decode('utf-8'))
        
        #self.xl_nearest_arch = self.get_values_from_field(fc_name,"nearest_arch","nearest_arch_name")
        self.xl_nearest_arch = self.get_values_from_field(fc_name,"nearest_arch")
        print("   self.xl_nearest_arch " , self.xl_nearest_arch)
        
        self.xl_nearest_ir_name = self.get_values_from_field(fc_name,"nearest_ir_name")
        for ir in self.xl_nearest_ir_name:
            arcpy.AddMessage("      {}, {}".format(ir, ir.decode('utf-8')))
        #print("   self.xl_nearest_ir_name " , self.xl_nearest_ir_name, self.xl_nearest_ir_name.decode('utf-8'))

        self.xl_nearest_arch_name = self.get_values_from_field(fc_name,"nearest_arch_name")
        print("   self.xl_nearest_arch_name " , self.xl_nearest_arch_name)

        self.xl_in_alr = 'No'
        fc_name = os.path.join(self.work_gdb, "parcel_alr")
        if int(str(arcpy.GetCount_management(fc_name))) > 0:
            self.xl_in_alr = "Yes"
        print("   self.xl_in_alr " , self.xl_in_alr)
        

    def create_spreadsheet_openpyxl(self):
        '''
        Creates the spreadsheet, adds worksheets, names the worksheets, and sets up
        some excel constants for aligning the text in the output cells.
        '''

        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Creating the final output spreadsheet')
   
        #self.xls_to_save = os.path.join(self.work_directory,self.xls_for_analysis + "_" + self.featureclass_to_analyize + ".xlsx")
        #print('{:10}{:250}'.format('Saving: ', self.xls_to_save))
        self.xls_to_save = self.xls_name_to_save_as

        # Create the workbook
        self.workbook = openpyxl.Workbook()
        
        #worksheet_name_dict = {"Conflicts & Constraints": 0, "Status of Conflict": 1, "blank": 2}
        self.newline = 1

        disclaimer = "This product contains sensitive information.  INTERNAL GOVERNMENT USE ONLY"
        for name, index in {"Crown Land Status": 0, "Status of Conflict": 1, "DEPRECATED Status of Conflict": 2, "blank": 3}.items():
            # Create Worksheet
            sheet = self.workbook.create_sheet(name, index)
            
            # Prepare and write header cell
            if sheet.title != "blank":
                for column, width in {"A": 30, 
                                      "B": 50, 
                                      "C": 25}.items():
                    sheet.column_dimensions[column].width = width
                sheet.merge_cells('A1:C1')
                # Header
                self.newcolumn =  1
                
                sheet.cell(self.newline, self.newcolumn, disclaimer).font =  Font(color='ff0000',size=12,name='Arial')
        self.newline += 1
        
        # save XLSX
        self.workbook.save(self.xls_to_save)
                
        # Worksheet variables
        self.tab1_sheet = self.workbook.get_sheet_by_name("Crown Land Status")
        self.tab2_sheet = self.workbook.get_sheet_by_name("Status of Conflict")
        self.tab3_sheet = self.workbook.get_sheet_by_name("DEPRECATED Status of Conflict")

       
    def apply_grey_fill(self, start_row, end_row, step=2):
        '''
        Applies grey fill to a range of cells
        '''
        
        if start_row == end_row:
            for column in range(1, 3):
                self.this_sheet.cell(start_row, column).fill = PatternFill("solid", fgColor="e6e6e6")
        elif start_row < end_row:
            for row in range(start_row+2, end_row+1, step):
                #print("   {}".format(row))
                for column in range(1, 3):
                    self.this_sheet.cell(row, column).fill = PatternFill("solid", fgColor="e6e6e6")
    
                
    def apply_border2(self, ws, start_row, end_row, start_column, end_column):
        '''
        Applies a thick black border to a range of cells
        '''
        
        cell_range = start_column + str(start_row+1) + ":" + end_column + str(end_row)
        rows = ws[cell_range]
        for row in rows:
            if row == rows[0][0] or row == rows[0][-1] or row == rows[-1][0] or row == rows[-1][-1]:
                pass
            else:
                row[0].border = Border(left=Side(style='thick'))
                row[-1].border = Border(right=Side(style='thick'))
            for c in rows[0]:
                c.border = Border(top=Side(style='thick'))
            for c in rows[-1]:
                c.border = Border(bottom=Side(style='thick'))
        rows[0][0].border = Border(left=Side(style='thick'), top=Side(style='thick'))
        rows[0][-1].border = Border(right=Side(style='thick'), top=Side(style='thick'))
        rows[-1][0].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
        rows[-1][-1].border = Border(right=Side(style='thick'), bottom=Side(style='thick'))

    
    def create_xls_files_section(self):
        '''
        Creates the FILES section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating FILES section of the spreadsheet")

        style1 = NamedStyle(name="style1")
        style1.font = Font(size=11)
        style1.alignment = Alignment(wrapText=True)

        style2 = NamedStyle(name="style2")
        style2.font = Font(color="31869b", size=9)
        style2.alignment = Alignment(wrapText=True)

        self.this_sheet = self.tab1_sheet
        self.newline = 2
        self.newcolumn = 1
        
        self.this_sheet.merge_cells('A2:B2')
        self.this_sheet.cell(self.newline, self.newcolumn, "Status Summary").font =  Font(color='31869b',size=14,name='Arial')
        self.this_sheet.cell(self.newline, self.newcolumn).alignment = Alignment(horizontal="center")
        self.newline += 1
        
        # File
        start_row = self.newline
        self.this_sheet.cell(self.newline, self.newcolumn, "Files").style = style1
        self.this_sheet.cell(self.newline, self.newcolumn+1, "" + self.crown_lands_file).style = style2
        self.newline += 1
        
        for line_title in ["Application Agency -> Mines",
                           "Application Agency -> Lands",
                           "Application Agency -> Water",
                           "Application Agency -> Forests",
                           "Application Received Date",
                           "Completed by NRS / Date",
                           "Status Run By / Date"]:
            self.this_sheet.cell(self.newline, self.newcolumn, line_title).style = style1
            if line_title == "Status Run By / Date":
                run_string = os.environ['USERNAME'] +"    -    " + str(datetime.date.today())
                self.this_sheet.cell(self.newline, self.newcolumn+1, run_string).style = style2
            self.newline += 1
        
        
        purpose_string = "Purpose: The purpose of this status summary is to present the " \
            "information gathered by each person involved in the status phase to form the complete " \
            "status package. The status summary sheet should be used in conjunction with the ILRR " \
            "Custom Theme Report to provide a full status package. The information on the summary " \
            "is meant to benefit the business leads, decision makers and others who are preparing " \
            "the tenure documents and decision summary to streamline the process by avoiding " \
            "duplicate parties researching information at multiple stages in the process."
        
        self.this_sheet.merge_cells(start_row=self.newline, 
                                    start_column=self.newcolumn, 
                                    end_row=self.newline, 
                                    end_column=self.newcolumn+1)
        self.this_sheet.cell(self.newline, self.newcolumn, purpose_string).style = style1
        #self.this_sheet.cell(self.newline, self.newcolumn).alignment = Alignment(wrap_text=True)
        #self.this_sheet.cell(self.newline, self.newcolumn).alignment = Alignment(wrapText=True)
        self.this_sheet.row_dimensions[self.newline].height = 115


        # apply grey fill        
        end_row = self.newline
        self.apply_grey_fill(start_row, end_row, 2)
        
        # apply border
        #self.apply_border(start_row, end_row)
        self.apply_border2(self.this_sheet, start_row, end_row, "A", "B")
        
        # Save worksheet
        self.workbook.save(self.xls_to_save)
        
 
    def create_xls_applicant_section2(self):
        '''
        Creates the applicant section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating Applicant's section of the spreadsheet")
 
        self.this_sheet = self.tab1_sheet
        self.newline += 3
        
        applicant_dict = {"Applicant": self.xl_applicant_name,
                          "Agent's Name": "",
                          "Project Lead": ""}
        for title_string, v in applicant_dict.items():
            self.newline +=1
            self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
            self.this_sheet.cell(self.newline, self.newcolumn+1, v).style = "style2"
            
            
        title_count = 0
        start_row = self.newline
                
        working_dict = {"Company Incorporation No.":    "",
                        "ATS Project Number":           "",
                        "Application Area (ha)":        self.xl_tenure_area_ha,
                        "Client Applied Area":          "",
                        "Dimensions":                   "",
                        "Amendments Existing Area":     "",
                        "Amendments New Area":          "",
                        "Location":                     self.xl_tenure_location,
                        "Map Sheet":                    self.xl_mapsheet,
                        "Shapefile Source":             "",
                        "Geomark":                      "",
                        "Legal Description":            self.xl_tenure_legal_description,
                        "PIN Intersects":               self.xl_pins,
                        "PIN Adjacent":                 self.xl_pins_adjacent,
                        "PIN Right of Ways":            self.xl_pins_right_of_way,
                        "PIN Reverted or Acquired":     "",
                        "PID Number (if titled)":       ""
                        }
        
        for title_string, v in working_dict.items():
            self.newline += 1
            # For Alternating grey fill
            title_count += 1
            line_numbers_written = [self.newline, self.newline]

            # Catch values that need further processing and formating
            # Area in RAT
            if title_string == "Application Area (ha)":
                if v != []:
                    v = "{} hectares".format(v[0].decode('utf-8'))
                    
            # Var is string
            if isinstance(v, str):
                # Write column 1
                self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                # Write value to column 2
                self.this_sheet.cell(self.newline, self.newcolumn+1, v).style = "style2"
                # Wrap Text
                #self.this_sheet.cell(self.newline, self.newcolumn+1).alignment = Alignment(wrap_text=True)
                #self.this_sheet.cell(self.newline, self.newcolumn+1).alignment = Alignment(wrapText=True)
                # Adjust row height, if necessary
                #if len(v) > 60:
                #    self.this_sheet(self.newline).height = 30
                line_numbers_written = [self.newline, self.newline]
            # Var is list
            elif isinstance(v, list):
                if v != []:
                    start_line_2 = self.newline
                    # Write Column 1
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    # Loop through list, enter value in column 2
                    for x in v:
                        # fist item in list goes on line with title
                        if x == v[0]:
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        # other items in list get their own line
                        else:
                            self.newline += 1
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        # Wrap text
                        #self.this_sheet.cell(self.newline, self.newcolumn+1).alignment = Alignment(wrap_text=True)
                        #self.this_sheet.cell(self.newline, self.newcolumn+1).alignment = Alignment(wrapText=True)
                        # Adjust row height, if necessary 
                        if len(v) > 60:
                        #    self.this_sheet(self.newline).height = 30
                            self.this_sheet.row_dimensions[self.newline].height = 30
                        end_line_2 = self.newline
                        line_numbers_written = [start_line_2, end_line_2]
                # Var is empty list
                elif v == []:
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    # Write value to cell
                    self.this_sheet.cell(self.newline, self.newcolumn+1, "").style = "style2"
                    # Wrap Text
                    #self.this_sheet.cell(self.newline, self.newcolumn+1).alignment = Alignment(wrap_text=True)
                    #self.this_sheet.cell(self.newline, self.newcolumn+1).alignment = Alignment(wrapText=True)
                    # Adjust row height, if necessary
                    if len(v) > 60:
                    #    self.this_sheet(self.newline).height = 30
                        self.this_sheet.row_dimensions[self.newline].height = 30
            end_row = self.newline
            
            #print(" # {} is count {}".format(title_string, title_count))
            # Apply grey fill
            if title_count%2 == 0:
                start_row_2, end_row_2 = line_numbers_written
                #print("*** {} *** {}:{}  ({})".format(title_string, start_row_2, end_row_2, self.newline))
                self.apply_grey_fill(start_row_2-2, end_row_2, 1)
                
        # Apply border
        self.apply_border2(self.this_sheet, start_row, end_row, "A", "B")
    
        # save Worksheet
        self.workbook.save(self.xls_to_save)


    def create_xls_lands_section2(self):
        '''
        Creates the LANDS section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating the LANDS section of the spreadsheet")

        self.this_sheet = self.tab1_sheet
        self.newline += 2
        
        title_string = "Lands note: this information may be applied to all agency application types"
        self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1" 
        
        title_count = 0
        start_row = self.newline
        
        working_dict = {"File Number":                  self.crown_lands_file,
                        "DID Number":                   self.disposition_transaction,
                        "DID Received Date":            "",
                        "IP Number":                    "",
                        "Parcel Number":                self.xl_intrid_sid,
                        "Region":                       self.region,
                        "Type":                         self.xl_tenure_type,
                        "Sub Type":                     self.xl_tenure_subtype,
                        "Purpose":                      self.xl_tenure_purpose,
                        "Sub Purpose":                  self.xl_tenure_subpurpose,
                        "Assessment Area":              self.xl_parcel_assessment_area,
                        "Regional District":            self.xl_regional_district,
                        "Municipality":                 self.xl_municipality,
                        "Electoral District":           self.xl_electoral_district,
                        "Provincial Forest":            self.xl_provincial_forest,
                        "Provincial Forest Addition":   self.xl_provincial_forest_addition,
                        "Provincial Forest Addition2":  self.xl_provincial_forest_addition2,
                        "Provincial Forest Deletion":   self.xl_provincial_forest_deletion,
                        "Provincial Forest Deletion2":  self.xl_provincial_forest_deletion2,
                        "Provincial Forest Exclusion":  self.xl_provincial_forest_exclusion,
                        "Provincial Forest Exclusion":  self.xl_provincial_forest_exclusion2,
                        "Forest District":              self.xl_forest_district,
                        "Land Title Office":            self.xl_land_title,
                        "In Agricultural Land Reserve": self.xl_in_alr,
                        "Nearest Indian Res":           "",
                        }
                        
        for title_string, v in working_dict.items():
            self.newline += 1
            # For Alternating grey fill
            title_count += 1
            line_numbers_written = [self.newline, self.newline]
            
            # Catch values that need further processing and formating
            if title_string == "DID Received Date":
                if self.disposition_transaction == 0:
                    v = ""
                if self.disposition_transaction != "" and self.disposition_transaction != '#' and self.disposition_transaction != '0':
                    v = self.xl_received_date
            # Catch IR Information
            elif title_string == "Nearest Indian Res":
                try:
                    x = self.xl_nearest_ir_name[0]
                    y = self.xl_nearest_ir[0]
                    #arcpy.AddMessage("- {}, {}".format(type(x), x))
                    x_decoded = x.decode('utf-8')
                    y_decoded = y.decode('utf-8')
                    v = "{} | {} km(s)".format(x_decoded, y_decoded)
                    v = ''.join([s for s in v if s in string.printable])
                except Exception as e:
                    v = "No Indian Reserve within 25km"

            #arcpy.AddMessage("Type: {}; \n\tValue:{}".format(type(v), v))
            
            # Write column 1 (if applicable)
            self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
            
            # Write column 2
            # string value
            if isinstance(v, str):
                self.this_sheet.cell(self.newline, self.newcolumn+1, v).style = "style2"
            # List value            
            elif isinstance(v, list):
                if v != []:
                    start_line_2 = self.newline
                    #self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    for x in v:
                        #if isinstance(x, bytes):
                        #    arcpy.AddMessage("decoding: {}".format(x))    
                            # fist item in list goes on line with title
                        #arcpy.AddMessage("l- {}, {}".format(type(x), x))
                        x_decoded = x.decode('utf-8')
                        if x == v[0]:
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x_decoded).style = "style2"  #@UndefinedVariable 
                        else:
                            self.newline += 1
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x_decoded).style = "style2"  #@UndefinedVariable

                        end_line_2 = self.newline
                        line_numbers_written = [start_line_2, end_line_2]
                # Catch empty list
                elif v == []:
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    self.this_sheet.cell(self.newline, self.newcolumn+1, "").style = "style2"
                        
            end_row = self.newline

            # Apply grey fill
            if title_count%2 == 0:# and line_numbers_written != None:
                start_row_2, end_row_2 = line_numbers_written
                self.apply_grey_fill(start_row_2-2, end_row_2, 1)
    
        # Apply border
        self.apply_border2(self.this_sheet, start_row, end_row, "A", "B")
    
        # save Worksheet
        self.workbook.save(self.xls_to_save)


    def create_xls_adjudication_section2(self):
        '''
        Creates the ADJUDICATION section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating the ADJUDICATION section of the spreadsheet")
        
        self.this_sheet = self.tab1_sheet
        self.newline += 2
        
        title_string = "Lands Adjudication Information"
        self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1" 
        
        title_count = 0
        start_row = self.newline
        
        working_dict = {"Upland Parcel Fronts Body of Water":   "",
                        "Is Upland Titled":                     "",
                        "Foreshore Parcel Adjoins Upland":      "",
                        "Nearest Arch Site within 25km":        "",
                        "Overview Assessment Potential":        self.xl_arch_potential
                        }
                        
        for title_string, v in working_dict.items():
            self.newline += 1
            # For Alternating grey fill
            title_count += 1
            line_numbers_written = [self.newline, self.newline]
            
            try:
                # Catch values that need further processing and formating
                if title_string == "Nearest Arch Site within 25km":
                    ires_pt1 = self.xl_nearest_arch[0].decode('utf-8')
                    #ires_pt2 = "(km)"
                    ires_pt2 = self.xl_nearest_arch_name[0].decode('utf-8')
                    #v = ires_pt1 + " " +  ires_pt2 + " - " + ires_pt3
                    v = f"{ires_pt1} (km) - {ires_pt2}"
            except:
                v = "No Arch Site within 25km"

            # Evaluate "v" and branch accordingly
            if isinstance(v, str):
                self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                self.this_sheet.cell(self.newline, self.newcolumn+1, v).style = "style2"
                line_numbers_written = [self.newline, self.newline]
            elif isinstance(v, list):
                if v != []:
                    start_line_2 = self.newline
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    for x in v:
                        # fist item in list goes on line with title
                        if x == v[0]:
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        else:
                            self.newline += 1
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        end_line_2 = self.newline
                        line_numbers_written = [start_line_2, end_line_2]
                elif v == []:
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    self.this_sheet.cell(self.newline, self.newcolumn+1, "").style = "style2"
            
            end_row = self.newline
            
            # Apply grey fill
            if title_count%2 == 0:
                start_row_2, end_row_2 = line_numbers_written
                #print("*** {} *** {}:{}  ({})".format(title_string, start_row_2, end_row_2, self.newline))
                self.apply_grey_fill(start_row_2-2, end_row_2, 1)
    
        # Apply border
        self.apply_border2(self.this_sheet, start_row, end_row, "A", "B")
   
        # save Worksheet
        self.workbook.save(self.xls_to_save)


    
    def create_xls_mines_section2(self):
        '''
        Creates the XLS section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating the MINES section of the spreadsheet")
        
        self.this_sheet = self.tab1_sheet
        self.newline += 2
        
        title_string = "Mines  -  See status number at top of report"
        self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1" 
        
        title_count = 0
        start_row = self.newline
        
        working_dict = {"Mine Number":                          "",
                        "Permit Number":                        "",
                        "Mine Type":                            "",
                        "Mine Number":                          "",
                        "Mine Name":                            "",
                        "Mine Manager":                         "",
                        "Inspector":                            "",
                        "New Permit":                           "",
                        "Crown":                                "",
                        "Private":                              "",
                        "License Number":                       ""
                        }
                        
        for title_string, v in working_dict.items():
            self.newline += 1
            # For Alternating grey fill
            title_count += 1
            line_numbers_written = [self.newline, self.newline]
            
            # Catch values that need further processing and formatting
            # NONE

            # Evaluate "v" and branch accordingly
            if isinstance(v, str):
                self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                self.this_sheet.cell(self.newline, self.newcolumn+1, v).style = "style2"
                line_numbers_written = [self.newline, self.newline]
                
            elif isinstance(v, list):
                if v != []:
                    start_line_2 = self.newline
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    for x in v:
                        # fist item in list goes on line with title
                        if x == v[0]:
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        else:
                            self.newline += 1
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        end_line_2 = self.newline
                        line_numbers_written = [start_line_2, end_line_2]
                elif v == []:
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    self.this_sheet.cell(self.newline, self.newcolumn+1, "").style = "style2"
            
            end_row = self.newline

            # Apply grey fill
            if title_count%2 == 0:
                start_row_2, end_row_2 = line_numbers_written
                #print("*** {} *** {}:{}  ({})".format(title_string, start_row_2, end_row_2, self.newline))
                self.apply_grey_fill(start_row_2-2, end_row_2, 1)
    
        # Apply border
        self.apply_border2(self.this_sheet, start_row, end_row, "A", "B")
        
        # save Worksheet
        self.workbook.save(self.xls_to_save)
        

    def create_xls_forests_section2(self):
        '''
        Creates the FORESTS section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating the FORESTS section of the spreadsheet")
        
                
        self.this_sheet = self.tab1_sheet
        self.newline += 2
        
        title_string = "Forests"
        self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1" 
        
        title_count = 0
        start_row = self.newline
        
        working_dict = {"License Number":                       "",
                        "License Type":                         "",
                        "Area":                                 "",
                        "Cubic Metres":                         "",
                        "Client Number":                        "",
                        "Forest Professional":                  ""
                        }
                        
        for title_string, v in working_dict.items():
            self.newline += 1
            # For Alternating grey fill
            title_count += 1
            line_numbers_written = [self.newline, self.newline]
            
            # Catch values that need further processing and formating
            # NONE

            # Evaluate "v" and branch accordingly
            if isinstance(v, str):
                self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                self.this_sheet.cell(self.newline, self.newcolumn+1, v).style = "style2"
                line_numbers_written = [self.newline, self.newline]
                
            elif isinstance(v, list):
                if v != []:
                    start_line_2 = self.newline
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    for x in v:
                        # fist item in list goes on line with title
                        if x == v[0]:
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        else:
                            self.newline += 1
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        end_line_2 = self.newline
                        line_numbers_written = [start_line_2, end_line_2]
                elif v == []:
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    self.this_sheet.cell(self.newline, self.newcolumn+1, "").style = "style2"
            
            end_row = self.newline
            
            
            #print(" # {} is count {}".format(title_string, title_count))
            # Apply grey fill
            if title_count%2 == 0:
                start_row_2, end_row_2 = line_numbers_written
                #print("*** {} *** {}:{}  ({})".format(title_string, start_row_2, end_row_2, self.newline))
                self.apply_grey_fill(start_row_2-2, end_row_2, 1)
    
        # Apply border
        self.apply_border2(self.this_sheet, start_row, end_row, "A", "B")
    
        # save Worksheet
        self.workbook.save(self.xls_to_save)
    
    
    def create_xls_water_section2(self):
        '''
        Creates the WATER section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating the WATER section of the spreadsheet")
        
        self.this_sheet = self.tab1_sheet
        self.newline += 2

        title_string = "Water"
        self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1" 
        
        title_count = 0
        start_row = self.newline
        
        working_dict = {"File Number":                          "",                           
                        "Stream Name":                          "",
                        "Purpose":                              "",
                        "Authorization Type":                   "",
                        "e-Licensing Report Attached":          ""
                        }
                        
        for title_string, v in working_dict.items():
            self.newline += 1
            # For Alternating grey fill
            title_count += 1
            line_numbers_written = [self.newline, self.newline]
            
            # Catch values that need further processing and formating
            # NONE

            # Evaluate "v" and branch accordingly
            if isinstance(v, str):
                self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                self.this_sheet.cell(self.newline, self.newcolumn+1, v).style = "style2"
                line_numbers_written = [self.newline, self.newline]
                
            elif isinstance(v, list):
                if v != []:
                    start_line_2 = self.newline
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    for x in v:
                        # fist item in list goes on line with title
                        if x == v[0]:
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        else:
                            self.newline += 1
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        end_line_2 = self.newline
                        line_numbers_written = [start_line_2, end_line_2]
                elif v == []:
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    self.this_sheet.cell(self.newline, self.newcolumn+1, "").style = "style2"
            
            end_row = self.newline
            
            
            #print(" # {} is count {}".format(title_string, title_count))
            # Apply grey fill
            if title_count%2 == 0:
                start_row_2, end_row_2 = line_numbers_written
                #print("*** {} *** {}:{}  ({})".format(title_string, start_row_2, end_row_2, self.newline))
                self.apply_grey_fill(start_row_2-2, end_row_2, 1)
                ##print("    is greyed...")
    
        # Apply border
        self.apply_border2(self.this_sheet, start_row, end_row, "A", "B")
    
        # save Worksheet
        self.workbook.save(self.xls_to_save)
        
        
    def create_xls_comments_section2(self):
        '''
        Creates the COMMENTS section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating the COMMENTS section of the spreadsheet")
        
        self.this_sheet = self.tab1_sheet
        self.newline += 2
        
        title_string = "Additional Comments"
        self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1" 
        self.newline += 1
        
        merge_cells = "A" + str(self.newline) + ":" + "B" + str(self.newline)
        self.this_sheet.merge_cells(merge_cells)
        #self.this_sheet.cell(self.newline, self.newcolumn).alignment = Alignment(wrap_text=True)
        #self.this_sheet.cell(self.newline, self.newcolumn).alignment = Alignment(wrapText=True)
        self.this_sheet.row_dimensions[self.newline].height = 95
        self.this_sheet.cell(self.newline, self.newcolumn, "").style = "style1"
        
        # apply border
        self.apply_border2(self.this_sheet, self.newline-1, self.newline+1, "A", "B")

        # Add additional text
        self.newline += 2
        title_string = "SEE TAB 3  CONFLICTS & CONSTRAINTS FOR COMPLETE STATUS"              #@UnusedVariable
        self.this_sheet.cell(self.newline, self.newcolumn, title_string).font =  Font(color='ff0000',size=12,name='Arial')
        
        # Add this run to run count in logfile
        logfile = r"\\spatialfiles.bcgov\work\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\automated_status_log\automated_runs_log_file.txt"
        file = open(logfile)
        this_many_runs = len(file.readlines())
        file.close()

        # Add to Spreadsheet 
        self.newline += 2
        run_count_string = "Automated Status Run # - " + str(this_many_runs + 1)
        self.this_sheet.cell(self.newline, self.newcolumn, run_count_string).style = "style2"

        self.workbook.save(self.xls_to_save)


    def create_xls_omineca_section2(self):
        '''
        Creates the OMINECA section of the spreadsheet
        '''
        
        arcpy.AddMessage(".  Creating the OMINECA section of the spreadsheet")

        self.this_sheet = self.tab1_sheet
        self.newline += 2

        title_string = "Omineca Manual Entries  -  yes / no / unk / exist / prop / adj"
        self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1" 
        
        title_count = 0
        start_row = self.newline
        
        working_dict = {"Stream Crossings":                     "",
                        "Within 30m of Waterbody":              "",
                        "New Access":                           "",
                        "Bulk Sample (Mines)":                  "",
                        "Underground Works (Mines)":            "",
                        "Waste Discharge":                      "",
                        }
                        
        for title_string, v in working_dict.items():
            self.newline += 1
            # For Alternating grey fill
            title_count += 1
            line_numbers_written = [self.newline, self.newline]
            
            # Catch values that need further processing and formating
            # NONE

            # Evaluate "v" and branch accordingly
            if isinstance(v, str):
                self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                self.this_sheet.cell(self.newline, self.newcolumn+1, v).style = "style2"
                line_numbers_written = [self.newline, self.newline]
                
            elif isinstance(v, list):
                if v != []:
                    start_line_2 = self.newline
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    for x in v:
                        # fist item in list goes on line with title
                        if x == v[0]:
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        else:
                            self.newline += 1
                            self.this_sheet.cell(self.newline, self.newcolumn+1, x).style = "style2"
                        end_line_2 = self.newline
                        line_numbers_written = [start_line_2, end_line_2]
                elif v == []:
                    self.this_sheet.cell(self.newline, self.newcolumn, title_string).style = "style1"
                    self.this_sheet.cell(self.newline, self.newcolumn+1, "").style = "style2"
            
            end_row = self.newline
            
            
            #print(" # {} is count {}".format(title_string, title_count))
            # Apply grey fill
            if title_count%2 == 0:
                start_row_2, end_row_2 = line_numbers_written
                #print("*** {} *** {}:{}  ({})".format(title_string, start_row_2, end_row_2, self.newline))
                self.apply_grey_fill(start_row_2-2, end_row_2, 1)
    
        # Apply border
        self.apply_border2(self.this_sheet, start_row, end_row, "A", "B")
    
        # save Worksheet
        self.workbook.save(self.xls_to_save)


    def create_spreadsheet_status_of_conflict2(self):
        '''
        Creates the 2nd tab in spreadsheet: STATUS OF CONFLICT
        '''
        
        arcpy.AddMessage(".  Creating the STATUS OF CONFLICT tab")
        
        # Create Styles
        active_style = NamedStyle(name="active")
        active_style.font = Font(size=9)
        active_style.fill = PatternFill("solid", fgColor="ff7c80")
        active_style.alignment = Alignment(horizontal='center')

        inactive_style = NamedStyle(name="inactive")
        inactive_style.font = Font(size=9)
        inactive_style.fill = PatternFill("solid", fgColor="ffff00")
        inactive_style.alignment = Alignment(horizontal='center')
        
        requested_style = NamedStyle(name="requested")
        requested_style.font = Font(size=9)
        requested_style.fill = PatternFill("solid", fgColor="66ff66")
        requested_style.alignment = Alignment(horizontal='center')
        
        self.this_sheet = self.tab2_sheet
        self.newline = 3
        
        # column width
        self.this_sheet.column_dimensions["A"].width = 20
        self.this_sheet.column_dimensions["B"].width = 60
        self.this_sheet.column_dimensions["C"].width = 10
        
        # File Number
        self.this_sheet.cell(self.newline, self.newcolumn, "File:").style = "style1"
        self.this_sheet.cell(self.newline, self.newcolumn+1, self.crown_lands_file).style = "style2"
        self.newline += 1

        # arcpy.AddMessage(f"inactives = {self.inactive_features}/{self.business_identifier}")
        if self.inactive_features is True:
            self.newline += 1
            inactive_disclaimer = "'Status of Conflict' only evaluates against inactive crown land tenures."
            self.this_sheet.cell(self.newline, self.newcolumn, inactive_disclaimer).font =  Font(color='ff0000',size=12,name='Arial')
            y = 0
            for x in self.business_identifier:      #@UnusedVariable
                self.newline += 1
                i = 0
                start_row = self.newline
                
                # top line              
                # Write Headers for Business Identifier
                self.this_sheet.cell(self.newline, self.newcolumn, "Record")#column_1_order[i])
                stat_val = self.interest_status[y]
                
                if stat_val == "ACTIVE":
                    self.this_sheet.cell(self.newline, self.newcolumn+1, stat_val).style = active_style
                elif stat_val == "INACTIVE":
                    self.this_sheet.cell(self.newline, self.newcolumn+1, stat_val).style = inactive_style
                elif stat_val == "REQUESTED":
                    self.this_sheet.cell(self.newline, self.newcolumn+1, stat_val).style = requested_style
                else:
                    self.this_sheet.cell(self.newline, self.newcolumn+1, stat_val)
                
                self.this_sheet.cell(self.newline, self.newcolumn+2, self.responsible_agency[y])
                i += 1
                self.newline += 1
                
                ## Populate rows
                # Disp Trans SID and File Number
                self.this_sheet.cell(self.newline, self.newcolumn, "Business Identifier:")#column_1_order[i])
                self.this_sheet.cell(self.newline, self.newcolumn+1, self.business_identifier[y])
                i += 1
                self.newline += 1
                # NRS Business Area
                self.this_sheet.cell(self.newline, self.newcolumn, "DPR Registry Name:")#column_1_order[i])
                self.this_sheet.cell(self.newline, self.newcolumn+1, self.dpr_registry_name[y])
                i += 1
                self.newline += 1
                # Interest Type
                self.this_sheet.cell(self.newline, self.newcolumn, "Interest Type:")#column_1_order[i])
                self.this_sheet.cell(self.newline, self.newcolumn+1, self.interest_type[y])
                i += 1
                self.newline += 1                
                        
                self.this_sheet.cell(self.newline, self.newcolumn, "Name:")#column_1_order[i])
                self.this_sheet.cell(self.newline, self.newcolumn+1, self.summary_holders_ilrr_identifier[y])
                i += 1
                self.newline += 1
                
                # Apply border
                end_row = self.newline-1
                self.apply_border2(self.this_sheet, start_row, end_row, "A", "C")
                y += 1
                
            # save Worksheet
            self.workbook.save(self.xls_to_save)

        elif self.inactive_features is False:
            self.newline += 1
            no_inactive_disclaimer = "'Status of Conflict' only evaluates against inactive crown land tenures. No inactive crown land tenures reported."
            self.this_sheet.cell(self.newline, self.newcolumn, no_inactive_disclaimer).font =  Font(color='ff0000',size=12,name='Arial')
            self.workbook.save(self.xls_to_save)
        else:
            self.newline += 1
            error_disclaimer = "ERROR: The AST was unable to evaluate the AOI against inactive crown land tenures. Please contact FCBCIntranet-SharePointSupport@gov.bc.ca for support"
            self.this_sheet.cell(self.newline, self.newcolumn, error_disclaimer).font =  Font(color='ff0000',size=12,name='Arial')
            self.workbook.save(self.xls_to_save)
    

    def create_spreadsheet_status_of_conflict2_csv(self):
        '''
        Remove for production
        Creates the 2nd tab in spreadsheet from CSV: STATUS OF CONFLICT
        '''
        
        arcpy.AddMessage(".  Creating the DEPRECATED CSV - STATUS OF CONFLICT tab")
        
        # Create Styles
        active_style_csv = NamedStyle(name="active_csv")
        active_style_csv.font = Font(size=9)
        active_style_csv.fill = PatternFill("solid", fgColor="ff7c80")
        active_style_csv.alignment = Alignment(horizontal='center')

        inactive_style_csv = NamedStyle(name="inactive_csv")
        inactive_style_csv.font = Font(size=9)
        inactive_style_csv.fill = PatternFill("solid", fgColor="ffff00")
        inactive_style_csv.alignment = Alignment(horizontal='center')
        
        requested_style_csv = NamedStyle(name="requested_csv")
        requested_style_csv.font = Font(size=9)
        requested_style_csv.fill = PatternFill("solid", fgColor="66ff66")
        requested_style_csv.alignment = Alignment(horizontal='center')
        
        self.this_sheet = self.tab3_sheet
        self.newline = 3
        
        # column width
        self.this_sheet.column_dimensions["A"].width = 20
        self.this_sheet.column_dimensions["B"].width = 60
        self.this_sheet.column_dimensions["C"].width = 10
        
        # File Number
        self.this_sheet.cell(self.newline, self.newcolumn, "File:").style = "style1"
        self.this_sheet.cell(self.newline, self.newcolumn+1, self.crown_lands_file).style = "style2"
        self.newline += 1
        
        y = 0
        for x in self.business_identifier:      #@UnusedVariable
            #print(x)
            if y >= 1:
                self.newline += 1
                i = 0
                start_row = self.newline
                
                # top line              
                # Write Headers for Business Identifier
                self.this_sheet.cell(self.newline, self.newcolumn, "Record")#column_1_order[i])
                stat_val = self.interest_status[y]
                
                if stat_val == "ACTIVE":
                    self.this_sheet.cell(self.newline, self.newcolumn+1, stat_val).style = active_style_csv
                elif stat_val == "INACTIVE":
                    self.this_sheet.cell(self.newline, self.newcolumn+1, stat_val).style = inactive_style_csv
                elif stat_val == "REQUESTED":
                    self.this_sheet.cell(self.newline, self.newcolumn+1, stat_val).style = requested_style_csv
                else:
                    self.this_sheet.cell(self.newline, self.newcolumn+1, stat_val)
                
                self.this_sheet.cell(self.newline, self.newcolumn+2, self.responsible_agency[y])
                i += 1
                self.newline += 1
                
                ## Populate rows
                # Disp Trans SID and File Number
                self.this_sheet.cell(self.newline, self.newcolumn, "Business Identifier:")#column_1_order[i])
                self.this_sheet.cell(self.newline, self.newcolumn+1, self.business_identifier[y])
                i += 1
                self.newline += 1
                # NRS Business Area
                self.this_sheet.cell(self.newline, self.newcolumn, "DPR Registry Name:")#column_1_order[i])
                self.this_sheet.cell(self.newline, self.newcolumn+1, self.dpr_registry_name[y])
                i += 1
                self.newline += 1
                # Interest Tyle
                self.this_sheet.cell(self.newline, self.newcolumn, "Interest Type:")#column_1_order[i])
                self.this_sheet.cell(self.newline, self.newcolumn+1, self.interest_type[y])
                i += 1
                self.newline += 1
                # Interest Holders
                this_person = ""
                for xx in self.interest_holders_ilrr_identifier:
                    array = xx.split(',')
                    first_item = array[0].replace('"', '') 
                    second_item = array[1].replace('"', '')
                    if first_item == self.summary_holders_ilrr_identifier[y]:
                        this_person = this_person + "  |  " + second_item
                        
                self.this_sheet.cell(self.newline, self.newcolumn, "Name:")#column_1_order[i])
                self.this_sheet.cell(self.newline, self.newcolumn+1, this_person.strip("  |  "))
                i += 1
                self.newline += 1
                
                # Apply border
                end_row = self.newline-1
                self.apply_border2(self.this_sheet, start_row, end_row, "A", "C")
            y += 1
            
            # save Worksheet
            self.workbook.save(self.xls_to_save)


#_________________________________________________________________#

        
if __name__ == '__main__':
    print("trying to run it from local")

    run_one_status_part2 = True
    test_spreadsheet = True
    test_make_unique_list_of_field_values = False


    if test_make_unique_list_of_field_values:
        
        fc_name = r"\\spatialfiles.bcgov\work\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\automated_status_ARCPRO\wes_python_development\Test_run\UOT_TestPoly.shp"
        definition_query = ""
        field3 = 'Block_Number'
        field2 = 'Group'
        field1 = 'asdfasdfasdf'
        
        dataObj = revolt_tool()     #@UndefinedVariable
        mylist = dataObj.make_unique_list_of_field_values(fc_name,definition_query,field1,field2,field3)
        print("mylist is ")
        print(mylist)


    if run_one_status_part2:
        #data_path = passed_input_list[0]
        #dont_overwrite_outputs = passed_input_list[1]   
        #region = passed_input_list[2]   
        #crown_lands_file = passed_input_list[3]   
        #disposition_transaction = passed_input_list[4]   

        part2_criteria_to_pass = []
        part2_criteria_to_pass.append(r"\\spatialfiles.bcgov\work\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\automated_status_ARCPRO\test_runs\wes_test_v3\Wes_Test_v3.shp")
        #part2_criteria_to_pass.append(r"\\spatialfiles.bcgov\work\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\automated_status_ARCPRO\wes_python_development\Test_run\UOT_TestPoly.shp")
        #part2_criteria_to_pass.append(r"\\granite\work\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\1414073_csv_mmtest")   # TEXT - analyize_this_featureclass 
        #part2_criteria_to_pass.append(r"\\granite\work\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\ATS 104512")   # TEXT - analyize_this_featureclass 
        #part2_criteria_to_pass.append(r"\\granite\work\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\5407649_902281")   # TEXT - analyize_this_featureclass 
        part2_criteria_to_pass.append(r"true")          # dont_overwrite_outputs
        part2_criteria_to_pass.append(r"cariboo")       # region
        part2_criteria_to_pass.append(r"#")             # crown_lands_file
        part2_criteria_to_pass.append(r"#")             # disposition_transaction
        part2_criteria_to_pass.append(r"#")             # crown_lands_file
        part2_criteria_to_pass.append(r"#")             # disposition_transaction

        dataObj = one_status_part2_tool()
        dataObj.run_tool(part2_criteria_to_pass)
        print("_____________________________________________________________________")
        