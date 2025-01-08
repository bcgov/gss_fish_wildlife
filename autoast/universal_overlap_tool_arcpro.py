'''
Author: Mark McGirr
Purpose: This script takes an input feature class, from either a geodatabase or a shape file
            which can be a point line or polygon and overlays it will all the files specified 
            in an analysis_type spreadsheet.
            It generate an XLS file of all the overlaps and creates a small overlap map for each
            of the overlap types.
            
Date: Feb 13, 2012

Arguments: argv[1] = feature class to analyze
           argv[2] = create sub-reports on this field in the above feature class
           argv[3] = the type of analysis to run.  This is a pick list of xls spreadsheets stored on the P: drive
           argv[4] = an xls file the you can use for inputs if you don't want one of the ones from the pick list
           argv[5] = header information to print on the output xls file
           argv[6] = header information to print on the output xls file
           argv[7] = header information to print on the output xls file
           argv[8] = header information to print on the output xls file 
           arcv[9] = boolean if you want the individual sub-reports to be on the same page, or different xls tabs
           arcv[10] = path to create the output xls in if not in same directory as the input feature class
           arcv[11] = boolean if you want report fields to be split onto multiple lines of the xls



Outputs: An xls file of all the overlaps, listing details of each type of overlap, and producing
            maps to view each of the overlaps in. 

Dependencies: MUST BE RUN IN ArcGIS PRO
              Must have access to p:\corp\script_whse\python\\Utility_Misc\Ready\

History: 
----------------------------------------------------------------------------------------------
Date:
Author:
Modification:
'----------------------------------------------------------------------------------'

''' 
import sys, os, time, datetime, arcpy, csv, runpy, shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from openpyxl.styles import Color
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side

from fc_to_html import HTMLGenerator

arcpy.env.overwriteOutput = True

log_file_directory = r'\\spatialfiles.bcgov\work\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\automated_status_ARCPRO\status_logs'


class revolt_tool(object):
    
    # this is the data utilities and applications object
    def __init__(self):
        pass


    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    def run_revolt_tool(self, passed_input_list):

        StartTime = time.perf_counter()   
        
        self.sde_connection = os.getenv("SDE_FILE_PATH")

        self.assign_english_names_to_list_position_numbers()
        self.split_passed_in_argument_into_individual_variables(passed_input_list)
        self.write_log_file_for_debugging()

        EnviroEndTime = time.perf_counter()
        EnviroEndTimeStr = "Set Environment Time is " + str(int(EnviroEndTime - StartTime))
        arcpy.AddMessage(EnviroEndTimeStr)

        self.create_working_directories_geodatabases_and_variables()
        self.copy_aoi_into_analysis_gdb()
        CreateGdbEndTime = time.perf_counter()
        CreateGdbEndTimeStr = "Create GDB Time is " + str(int(CreateGdbEndTime - EnviroEndTime))
        arcpy.AddMessage(CreateGdbEndTimeStr)

        self.read_input_spreadsheet()
        ReadInputsEndTime = time.perf_counter()
        ReadInputsEndTimeStr = "Read Inputs Time is " + str(int(ReadInputsEndTime - CreateGdbEndTime))
        arcpy.AddMessage(ReadInputsEndTimeStr)     
        
        self.clip_input_datasets_to_aoi()
        ClipDataEndTime = time.perf_counter()
        ClipDataEndTimeStr = "Clip Data Time is " + str(int(ClipDataEndTime - ReadInputsEndTime))
        arcpy.AddMessage(ClipDataEndTimeStr)
        
        if self.suppress_map_creation != 'true':
            arcpy.AddMessage('Generating HTML maps' )
            self.make_html_maps()
            # self.aprx = self.create_set_aprx()
            # self.make_overview_maps()
            # self.make_the_maps_with_labels()
            # self.delete_project()

        MakeMapsEndTime = time.perf_counter()
        MakeMapsEndTimeStr = "Make Maps Time is " + str(int(MakeMapsEndTime - ClipDataEndTime))
        arcpy.AddMessage(MakeMapsEndTimeStr)

        
        self.create_spreadsheet()

        MakeXlsEndTime = time.perf_counter()
        MakeXlsEndTimeStr = "Make Spreadsheet Time is " + str(int(MakeXlsEndTime - MakeMapsEndTime))
        arcpy.AddMessage(MakeXlsEndTimeStr)

        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Universal Overlap Tool finished running')

        EndTime = time.perf_counter()
        TotalRunTimeStr = "Total Run Time is " + str(int(EndTime - StartTime))

        arcpy.AddMessage(EnviroEndTimeStr)
        arcpy.AddMessage(CreateGdbEndTimeStr)
        arcpy.AddMessage(ReadInputsEndTimeStr)
        arcpy.AddMessage(ClipDataEndTimeStr)
        arcpy.AddMessage(MakeMapsEndTimeStr)
        arcpy.AddMessage(MakeXlsEndTimeStr)
        arcpy.AddMessage(TotalRunTimeStr)
        
    
    def write_log_file_for_debugging(self):
        '''
        This writes the relevant arguments into a csv file to make debugging easier
        '''
        
        try:
            arcpy.AddMessage("======================================================================")
            arcpy.AddMessage('Writing log file ')
            
            today = str(datetime.date.today())
            user = os.environ['USERNAME']
            
            log_data = []
            log_data.append(today)
            log_data.append(user)
            log_data.append(self.analyize_this_featureclass)
            log_data.append(self.directory_to_store_output)
            log_data.append(self.create_subreports_on_this_field)
            log_data.append(self.what_type_of_overlap_to_run)
            log_data.append(self.xls_file_for_analysis_input)
            log_data.append(self.xls_file_for_analysis_input2)
        
            log_name = "revolt_log_file.csv"
            log_file = os.path.join(log_file_directory,log_name)
        
            out = csv.writer(open(log_file,"a", newline=''),delimiter=",",quoting=csv.QUOTE_ALL)
            out.writerow(log_data)

        except:
            pass

    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    
    def get_size_of_log_file(self):
        '''
        This returns the size of the log file.  Just for interest.
        '''
        try:  
            log_name = "revolt_log_file.csv"
            log_file = os.path.join(log_file_directory,log_name)
            with open(log_file) as f:
                size_of_log = sum(1 for line in f)
        except:
            size_of_log = -99  
        
        return(size_of_log)
        
    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    
    def assign_english_names_to_list_position_numbers(self):
        '''
        these are the column locations in the input spreadsheet, and as well in the input spreadsheet list[]
        that it's read into.  It is easier to refer to these names in the list than the position number in the list.
        '''

        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Assigning English names to list positions')
         
        self.rpt_category = 0
        self.rpt_data_name = 1
        self.rpt_data_source = 2
        self.rpt_fld_to_summarize1 = 3
        self.rpt_def_query = 4
        self.rpt_buf_distance = 5
        self.rpt_fld_to_summarize2 = 6
        self.rpt_fld_to_summarize3 = 7
        self.rpt_fld_to_summarize4 = 8
        self.rpt_fld_to_summarize5 = 9
        self.rpt_fld_to_summarize6 = 10
        self.rpt_label_field = 11
        self.rpt_aoi_for_clip = 12
        self.rpt_clipped_fc_name = 14
        self.rpt_error_flag = 15
    
    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    
    def split_passed_in_argument_into_individual_variables(self, passed_input_list):
        '''
        this splits the passed in argument into individual variables
        '''
        
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Splitting the arguments into individual variables')
        
        self.analyize_this_featureclass = passed_input_list[0]
        self.create_subreports_on_this_field = passed_input_list[1]
        self.what_type_of_overlap_to_run = passed_input_list[2]
        self.xls_file_for_analysis_input = passed_input_list[3]
        self.report_header1 = passed_input_list[4]
        self.report_header2 = passed_input_list[5]
        self.report_header3 = passed_input_list[6]
        self.report_header4 = passed_input_list[7]
        self.subreports_on_seperate_sheets = passed_input_list[8]
        self.directory_to_store_output = passed_input_list[9]
        self.summary_fields_on_seperate_lines = passed_input_list[10]
        self.dont_overwrite_data_and_maps = passed_input_list[11]
        self.xls_file_for_analysis_input2 = passed_input_list[12]
        self.disclaimer = passed_input_list[13]
        self.suppress_map_creation = passed_input_list[14]
        self.region = passed_input_list[15]
        self.crown_file_number = passed_input_list[16]
        self.disposition_number = passed_input_list[17]
        self.parcel_number = passed_input_list[18]
        self.run_as_fcbc = passed_input_list[19]
        self.add_maps_to_current = passed_input_list[20]
        

        self.subreports_on_seperate_sheets = self.subreports_on_seperate_sheets.lower()
        self.summary_fields_on_seperate_lines = self.summary_fields_on_seperate_lines.lower()
        self.dont_overwrite_data_and_maps = self.dont_overwrite_data_and_maps.lower()
    
        
        arcpy.AddMessage("The input criteria are:")
        arcpy.AddMessage("    analyize_this_featureclass = " + str(self.analyize_this_featureclass))
        arcpy.AddMessage("    create_subreports_on_this_field = " + self.create_subreports_on_this_field)
        arcpy.AddMessage("    what_type_of_overlap_to_run = " + self.what_type_of_overlap_to_run)
        arcpy.AddMessage("    xls_file_for_analysis_input = " + self.xls_file_for_analysis_input)
        arcpy.AddMessage("    report_header1 = " + self.report_header1)
        arcpy.AddMessage("    report_header2 = " + self.report_header2)
        arcpy.AddMessage("    report_header3 = " + self.report_header3)
        arcpy.AddMessage("    report_header4 = " + self.report_header4)
        arcpy.AddMessage("    subreports_on_seperate_sheets = " + self.subreports_on_seperate_sheets)#, ; arcpy.AddMessage(subreports_on_seperate_sheets)
        arcpy.AddMessage("    directory_to_store_output = " + self.directory_to_store_output)
        arcpy.AddMessage("    summary_fields_on_seperate_lines = " + self.summary_fields_on_seperate_lines)# , ; arcpy.AddMessage(summary_fields_on_seperate_lines) 
        arcpy.AddMessage("    dont_overwrite_data_and_maps = " + self.dont_overwrite_data_and_maps)# , ; arcpy.AddMessage(dont_overwrite_data_and_maps) 
        arcpy.AddMessage("    xls_file_for_analysis_input2 = " + self.xls_file_for_analysis_input2)# , ; arcpy.AddMessage(dont_overwrite_data_and_maps) 
        arcpy.AddMessage("    disclaimer = " + self.disclaimer)# , ; arcpy.AddMessage(dont_overwrite_data_and_maps) 
        arcpy.AddMessage("    suppress_map_creation = " + self.suppress_map_creation)# , ; arcpy.AddMessage(dont_overwrite_data_and_maps) 
        arcpy.AddMessage("    region = " + self.region)
        arcpy.AddMessage("    crown_file_number = " + self.crown_file_number)
        arcpy.AddMessage("    disposition_number = " + self.disposition_number)
        arcpy.AddMessage("    parcel_number = " + self.parcel_number)
        arcpy.AddMessage("    run_as_fcbc = " + self.run_as_fcbc)
        arcpy.AddMessage("    add_maps_to_current = " + self.add_maps_to_current)

    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    
    def create_working_directories_geodatabases_and_variables(self):
        '''
        Creates the working directory.  Deletes the existing working GDB and MAP directory 
        if the 'dont_overwrite_flag' is not set to true.  It then creates the MAP directory.
        
        
        Also sets the following variables for universal use. 
        
        self.work_directory     -    the directory where the work will be performed
                                ie.  \\granite\work\srm\wml\workarea\arcproj\!williams_lake_toolbox_development\5407649
        
        self.work_gdb           -    the complete path to the working GDB file, with a .GDB extension
                                ie.  \\granite\work\srm\wml\workarea\arcproj\!williams\5407649\one_status_common_datasets_aoi.gdb
        
        xls_for_analysis        -    the name of the spreadsheet to be read without the XLS extension.
                                ie.  one_status_common_datasets_debug_version

        featureclass_to_analyize  -  aoi (should always be this because the parcel boundary is copied into the GDB with this name
        
        self.map_directory      -    the complete path to the directory where the maps will be created.
                                ie.  \\granite\work\srm\wml\workarea\arcproj\!williams\5407649\maps_for_one_status_common_datasets_aoi
        
        '''
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Creating working directories, geodatabases, variables')

        #this is the default directory for where the input XLS files are stored (p:)
        self.input_criteria_xls_directory = r"\\Giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\statusing_input_spreadsheets"
        
        # if the type of overlap to run is selected in the dropdown box in the tool, and the xls not specified 
        if self.what_type_of_overlap_to_run != "#"  and self.what_type_of_overlap_to_run != "":
            self.xls_file_for_analysis_input = os.path.join(self.input_criteria_xls_directory, self.what_type_of_overlap_to_run)# + ".xls")
    
        # this just returns the name of the xls file without the .xls extension to print on the top of output report
        self.xls_for_analysis = os.path.split(self.xls_file_for_analysis_input)[1]
        self.xls_for_analysis = os.path.splitext(self.xls_for_analysis)[0]
        
        self.excel_input = self.xls_file_for_analysis_input

        # determine what the output (working) directory is.  If it is specified by the user in the tool use that directory instead
        try:
            if self.directory_to_store_output != "#" and self.directory_to_store_output != "":
                if not os.path.exists(self.directory_to_store_output):
                    try:
                        os.makedirs(self.directory_to_store_output)
                    except:
                        arcpy.AddError("The Output Folder Directory Does Not Exist and Could Not Be Created")
                        sys.exit()
                self.work_directory = self.directory_to_store_output  # choose the user defined directory to store the output xls and maps
            # Output directory set to where the input shapefile/feature class resides
            else:
                desc = arcpy.Describe(self.analyize_this_featureclass)
                if desc.dataType == "FeatureLayer":
                    self.analyize_this_featureclass = desc.catalogPath
                self.directory_to_store_output = get_fc_directory_name(str(self.analyize_this_featureclass))
                self.work_directory = self.directory_to_store_output
        except Exception as e:
            arcpy.AddWarning(e)
            sys.exit()

        #just the name of the fc to analyze with no path prefix, and no .shp extension
        self.featureclass_to_analyize = os.path.split(self.analyize_this_featureclass)[1]
        prefix, extension_type = os.path.splitext(self.featureclass_to_analyize)
        if extension_type == ".shp":
            self.featureclass_to_analyize = prefix

        #create the paths for the output folders and working geodatabase
        self.work_gdb = os.path.join(self.work_directory, self.xls_for_analysis + "_" + self.featureclass_to_analyize + ".gdb")
        self.map_directory = os.path.join(self.work_directory, "maps")
        self.mapx_directory = os.path.join(self.work_directory, "mapx_files")
        
        
        # delete the maps directory, and output GDB if flag is not set to true.
        dirs = [self.work_gdb, self.map_directory, self.mapx_directory]
        for d in dirs:
            try:
                #always delete the mapx folder if it exists
                if os.path.isdir(d) and d == self.mapx_directory:
                    shutil.rmtree(d)
                #if the folder exists and Run Again Mode is not checked, delete the dataset or folder
                if os.path.isdir(d) and self.dont_overwrite_data_and_maps != 'true' :
                    arcpy.AddMessage(f"deleting the existing file: {d}")
                    shutil.rmtree(d)
            except Exception as e:
                arcpy.AddError(f"Could not delete existing files. {e}")
                sys.exit()
        dirs = dirs[1:]
        for f in dirs:
            if not os.path.exists(f):
                os.makedirs(f)  # make the directory to store the maps and MAPX files.

        # create the GDB and feature dataset to do the work in
        self.input_dataset = os.path.join(self.work_gdb, "input_of_raw_data")
        create_feature_dataset_if_needed(self.input_dataset)

    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################

    # def create_set_aprx(self):
    #     '''
    #     Deletes the existing APRX file as it would cause conflict with generating
    #     new maps. A new aprx is then created on the T Drive for generating PDF maps.
    #     '''
    #     try:
    #         username = os.getenv('username')
    #         tmpDir = r"T:\StatusMaps_" + username
    #         template = r"\\Giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\map_files\dev_template.aprx"

    #         if not os.path.exists(tmpDir):
    #             os.mkdir(tmpDir)
    #         counter = 0
    #         filename = os.path.join(tmpDir,f"StatusMap_{username}{counter}.aprx")
    #         while os.path.isfile(filename):
    #             counter += 1
    #             filename = os.path.join(tmpDir,f"StatusMap_{username}{counter}.aprx")

    #         aprx = arcpy.mp.ArcGISProject(template)
    #         aprx.saveACopy(filename)
    #         aprx = arcpy.mp.ArcGISProject(filename)
    #         return(aprx)

    #     except Exception as e:
    #         arcpy.AddMessage(e)
    #         sys.exit()

    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    def add_new_fields(self, aoi):
        def fieldInfoUnpack(fieldInfo,name,alias):
            return  {"field_name":name,
            "field_type":fieldInfo.type,
            "field_precision":fieldInfo.precision,
            "field_scale":fieldInfo.scale,
            "field_length":fieldInfo.length,
            "field_alias":alias,
            "field_is_nullable":fieldInfo.isNullable,
            "field_is_required":fieldInfo.required,
            "field_domain":fieldInfo.domain}

        #add labelling field
        try:
            arcpy.AddField_management(aoi, "label_field", "TEXT", "", "", "40", "", "NULLABLE", "NON_REQUIRED", "")
        except Exception as e:
            pass

        #if the script is to report on a specific field, create a new field to be used for reporting to ensure
        #the field is not overwritten when intersecting with conflicts datasets
        if self.create_subreports_on_this_field != "" and self.create_subreports_on_this_field != "#":
            try:
                self.reporting_field = f"report_{self.create_subreports_on_this_field}"
                fieldInfo = arcpy.ListFields(aoi, self.create_subreports_on_this_field)[0]
                arcpy.AddField_management(aoi,**fieldInfoUnpack(fieldInfo, self.reporting_field,"Reporting"))
                arcpy.CalculateField_management(aoi, self.reporting_field, f"!{self.create_subreports_on_this_field}!")
            except Exception as e:
                pass

    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    def copy_aoi_into_analysis_gdb(self):
        '''
        Copies the parcel boundary into the working gdb.  Buffers it at 1 meter if it a point or line
        because identities need a polygon.
        
        Adds a label field to the created AOI to be used in writing labels on the maps.
        '''
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Copying the AOI into the working .GDB')

        # copy the input data shape to be the AOI.  If it's a point or line,  buffer it to be the AOI
        arcpy.AddMessage("    Creating the raw AOI ")
        desc = arcpy.Describe(self.analyize_this_featureclass)
        input_data_type = desc.ShapeType
        raw_output = os.path.join(self.input_dataset, "raw_aoi")
        if not arcpy.Exists(raw_output):
            arcpy.CopyFeatures_management(self.analyize_this_featureclass, raw_output)
        
        #Copy the AOI feature class to the 'aoi' feature class if polygon
        #if not a polygon, buffer it.
        the_output = os.path.join(self.work_gdb, "aoi")
        if not arcpy.Exists(the_output):
            arcpy.AddMessage("    Creating the final AOI ")
            if input_data_type == "Polygon":
                arcpy.CopyFeatures_management(raw_output, the_output)
            # buffer the shape at 1 meter if it's not a polygon because identities need polygons.  This is now the AOI
            else:
                arcpy.AddMessage("    Buffering the line or point feature class into final AOI ")
                arcpy.Buffer_analysis(raw_output, the_output, 1)
        #check for and add fields, if necessary.
        self.add_new_fields(the_output)

    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################
    def read_input_spreadsheet(self):
        '''
        This reads the input spreadsheets.  If two are specified it reads them both and
        merges them together.
        
        It then creates a master_control_list that is used for all the clipping, error flags,
        map_generation etc 
        
        self.input_datasources_list[]    is the master control list name
        '''

        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Reading the input XLS files')
    

        '''This reads both the input featureclass spreadsheets.
            The values are read into two different lists, then these
            two lists are appended together.
        ''' 
        if self.xls_file_for_analysis_input == ""  and self.xls_file_for_analysis_input2 == "":
            arcpy.AddError("    No Input spreadsheets specified")
            sys.exit()
        
        if self.xls_file_for_analysis_input != "":
            arcpy.AddMessage("    Reading input spreadsheet #1 - ")# + self.xls_file_for_analysis_input)
            data_1 = read_xls_into_list_of_lists(self.xls_file_for_analysis_input)
            del data_1[0] # delete the first row of the list because its just the field descriptions.

        if self.xls_file_for_analysis_input2 != "":
            arcpy.AddMessage("    Reading input spreadsheet #2 - ")# + self.xls_file_for_analysis_input2)
            data_2 = read_xls_into_list_of_lists(self.xls_file_for_analysis_input2)
            del data_2[0] # delete the first row of the list because its just the field descriptions.
        
        xls_to_read_from = [] # both spreadsheets appended into this list
        arcpy.AddMessage("    Merging Spreadsheets")
        for line in data_1:
            xls_to_read_from.append(line)
        if self.xls_file_for_analysis_input2 != "" and self.xls_file_for_analysis_input2 != "#"  :
            for line in data_2:
                xls_to_read_from.append(line)
        # end of Read the input data excel spreadsheets
        #----------------------------------------------------------
    
    
        #----------------------------------------------------------
        # Make Master Control List
        '''This reads the merged spreadsheet list and creates
            a master_control list.  There are blank spaces added
            to the end of each list row that can be used as flags,
            hold extra variables etc.
        ''' 
        self.input_datasources_list = [] # list that holds the values read in from the spreadsheet
        this_category = 'blank'
        arcpy.AddMessage("    Creating Master Control List")
        for this_row in  xls_to_read_from:
            category = this_row[0]
            data_name = this_row[1]
            data_source = this_row[2]
            definition_query = this_row[3]
            buffer_distance = this_row[4]
            field_to_summarize = this_row[5]
            field_to_summarize2 = this_row[6]
            field_to_summarize3 = this_row[7]
            field_to_summarize4 = this_row[8]
            field_to_summarize5 = this_row[9]
            field_to_summarize6 = this_row[10]
            label_field = this_row[11]
            
            
            # check to see if the category is different from the current category, and change if it is
            if this_category == 'blank' and category:
                this_category = category    
            if this_category != 'blank' and category:
                this_category = category
    
            excel_line = [this_category, data_name, data_source, field_to_summarize, definition_query, buffer_distance, field_to_summarize2, field_to_summarize3, field_to_summarize4, field_to_summarize5, field_to_summarize6, label_field, "", "", "", "", "", ""]
            self.input_datasources_list.append(excel_line)

    
    def clip_input_datasets_to_aoi(self):
        '''
        This clips out all the input datasets.
        
        Step 1 is to loop through the input list and see what distances are used in the buffer_by column.
        Then buffer the AOI by that distance into its own featureclass. 

        Step 2 is to loop through the input list and determine if the data is local, or on the BCGW.  If it
        is on the BCGW append the SDE connection string to the data field. 

        Step 3 is to loop through the input list doing an identity of all the input data fields.  At the same
        time the label_field is calculated so that maps can have a label on them. 
        '''
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Clipping the input datasets to the AOI')
 
        the_aoi = os.path.join(self.work_gdb, "aoi")
 
        # buffer the aoi at the values in the input spreadsheet, and set x[rpt_aoi_for_clip] field to the value to clip input data to
        arcpy.AddMessage("    Creating the buffered AOI's")
        for input_list_line in self.input_datasources_list:
            input_list_line[self.rpt_aoi_for_clip] = the_aoi  # this will be the dataset to clip the source to
            buffer_distance_for_aoi = input_list_line[self.rpt_buf_distance]
            try:
                buffer_distance_for_aoi = int(float(buffer_distance_for_aoi))
            except:
                buffer_distance_for_aoi = 0 # set buf dist to 0 if it's a blank
            
            if buffer_distance_for_aoi > 0 :
                # change the AOI that the input will be clipped to if there is a buffer distance.
                input_list_line[self.rpt_aoi_for_clip] = the_aoi + "_" + str(buffer_distance_for_aoi)
                the_input = the_aoi
                the_output = input_list_line[self.rpt_aoi_for_clip]
                
                # create the buffered AOIs
                if not arcpy.Exists(the_output):
                    arcpy.AddMessage("    Buffering the AOI at " + str(buffer_distance_for_aoi) + " meters")
                    raw_output = os.path.join(self.input_dataset, "raw_aoi")
                    if self.create_subreports_on_this_field != "" and self.create_subreports_on_this_field != "#":
                        arcpy.Buffer_analysis(raw_output, the_output, buffer_distance_for_aoi,"FULL")
                    else:
                        arcpy.Buffer_analysis(raw_output, the_output, buffer_distance_for_aoi, "OUTSIDE_ONLY")
                    arcpy.RepairGeometry_management(the_output)
                    self.add_new_fields(the_output)
                    #arcpy.AddField_management(the_output, "label_field", "TEXT", "", "", "40", "", "NULLABLE", "NON_REQUIRED", "")



        # get the sde connection string
        arcpy.AddMessage("    SDE = " + self.sde_connection)

                       
        # append the SDE path to the LRDW/BCGW if it is not a local dataset
        for input_list_line in self.input_datasources_list:
            data_source = input_list_line[self.rpt_data_source]
            input_list_line[self.rpt_clipped_fc_name] = 'local_data'
            # considered it a local data source if like \\granite or like w:\
            # if not local append the sde_datasource to the front of string         
            if not(data_source[0:2] == r'\\' or data_source[2:3] == "\\"):
                input_list_line[self.rpt_data_source] = os.path.join(self.sde_connection, input_list_line[self.rpt_data_source])
                input_list_line[self.rpt_clipped_fc_name] = 'bcgw_data'

    
        arcpy.AddMessage("    Intersecting the input featuresets")
        items_in_list = len(self.input_datasources_list) # the num of items in master_control list to display in messages
        xxx_count = 0 # just a counter do display in messages
        for input_list_line in self.input_datasources_list:
            xxx_count += 1
            the_input = input_list_line[self.rpt_data_source]
            the_output = os.path.join(self.work_gdb, input_list_line[self.rpt_data_name].replace (" ", "_"))
            if input_list_line[self.rpt_clipped_fc_name] == 'bcgw_data':
                intersect_message = "    intersecting " + str(xxx_count) + " of " + str(items_in_list) + "     " + "\\".join(the_input.split('\\')[-2:]) #print as bcgw.sde\FEATURE_CLASS without full path
            else:
                intersect_message = "    intersecting " + str(xxx_count) + " of " + str(items_in_list) + "     " + the_input
            arcpy.AddMessage(intersect_message)
            input_list_line[self.rpt_clipped_fc_name] = the_output
            arcpy.MakeFeatureLayer_management(input_list_line[self.rpt_aoi_for_clip], "aoi_layer")

            if not arcpy.Exists(the_output):
                try:
                    arcpy.MakeFeatureLayer_management(the_input, "input_layer")
                    arcpy.SelectLayerByLocation_management("input_layer", 'INTERSECT', "aoi_layer")
                    if input_list_line[self.rpt_def_query]: # if there is a definition query field in the master_control_list
                        arcpy.SelectLayerByAttribute_management("input_layer", "SUBSET_SELECTION", input_list_line[self.rpt_def_query])
                    #arcpy.Intersect_analysis("aoi_layer #;input_layer #", the_output)
                    arcpy.Intersect_analysis(["input_layer", "aoi_layer"], the_output)

                    # calculate the label field to be displayed on the maps
                    try:
                        arcpy.CalculateField_management(input_list_line[self.rpt_clipped_fc_name], "label_field", f"!{input_list_line[self.rpt_label_field]}!", "PYTHON3")          # "\"denning\"")
                    except:
                        arcpy.AddMessage(f"Could not populate 'label_field' with {input_list_line[self.rpt_label_field]}")
                except Exception as e:
                    arcpy.AddWarning("Failure occurred: {0}".format(intersect_message))
                    input_list_line[self.rpt_error_flag] = "failed"
                    # arcpy.AddMessage("AOI: " + str(input_list_line[self.rpt_aoi_for_clip]))
                finally:
                    try:
                        arcpy.Delete_management("aoi_layer")
                        arcpy.Delete_management("input_layer")
                    except arcpy.ExecuteError as delete_error:
                        pass
                        # arcpy.AddWarning(f"Failed to delete feature layers: {delete_error}")

    
    
    def create_spreadsheet(self):
        arcpy.AddMessage("======================================================================")
        arcpy.AddMessage('Creating the final output spreadsheet')
   
        self.xls_to_save = os.path.join(self.work_directory,self.xls_for_analysis + "_" + self.featureclass_to_analyize + ".xlsx")

        self.book = openpyxl.Workbook()
        
        # Sheet 1
        self.sheet = self.book.active
        self.sheet.title = "Conflicts & Constraints"
        self.sheet.merge_cells('B1:D1')

        # Sheet 2
        self.sheet2= self.book.create_sheet("Data Sources")

        selection_string_list = []
        # get list of values to create sub_reports on
        if self.create_subreports_on_this_field != "" and self.create_subreports_on_this_field != "#":
            the_aoi = os.path.join(self.work_gdb, "aoi")
            #summary_list = self.make_unique_list_of_field_values(the_aoi,"",self.create_subreports_on_this_field)
            summary_list = self.unique_values(the_aoi, self.reporting_field)

            # create the selection_strings for all the sub_reports_on field values
            # ie. selection_string = "\"Group\" = 'Group A' "
            selection_string_list = []
            for x in summary_list:
                try:
                    if isinstance(x,str):
                        sql_exp = """{0} = '{1}'""".format(arcpy.AddFieldDelimiters(the_aoi, self.reporting_field), x)
                    else:
                        sql_exp = """{0} = {1}""".format(arcpy.AddFieldDelimiters(the_aoi, self.reporting_field), x)
                    selection_string_list.append(sql_exp)
                except Exception as e:
                    arcpy.AddWarning("Subreports could not be created with the input field.")
                    arcpy.AddError(e)
                    sys.exit()


                #selection_string_list.append('"' + self.reporting_field + '"  = ' + x +  "'")
        


        # if all the output is on one sheet
        if self.subreports_on_seperate_sheets != 'true':
            self.newline = 1  # newline is the row on the spreadsheet that is currently being written
            self.create_header_information()
            
            #if self.subreports_on_seperate_sheets == "" or self.subreports_on_seperate_sheets == "#": 
            self.write_report_header()
            count = 0
            if self.create_subreports_on_this_field != "" : # Create report details for each sub_group
                for selection_string in selection_string_list:
                    arcpy.AddMessage("Creating report for: " + selection_string)
                    self.make_excel_details(selection_string, count)# pass in the selection string
                    count =+1
            else: # create report details for the whole data clips
                selection_string = ''
                self.make_excel_details(selection_string, count)

            #self.data_source_details_on_report()
            self.data_source_details_on_report_newsheet()
            

        
        # if each sub report is on a separate sheet
        if self.subreports_on_seperate_sheets == "true":
            self.create_header_information()
            self.this_workbook = self.workbook

            # adds a blank sheet for every each sub-group
            self.this_workbook.Worksheets[2].Delete()
            self.this_workbook.Worksheets[1].Delete()
            for selection_string in selection_string_list:
                self.this_workbook.Worksheets.Add()


            y = 0
            for selection_string in selection_string_list: # loop through all the sub-groups
                
                # strip just the sup_group value so you can name the spreadsheets with it.
                sel_string = selection_string.split("=")
                sheet_name = sel_string[1]
                sheet_name = sheet_name.replace("'", "")
                
                
                self.this_sheet = self.workbook.Worksheets[y]
                self.this_workbook.Worksheets[y].Name = sheet_name
                self.set_xls_column_widths(y)
                self.newline = 1
                self.write_report_header()
                self.make_excel_details(selection_string, count = 0)# pass in the selection string
                y += 1
            
            #self.data_source_details_on_report()
            self.data_source_details_on_report_newsheet()


    def unique_values(self, table, field):
        with arcpy.da.SearchCursor(table, [field]) as cursor:
            return sorted({row[0] for row in cursor})

       
    def make_unique_list_of_field_values(self,fc_name,selection_query,fld1='',fld2='',fld3='',fld4='',fld5='',fld6=''):
        '''
        This returns a list of all the unique values in the feature class fields.
        
        @return: A list 
        @rtype: string list
        '''
    
        field1_exists = False
        field2_exists = False
        field3_exists = False
        field4_exists = False
        field5_exists = False
        field6_exists = False
         
        # determine if the fields exist 
        if arcpy.Exists(fc_name):
            fieldList = arcpy.ListFields(fc_name)
            for field in fieldList:
                fld_name = field.name
                fld_type = field.type

                if fld1 == fld_name:
                    field1_exists = True
                    field1_type = fld_type
                if fld2 == fld_name:
                    field2_exists = True
                    field2_type = fld_type
                if fld3 == fld_name:
                    field3_exists = True
                    field3_type = fld_type
                if fld4 == fld_name:
                    field4_exists = True
                    field4_type = fld_type
                if fld5 == fld_name:
                    field5_exists = True
                    field5_type = fld_type
                if fld6 == fld_name:
                    field6_exists = True
                    field6_type = fld_type
        
        
        uniqueList = []
        
        # add all the field values to a list.  The list is joined into one big
        # string and then that string append to the unique list that will be returned.
        if arcpy.Exists(fc_name) and int(str(arcpy.GetCount_management(fc_name))) > 0:
            arcpy.MakeFeatureLayer_management(fc_name, "fc_layer")
            if selection_query != "": # if there is a definition query field in the master_control_list
                arcpy.SelectLayerByAttribute_management("fc_layer", "NEW_SELECTION", selection_query)

            all_rows = arcpy.SearchCursor("fc_layer")
            for row in all_rows:
                field_values_list = []
                if field1_exists:
                    try:
                            
                        field_value = row.getValue(fld1)
                        if not field_value:
                            field_value = 'blank field'
                        field_value = "'" + str(field_value)
                        if field_value[-2:] == '.0': # trim of .0 from strings
                            field_value = field_value[:-2]
                        field_values_list.append(field_value) 
                    except:
                        field_values_list.append("!!! potential error - check nation name is displayed correctly !!! ")

                
                if field2_exists:
                    try:
                            
                        field_value = (str(row.getValue(fld2)))
                        if not field_value:
                            field_value = 'blank field'
                        field_value = str(field_value)
                        field_values_list.append(field_value)
                    except:
                        field_values_list.append("!!! potential error - check nation name is displayed correctly !!! ")
                        
                        
                if field3_exists:
                    try:

                        field_value = (str(row.getValue(fld3)))
                        if not field_value:
                            field_value = 'blank field'
                        field_value = str(field_value)
                        field_values_list.append(field_value) 
                    except:
                        field_values_list.append("!!! potential error - check nation name is displayed correctly !!! ")

                
                if field4_exists:
                    try:
                            
                        field_value = (str(row.getValue(fld4)))
                        if not field_value:
                            field_value = 'blank field'
                        field_value = str(field_value)
                        field_values_list.append(field_value) 
                    except:
                        field_values_list.append("!!! potential error - check nation name is displayed correctly !!! ")

                if field5_exists:
                    try:
                        field_value = (str(row.getValue(fld5)))
                        if not field_value:
                            field_value = 'blank field'
                        field_value = str(field_value)
                        field_values_list.append(field_value)
                    except:
                        field_values_list.append("!!! potential error - check nation name is displayed correctly !!! ")

                if field6_exists:
                    try:
                        field_value = (str(row.getValue(fld6)))
                        if not field_value:
                            field_value = 'blank field'
                        field_value = str(field_value)
                        field_values_list.append(field_value) 
                    except:
                        field_values_list.append("!!! potential error - check nation name is displayed correctly !!! ")


                result_string_list = (";".join(field_values_list))
                if result_string_list == 'None':
                    result_string_list = "blank field"

                uniqueList.append(result_string_list)

        uniqueList = list(set(uniqueList))  # makes the list unique
        uniqueList.sort() # sort the list
        
        return  uniqueList 


    def create_header_information(self):
        ''' This creates the header information that will be printed at the 
            top of the output spreadsheet.  It consists of the input_xls file name,
            the date, the path to the input shape being analyized, and any of the 
            header fields that have been passed in to this tool.
        ''' 
        #To differentiate between FCBC and MOF headers.
        #MOF Header
        header = []
        if self.run_as_fcbc != 'true':
            header.append(self.xls_for_analysis)  # the name of the xls that has the input criteria
            header.append("Date: " + str(datetime.date.today()))
            header.append("input:  " + self.analyize_this_featureclass)
            xls_to_save = os.path.join(self.work_directory,self.xls_for_analysis + "_" + self.featureclass_to_analyize + ".xlsx")
            header.append("output:  " + xls_to_save)
        
        #FCBC Header
        else:
            header_titles = {   "Date: ":datetime.date.today(),
                                "Region: ":self.region,
                                "Crown File Number: ":str(self.crown_file_number),
                                "Disposition Number: ":str(self.disposition_number),
                                "Parcel Number: ":str(self.parcel_number),
                                "Geomark: ":"""=IF('Crown Land Status'!B28=0,"",'Crown Land Status'!B28)"""}
            
            self.header_titles = header_titles               
                
                        
        if self.report_header1 != "#" and self.report_header1 != "":
            header.append (self.report_header1)
        if self.report_header2 != "#" and self.report_header2 != "":
            header.append (self.report_header2)
        if self.report_header3 != "#" and self.report_header3 != "":
            header.append (self.report_header3)
        if self.report_header4 != "#" and self.report_header4 != "":
            header.append (self.report_header4)
        
        self.report_header = header

 
    def write_report_header(self):
        ''' This writes the header information on the spreadsheet
        ''' 

        #if self.disclaimer != "" and self.disclaimer != "#":
        my_cell = 'B' + str(self.newline)
        self.sheet[my_cell] = self.disclaimer
        self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[2],size=12,name='Arial')
        self.newline += 1
        
                  
        #Write header information for MFLNRORD 
        if self.run_as_fcbc != 'true':
            self.newline2 = self.newline + 1
            y = 0
            for x in self.report_header:
                y += 1
                if y == 1:
                    my_cell = 'B' + str(self.newline)
                    self.sheet.merge_cells(f'B{str(self.newline)}:C{str(self.newline)}')
                    self.sheet[my_cell] = x
                    self.sheet[my_cell].font = Font(color=colors.COLOR_INDEX[21],size=12,name='Arial')
                    self.sheet[my_cell].alignment = Alignment(wrapText=True)
                    self.newline += 1
                else:
                    my_cell = 'B' + str(self.newline)
                    self.sheet.merge_cells(f'B{str(self.newline)}:C{str(self.newline)}')
                    self.sheet[my_cell] = x
                    self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[8],size=10,name='Arial')
                    self.sheet[my_cell].alignment = Alignment(wrapText=True)
                    self.newline += 1
            ##self.newline += 1
            
         
        #Write header information for FCBC
        else:
            #Write the header information that was input in the tool from ArcGIS Pro
            self.newline2 = self.newline
            for x in self.report_header:
                my_cell = 'B' + str(self.newline)
                self.sheet[my_cell] = x
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[8],size=10,name='Arial')
                self.newline += 1
                
            #Iterate through the FCBC header title keys and write the header names
            for key, val in self.header_titles.items():
                my_cell = 'B' + str(self.newline)
                self.sheet[my_cell] = key
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[8],size=10,name='Arial')
                self.sheet[my_cell].alignment = Alignment(horizontal='justify')    
                #Write out the values to the headers
                my_cell = 'C' + str(self.newline)
                self.sheet[my_cell] = val
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[8],size=10,name='Arial')
                self.sheet[my_cell].alignment = Alignment(horizontal='justify')
                self.newline += 1
                
            self.newline += 1
        
        if self.suppress_map_creation != 'true':
            #Write the links for the Overview maps
            for overview_map_name in ["Overview Map 1:50000", "Overview Map 1:100000", "Overview Map 1:300000"]:
                map_html_name = overview_map_name + ".html"
                for find_char, relpace_char in {"1:": "", 
                                                " ": "_", 
                                                "O": "o", 
                                                "M": "m"}.items():
                    map_html_name = map_html_name.replace(find_char, relpace_char)

                map_path = os.path.join(self.map_directory, map_html_name)
                path_resolve = str(Path(map_path).resolve())
                if os.path.exists(map_path):
                    my_cell = 'D' + str(self.newline2)
                    self.sheet[my_cell].hyperlink = path_resolve 
                    self.sheet[my_cell].value = overview_map_name
                    self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[12],size=10,name='Arial')
                self.newline2 += 1

            self.newline += 1

               
    def make_excel_details(self,selection_string, count):
        # selection_string is a definition query to week only those values that pass ie. "Group" = 'A'
        
#         pinkfill = PatternFill(fill_type='solid', start_color='F2DCDB', end_color='F2DCDB')
#         greyfill = PatternFill(fill_type='solid', start_color='E0E0E0', end_color='E0E0E0')
#         greyfill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
        
        
        current_category = "" # set the category to blank so that when it changes it can write the category title on the xls
        items_in_list = len(self.input_datasources_list) # the num of items in master_control list to display in messages
        xxx_count = 0 # just a counter do display in messages

        #  ---NEW---
        # Add headers {header: columnwidth}

        #self.sheet.column_dimensions['A'].width = 30

        newcolumn = 1
        
        #Do not write the column titles that pertain specifically to FCBC
        if self.run_as_fcbc != 'true':
            column_titles = {" ":1.5,
                             "  ":18,
                             "   ":60,
                             "List conflicts": 45,
                             "Map": 15}
        #Do not write the column titles that pertain specifically to FCBC
        else:
            column_titles = {" ":1.5,
                             "  ":18,
                             "   ":60,
                             "List conflicts": 45,
                             "Map": 15,
                             "Manual Entry of Contact information for Referrals": 35,
                             "Comprehensive Review Comments": 35,
                             "Conflict Overlap Snips": 30,
                             "Business Line Comments": 23}
        if count == 0:
            for header, width in column_titles.items():
                self.sheet.cell(self.newline, newcolumn, header).font = Font(size=14, 
                                                                            bold=True,
                                                                            italic=True)
                col_letter = openpyxl.utils.cell.get_column_letter(newcolumn)
                self.sheet.column_dimensions[col_letter].width = width
                self.sheet.cell(self.newline, newcolumn).alignment = Alignment(wrapText=True)
                #self.sheet.alignment = Alignment(wrapText=True)
                newcolumn += 1

                #Freeze Panes
                self.sheet.freeze_panes = "B" + str(self.newline+1)
    
        if selection_string != "":
            self.newline += 1
            my_cell = 'B' + str(self.newline)
            self.sheet[my_cell].value = "_".join(selection_string.split("_", 1)[1:])  #remove the 'report_' portion of reporting field name in AOI for spreadsheet#
            self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[21],size=12,name='Arial')
            ##self.newline += 1
        
        #Set the starting border range
        start_border_range = "B" + str(self.newline + 2)
        
        #Set the ending border range, depending if the spreadsheet is for MFLNORD or FCBC
        if self.run_as_fcbc != 'true':
            ltr_range = ["B", "C", "D", "E"]
            end_border_ltr = "E"
        else:
            ltr_range = ["B", "C", "D", "E", "F", "G", "H", "I"]
            end_border_ltr = "I"
            
        for input_list_line in self.input_datasources_list:
            #------Create a message with the number of iterations to track progress  ----------------------- 
            xxx_count += 1
            row_message = f"    writing spreadsheet row {str(xxx_count)} of {str( items_in_list)}      {input_list_line[self.rpt_data_name]}"
            arcpy.AddMessage(row_message)

            #---------Write the category in bigger text
            if input_list_line[self.rpt_category] != current_category:
                end_border_range = end_border_ltr + str(self.newline)
                if current_category != "":
                    self.newline += 1
                    range_string = start_border_range + ":" + end_border_range 
                    set_border(self.sheet, range_string)
                    self.newline += 1
                start_border_range = "B" + str(self.newline + 1)
                self.block_color = 0 # set the default color for each category block to white
                self.newline +=1

                
                my_cell = 'B' + str(self.newline)
                self.sheet[my_cell].value = input_list_line[self.rpt_category]
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[0],size=14,name='Arial',bold=True)
#                 fill_hex = 'FFFFFF'
                current_category = input_list_line[self.rpt_category]
            
            #--------write the feature class name
#             if fill_hex == 'E0E0E0':
#                 fill_hex = 'FFFFFF'
#             else:
#                 fill_hex = 'E0E0E0'
            fill_hex = 'E0E0E0'
            tmp_fill = PatternFill(fill_type='solid', start_color=fill_hex, end_color=fill_hex)
            
            self.newline += 1
            my_cell = 'C' + str(self.newline)
            
            #Set the restricted Arch layer input feature class names
            restrict_list = ['WHSE_ARCHAEOLOGY.RAAD_AOA_PROVINCIAL',
                             'WHSE_ARCHAEOLOGY.RAAD_INFORMED_CONTRIBUTORS_SV',
                             'WHSE_ARCHAEOLOGY.RAAD_TFM_SITES_SVW']
            
            #check the source and if it is a restricted layer write the disclaimer indicating this.
            source_fc = os.path.basename(input_list_line[self.rpt_data_source])
            if source_fc in restrict_list:
                restrictVal = "https://www2.gov.bc.ca/assets/gov/farming-natural-resources-and-industry/natural-resource-use/archaeology/forms-publications/archaeological_information_sharing_agreement.pdf"
                self.sheet[my_cell].hyperlink = restrictVal
                self.sheet[my_cell].value = "Archaeological Information Sharing Agreement (gov.bc.ca)"
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[53],size=8,name='Arial',bold=True)
                self.sheet[my_cell].alignment = Alignment(wrapText=True)
                self.newline += 1


            my_cell = 'C' + str(self.newline)       
            self.sheet[my_cell].value = input_list_line[self.rpt_data_name]
            self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[8],size=10,name='Arial',bold=True)
            self.sheet[my_cell].alignment = Alignment(wrapText=True)
            
            # -- New 
            for col_ltr in ltr_range:
                self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
#                 if fill_hex == 'E0E0E0':
#                     self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
            # -- End New (old code commented out below)
            #if fill_hex == 'E0E0E0':
            #    self.sheet[my_cell].fill = tmp_fill

            #--------write the summary fields, or "failed", or "overlaps with this value"
            summary_list = [] # the list to hold the summary values to be printed on the xls
            num_of_recs = 0  #how many records in the FC
            if arcpy.Exists(input_list_line[self.rpt_clipped_fc_name]):
                try:
                    arcpy.MakeFeatureLayer_management(input_list_line[self.rpt_clipped_fc_name], "fc_layer")
                    if selection_string != "": # if a selection_string was passed in   ie. "Group" = 'A'
                        arcpy.SelectLayerByAttribute_management("fc_layer", "NEW_SELECTION", selection_string)
                    num_of_recs = int(str(arcpy.GetCount_management("fc_layer"))) #how many records in the FC
                    arcpy.Delete_management("fc_layer")
                except:
                    if not input_list_line[self.rpt_error_flag] == "failed":
                        input_list_line[self.rpt_error_flag] = "failed"

            
            f1 = input_list_line[self.rpt_fld_to_summarize1]#Set shorter variable names
            f2 = input_list_line[self.rpt_fld_to_summarize2]
            f3 = input_list_line[self.rpt_fld_to_summarize3]
            f4 = input_list_line[self.rpt_fld_to_summarize4]
            f5 = input_list_line[self.rpt_fld_to_summarize5]
            f6 = input_list_line[self.rpt_fld_to_summarize6]
            
            if input_list_line[self.rpt_error_flag] == "failed":  # this flag set if the clip features part failed
                summary_list = "failed"
            elif f1 == "" and f2 == "" and f3 == "" and f4 == "" and f5 == "" and f6 ==  "" and num_of_recs > 0 : #if all fields to summarize are blank
                summary_list = "overlaps with this value"
            elif num_of_recs > 0: # make the unique list if not all fields to summarize are blank
                summary_list = self.make_unique_list_of_field_values(input_list_line[self.rpt_clipped_fc_name],selection_string,f1,f2,f3,f4,f5,f6)
                if len(summary_list) == 0:
                    summary_list = "overlaps with this value"
            elif num_of_recs == 0 :
                summary_list = "No data to display"

            # write the summary field values on the xls
            map_name  = os.path.split(input_list_line[self.rpt_clipped_fc_name])[1] #what the map was saved as
            map_path = os.path.join(self.map_directory,map_name+ ".html")
            path_resolve = str(Path(map_path).resolve())
            if os.path.exists(map_path):
                my_cell = 'E' + str(self.newline)
                self.sheet[my_cell].hyperlink = path_resolve 
                self.sheet[my_cell].value= "View Map"
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[12],size=10,name='Arial')


            ##start_color_range = "B" + str(self.newline)
            if summary_list == "failed":
                my_cell = 'D' + str(self.newline)
                self.sheet[my_cell].value = "Unsuccessful"
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[2],size=8,name='Arial',bold=True)
                self.sheet[my_cell].alignment = Alignment(wrapText=True)
                for col_ltr in ltr_range:
                    self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
#                     if fill_hex == 'E0E0E0':
#                         self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
                self.newline += 1
            elif summary_list == "No data to display":
                my_cell = 'D' + str(self.newline)
                self.sheet[my_cell].value = summary_list
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[0],size=8,name='Arial',bold=True)
                self.sheet[my_cell].alignment = Alignment(wrapText=True)
                for col_ltr in ltr_range:
                    self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
#                     if fill_hex == 'E0E0E0':
#                         self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
                self.newline += 1
            elif summary_list == "overlaps with this value":
                my_cell = 'D' + str(self.newline)
                self.sheet[my_cell].value = summary_list
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[4],size=8,name='Arial')
                self.sheet[my_cell].alignment = Alignment(wrapText=True)
                for col_ltr in ltr_range:
                    self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
#                     if fill_hex == 'E0E0E0':
#                         self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
                self.newline += 1
            else:
                for x in summary_list:
                    my_cell = 'D' + str(self.newline)
                    self.sheet[my_cell].value = x
                    self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[4],size=8,name='Arial')
                    self.sheet[my_cell].alignment = Alignment(wrapText=True)
                    for col_ltr in ltr_range:
                        self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
#                         if fill_hex == 'E0E0E0':
#                             self.sheet[col_ltr + str(self.newline)].fill = tmp_fill
                    self.newline += 1


        end_border_range = end_border_ltr + str(self.newline)
        range_string = start_border_range + ":" + end_border_range 
        set_border(self.sheet, range_string)
        self.newline +=1
        
        if self.run_as_fcbc == 'true':
            fcbc_footer = []
            fcbc_footer.append("input:  " + self.analyize_this_featureclass)
            xls_to_save = os.path.join(self.work_directory,self.xls_for_analysis + "_" + self.featureclass_to_analyize + ".xlsx")
            fcbc_footer.append("output:  " + xls_to_save)
            
            for x in fcbc_footer:
                my_cell = 'C' + str(self.newline)
                self.sheet[my_cell] = x
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[8],size=10,name='Arial')
                self.newline += 1
        self.newline +=1
        self.book.save(self.xls_to_save)

 
    def data_source_details_on_report(self):
        ''' This creates the cells with all the data sources in them.
        ''' 
        self.newline += 1
        new_line2 = 1

        my_cell = 'B' + str(self.newline)
        self.sheet[my_cell].value = "Data Sources"
        self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[0],size=10,name='Arial')
        #self.sheet[my_cell].alignment = Alignment(wrapText=True)
        
        my_cell = 'D' + str(self.newline)
        self.sheet[my_cell].value = "Definition Query"
        self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[0],size=10,name='Arial')
        #self.sheet[my_cell].alignment = Alignment(wrapText=True)
        
        self.newline += 1

        
        for input_list_line in self.input_datasources_list:
            data_source = input_list_line[self.rpt_data_source]
    
            # check to see if input is local dataset or from the LRDW
            #LOCAL DATA
            if (data_source[0:2]== "r'\\'" or data_source[0:1] == "\\" ):
                #input_list_line[self.rpt_clipped_fc_name]= 'local_data'
                my_cell = 'C' + str(self.newline)
                self.sheet[my_cell].value = data_source
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[48],size=7,name='Arial')
                self.sheet[my_cell].alignment = Alignment(wrapText=True)

                
            #BCGW DATA SOURCES
            else:
                my_cell = 'C' + str(self.newline)
                self.sheet[my_cell].value = data_source
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[19],size=7,name='Arial')
                self.sheet[my_cell].alignment = Alignment(wrapText=True)
                
            #write the definition query that pertains to the data source in a new column
            rpt_def_query = input_list_line[self.rpt_def_query]
            if rpt_def_query:
                my_cell = 'D' + str(self.newline)
                self.sheet[my_cell].value = rpt_def_query
                self.sheet[my_cell].font =  Font(color=colors.COLOR_INDEX[61],size=6,name='Arial')
                self.sheet[my_cell].alignment = Alignment(wrapText=True)
            self.newline += 1
        
        self.book.save(self.xls_to_save)


    def data_source_details_on_report_newsheet(self):
        ''' This creates the cells with all the data sources in them.
        ''' 
        newline = 2

        # Spreadsheets used for datasources
        my_cell = 'A' + str(newline)
        self.sheet2[my_cell].value = "Input Spreadsheet(s):"
        self.sheet2[my_cell].font =  Font(color=colors.COLOR_INDEX[0],
                                            bold=True,
                                            size=10,
                                            name='Arial')
        newline += 1
        for sheet in [self.xls_file_for_analysis_input, self.xls_file_for_analysis_input2]:
            if sheet != "":
                my_cell = 'A' + str(newline)
                self.sheet2[my_cell].value = sheet
                self.sheet2[my_cell].font =  Font(color=colors.COLOR_INDEX[0],
                                            bold=False,
                                            size=7,
                                            name='Arial')
                newline += 1
        newline += 2

        # Fields Names - Datasources and Queries
        my_cell = 'A' + str(newline)
        self.sheet2[my_cell].value = "Data Sources"
        self.sheet2[my_cell].font =  Font(color=colors.COLOR_INDEX[0],
                                            bold=True,
                                            size=10,
                                            name='Arial')
        self.sheet2[my_cell].alignment = Alignment(wrapText=True)
        
        my_cell = 'B' + str(newline)
        self.sheet2[my_cell].value = "Definition Query"
        self.sheet2[my_cell].font =  Font(color=colors.COLOR_INDEX[0],
                                            bold=True,
                                            size=10,
                                            name='Arial')
        self.sheet2[my_cell].alignment = Alignment(wrapText=True)
        
        # Adjust Column Width
        for col_letter, width in {'A': 55, 'B': 30}.items():
            self.sheet2.column_dimensions[col_letter].width = width

        newline += 1

        # Read input datasource list
        for input_list_line in self.input_datasources_list:
            data_source = input_list_line[self.rpt_data_source]
            rpt_def_query = input_list_line[self.rpt_def_query]
            #arcpy.AddMessage(f'{data_source}:  {rpt_def_query}')

            # check to see if input is local dataset or from the LRDW
            #LOCAL DATA
            if (data_source[0:2]== "r'\\'" or data_source[0:1] == "\\" ):  color_i = 48
            #BCGW DATA SOURCES
            else:  color_i = 19

            # Write value to cell
            my_cell = 'A' + str(newline)
            self.sheet2[my_cell].value = data_source
            self.sheet2[my_cell].font =  Font(color=colors.COLOR_INDEX[color_i],
                                                size=7,
                                                name='Arial')
            self.sheet2[my_cell].alignment = Alignment(wrapText=True)
                
            #write the definition query that pertains to the data source in a new column
            my_cell = 'B' + str(newline)
            self.sheet2[my_cell].value = rpt_def_query
            self.sheet2[my_cell].font =  Font(color=colors.COLOR_INDEX[61],
                                                size=6,
                                                name='Arial')
            self.sheet2[my_cell].alignment = Alignment(wrapText=True)

            newline += 1

            # merge source Excel cells
            for i in [3,4]:
                self.sheet2.merge_cells(f'A{str(i)}:B{str(i)}')
        
        self.book.save(self.xls_to_save)

    
    def make_html_maps(self):
        html = HTMLGenerator(self.xls_file_for_analysis_input, self.xls_file_for_analysis_input2, self.work_gdb, self.map_directory)
        html.generate_html_maps()


    # def make_the_maps_with_labels(self):
    #     global map_name
    #     arcpy.AddMessage("======================================================================")
    #     arcpy.AddMessage("Creating the maps")
    #     arcpy.AddMessage("======================================================================")

    
    #     # set the map, layout, and map frame variables
    #     aprx = self.aprx
    #     aprx.importDocument(r"\\giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\map_files\Status_Layout.pagx")
    #     revolt_map = aprx.listMaps("Status Map")[0]
    #     revolt_lyt = aprx.listLayouts("Status Layout")[0]
    #     mf_revolt = revolt_lyt.listElements("mapframe_element", "Revolt Frame")[0]
    #     layer_files = r"\\giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\lyrx_files"

        
    #     # set the data source for the AOI and zoom to it.
    #     aoi_layer = revolt_map.listLayers("Area_of_Interest")[0]
    #     connprop_aoi = aoi_layer.connectionProperties
    #     connprop_aoi['connection_info']['database'] = self.work_gdb
    #     aoi_layer.updateConnectionProperties(aoi_layer.connectionProperties, connprop_aoi)
        
     
    #     items_in_list = len(self.input_datasources_list) # the total number of items in the master control list
    #     xxx_count = 0 # a counter to display in messages
        
    #     for input_list_line in self.input_datasources_list:     # this will hold the values from the input xls file 
    #         xxx_count += 1
    #         name_with_spaces_replaced = os.path.join(self.work_gdb, input_list_line[self.rpt_data_name].replace (" ", "_"))
    #         map_name  = os.path.split(name_with_spaces_replaced)[1]  #what the map will be saved as
    #         self.map_path = os.path.join(self.map_directory,map_name + ".pdf") 
    
    
    #         the_clipped_fc = os.path.join(self.work_gdb,map_name)
    #         map_message = "Making map " + str(xxx_count) + " of " + str(items_in_list) + "  " +  map_name + ".pdf"
    #         arcpy.AddMessage(map_message)
            
            
    #         # iterate through each of the feature classes in the working gdb directory and
    #         # get the number of features that are contained in that feature class
    #         if arcpy.Exists(the_clipped_fc):
    #             print ("the_clipped_fc ", the_clipped_fc)
    #             count_of_recs_in_fc = int(str(arcpy.GetCount_management(the_clipped_fc)))
                
    #             # check to see if the count of features in the feature class is greater than 0.
    #             if count_of_recs_in_fc > 0:
    #                 desc = arcpy.Describe(the_clipped_fc)
    #                 input_data_type = desc.ShapeType

                    
    #                 # check if the input is a point or multipoint feature class. Make single part, if necessary
    #                 if input_data_type.upper() in ("POINT", "MULTIPOINT"):
    #                     if input_data_type.upper() == ("MULTIPOINT"):
    #                         arcpy.MultipartToSinglepart_management(the_clipped_fc, the_clipped_fc + "_singlepart")
    #                         arcpy.Delete_management(the_clipped_fc)
    #                         arcpy.Rename_management(the_clipped_fc + "_singlepart", the_clipped_fc)
                    
    #                 # add the excel group layer to the map. This will be used to store all layers of each of the 
    #                 # feature classes that appear under each category in the excel file.
    #                 category = input_list_line[self.rpt_category]
    #                 group = revolt_map.listLayers(category)

                    
    #                 # check if the category group exists. If not, create one and rename it.
    #                 if not group:
    #                     insert_layer = arcpy.mp.LayerFile(os.path.join(layer_files, "Revolt_Group.lyrx"))
    #                     revolt_map.addLayer(insert_layer, "TOP")
    #                     group = revolt_map.listLayers("Group")[0]
    #                     group.name = category
    #                     group.visible = True

    #                 else:
    #                     group = revolt_map.listLayers(category)[0]
                        
                        
    #                 # add clipped data from working gdb file using the layer file. Update the connection properties
    #                 overlap_name = os.path.join(layer_files, f"Revolt_Overlapping_{input_data_type}.lyrx")
    #                 insert_layer = arcpy.mp.LayerFile(overlap_name)
    #                 revolt_map.addLayerToGroup(group, insert_layer,"BOTTOM")
    #                 for overlap in revolt_map.listLayers("Overlapping*"):
    #                     if overlap.longName == f"{category}\Overlapping_Features":
    #                         overlap.name = map_name + "_Overlaps"
    #                         old_source = overlap.connectionProperties
    #                         base_fc = os.path.basename(name_with_spaces_replaced)
    #                         new_source = {'dataset': base_fc,
    #                                         'workspace_factory': 'File Geodatabase',
    #                                         'connection_info': {'database': self.work_gdb}}
    #                         overlap.updateConnectionProperties(old_source, new_source)
    #                         overlap.visible = True
    #                     if overlap.supports("SHOWLABELS"):
    #                         lblClass = overlap.listLabelClasses("Default")[0]
    #                         lblClass.expression = "[" + input_list_line[self.rpt_label_field] + "]"
    #                         lblClass.visible = True


                        
    #                 # Determine the source data type to reference the proper layer file.
    #                 insert_source = input_list_line[self.rpt_data_source]
    #                 in_desc = arcpy.Describe(insert_source)
    #                 data_type = in_desc.dataType
    #                 if data_type == "ShapeFile":
    #                     ref_layer = os.path.join(layer_files, f"Revolt_All_{input_data_type}_SHP.lyrx")
    #                 elif data_type == "FeatureClass":
    #                     if in_desc.geometryStorage == "SDO":
    #                         ref_layer = os.path.join(layer_files, f"Revolt_All_{input_data_type}_SDE.lyrx")
    #                     else:
    #                         ref_layer = os.path.join(layer_files, f"Revolt_All_{input_data_type}_GDB.lyrx")

                    
    #                 #Add the reference layer to the appropriate group category
    #                 insert_layer = arcpy.mp.LayerFile(ref_layer)
    #                 revolt_map.addLayerToGroup(group, insert_layer,"BOTTOM")
    #                 for feature in revolt_map.listLayers("All_Features"):
    #                     def_query = input_list_line[self.rpt_def_query]
    #                     if feature.longName == f"{category}\All_Features":
    #                         feature.visible = True
    #                         feature.name = map_name + "_All_Features"
    #                         old_con_prop = feature.connectionProperties
                            
                            
    #                         # Get the path and feature class name from the described data.
    #                         path_to_data = in_desc.path
    #                         base_of_data = in_desc.baseName
                            
    #                         #Update the shapefile connection properties
    #                         if data_type == "ShapeFile":
    #                             data_extension = "." + str(in_desc.extension)
    #                             shape_data = base_of_data + data_extension
    #                             newConnPropDict = {'connection_info': {'database': path_to_data},
    #                                                     'dataset': shape_data,
    #                                                     'workspace_factory': 'Shape File'}
    #                             feature.updateConnectionProperties(old_con_prop, newConnPropDict)
                                
    #                         #Determine whether the connection is from an SDE or FGDB and update
    #                         #the feature class connection properties.
    #                         elif data_type == "FeatureClass":
    #                             if in_desc.geometryStorage == "SDO":
    #                                 newConnPropDict = {'connection_info': {'database': self.sde_connection},
    #                                                     'dataset': base_of_data}
    #                                 feature.updateConnectionProperties(old_con_prop, newConnPropDict)
    #                             else:
    #                                 newConnPropDict = {'connection_info': {'database': path_to_data},
    #                                                     'dataset': base_of_data,
    #                                                     'workspace_factory': 'File Geodatabase'}
    #                                 feature.updateConnectionProperties(old_con_prop, newConnPropDict)
                                    
    #                         # If a query exists in the spreadsheet, set the definition query for the base data.                   
    #                         if def_query:
    #                             if feature.supports("DEFINITIONQUERY"):
    #                                 feature.definitionQuery = def_query


    #                 #Create the PDF file if it does not exist
    #                 if not os.path.isfile(self.map_path):
    #                     #capture extent of the aoi used to clip the input feature class
    #                     #and zoom to that layer's extent.
    #                     cbpath = input_list_line[12]
    #                     desc = arcpy.Describe(cbpath)
    #                     aoi_extent = desc.extent
    #                     mf_revolt.camera.setExtent(aoi_extent)

                        
    #                     # round up the scale and set the scale for the map frame
    #                     cbScale = mf_revolt.camera.scale
    #                     if cbScale < 10000:
    #                         scale = 10000
    #                     elif cbScale >= 10000 and cbScale < 20000:
    #                         scale = 20000
    #                     elif cbScale >= 20000 and cbScale < 50000:
    #                         scale = 50000
    #                     elif cbScale >= 50000 and cbScale < 75000:
    #                         scale = 75000
    #                     elif cbScale >= 75000 and cbScale < 125000:
    #                         scale = 125000    
    #                     elif cbScale >= 125000 and cbScale < 200000:
    #                         scale = 200000  
    #                     elif cbScale >= 200000 and cbScale < 500000:
    #                         scale = 500000
    #                     elif cbScale >= 500000 and cbScale < 1000000:
    #                         scale = 1000000
    #                     elif cbScale >= 1000000 and cbScale < 5000000:
    #                         scale = 5000000
    #                     else:
    #                         scale = 10000000
                            
                            
    #                     mf_revolt.camera.scale = scale
    #                     ##aprx.save()

    #                     # split the path into multiple lines of text to put on the map       
    #                     part1 = self.map_path[0:40]
    #                     part2 = self.map_path[40:80]
    #                     part3 = self.map_path[80:120]
    #                     part4 = self.map_path[120:160]            
                        
    #                     # change the text fields on the map to display the path of the pdf's
    #                     text_list = revolt_lyt.listElements("TEXT_ELEMENT")
    #                     for y in text_list:
    #                         if y.name == 'Title':
    #                             y.text = input_list_line[self.rpt_data_name]
    #                         elif y.name == 'Date':
    #                             y.text = str(datetime.date.today())
    #                         elif y.name == 'map_path_part1':
    #                             y.text = part1
    #                         elif y.name == 'map_path_part2':
    #                             y.text = part2
    #                         elif y.name == 'map_path_part3':
    #                             y.text = part3
    #                         elif y.name == 'map_path_part4':
    #                             if part4 != "":
    #                                 y.text = part4

    #                     #create PDF
    #                     revolt_lyt.exportToPDF(self.map_path,resolution=90)
                        
    #                 #turn off the group that contains the active layers exported to pdf
    #                 lyr = revolt_map.listLayers(category)[0]
    #                 if lyr.isGroupLayer:
    #                     off_list = lyr.listLayers()
    #                     for l in off_list:
    #                         if l.visible == True:
    #                             l.visible = False
                

    #             # arcpy.AddMessage("======================================================================")
    #         else:
    #             print("No data exists in the clipped feature class")
    #             # arcpy.AddMessage("======================================================================")
        

    #     print("All Maps Exported!")

    #     #create mapx files and add to current aprx file, if specified
    #     create_mapx_files(self, revolt_map, self.add_maps_to_current)



    # def delete_project(self):
    #     #delete the aprx files where the status and overview maps are created
    #     try:
    #         os.remove(self.aprx.filePath)
    #     except Exception as e:
    #         pass
    #         #arcpy.AddWarning(f"Could not delete the APRX file: " + str(e))
        

##########################################################################################################
##########################################################################################################
##########################################################################################################        
            
#     def make_overview_maps(self):
#         global map_name
#         arcpy.AddMessage("======================================================================")
#         arcpy.AddMessage("Creating the overview maps")
#         arcpy.AddMessage("======================================================================")
    
#         # set the project file and default geodatabase that the project will access.
#         aprx = self.aprx
#         aprx.importDocument(r"\\giswhse.env.gov.bc.ca\whse_np\corp\script_whse\python\Utility_Misc\Ready\statusing_tools_arcpro\map_files\Overview_Layout.pagx")
#         map_name = "the_overview_map"
#         self.map_path = os.path.join(self.map_directory,map_name+ ".pdf") 
    
#         # set the aoi feature class in the default geodatabase.
#         map_message = "Making overview maps"
#         arcpy.AddMessage(map_message) 
    
#         # set the map to "Revolt Overview" and set the layout to "Revolt Overview Layout". Then set the map
#         # frame to the "Revolt Overview Frame".
#         overview_map = aprx.listMaps("Overview Map")[0]
#         lyt = aprx.listLayouts("Overview Layout")[0]
#         mf_overview = lyt.listElements("mapframe_element", "Revolt Overview Frame")[0]
        
#         # set the data source for the AOI
#         lyr = overview_map.listLayers("Area_of_Interest")[0]
#         connprop_aoi = lyr.connectionProperties
#         connprop_aoi['connection_info']['database'] = self.work_gdb
#         lyr.updateConnectionProperties(lyr.connectionProperties, connprop_aoi)
    
#         # zoom to the extent of the AOI layer
#         layer = overview_map.listLayers("Area_of_Interest")[0]
#         mf_overview.camera.setExtent(mf_overview.getLayerExtent(layer, True, True))

        
#         # split the path into multiple lines of text to put on the map
#         part1 = self.map_path[0:40]
#         part2 = self.map_path[40:80]
#         part3 = self.map_path[80:120]
#         part4 = self.map_path[120:160]
        
#         # change the text fields on the map to display the path of the pdf's
#         text_list = lyt.listElements("TEXT_ELEMENT")
#         for y in text_list:
#             if y.name == 'Title':
#                 y.text = map_name
#             elif y.name == 'Date':
#                 y.text = str(datetime.date.today())
#             elif y.name == 'map_path_part1':
#                 y.text = part1
#             elif y.name == 'map_path_part2':
#                 y.text = part2
#             elif y.name == 'map_path_part3':
#                 y.text = part3
#             elif y.name == 'map_path_part4':
#                 if part4 != "":
#                     y.text = part4
                
#         # export the 3 overview maps to pdf.
#         output_dir = os.path.split(self.map_path)[0]
#         scale_list = [300000, 100000, 50000]
#         for scale in scale_list:
#             output_map = os.path.join(output_dir,"overview_map_" + str(scale) + ".pdf")
#             if not arcpy.Exists(output_map):
#                 mf_overview.camera.scale = scale
#                 lyt.exportToPDF(output_map,resolution=90)
#             else:
#                 print("Map Already Exists")
                
#         #save the project
#         aprx.saveACopy(aprx.filePath + "_overviews")

#         #create mapx files and add to current aprx file, if specified
#         create_mapx_files(self, overview_map, self.add_maps_to_current)
        
#         arcpy.AddMessage("======================================================================")
# ##########################################################################################################
# ##########################################################################################################
# ##########################################################################################################       
        
# def create_mapx_files(self, map_input, add_maps):
#     '''create MAPX files of the input maps
#     '''
#     map_name = map_input.name
#     mapx_name = map_name.replace(" ", "_")
#     out_mapx = os.path.join(self.mapx_directory, mapx_name) + ".mapx"
#     map_input.exportToMAPX(out_mapx)
#     arcpy.AddMessage(f"MAPX file created for: {mapx_name}")
    
#     #if user wants maps added to current project, add the MAPX files to it
#     if add_maps == 'true':
#         try:
#             aprx = arcpy.mp.ArcGISProject('CURRENT')
#             aprx.importDocument(out_mapx)
#             arcpy.AddMessage(f"Adding the following map to the current project: {map_name}")
#             ##aprx.save()
#         except Exception as e:
#             arcpy.AddMessage(e)
#     else:
#         arcpy.AddMessage(f"{map_name} was NOT added to the project")
##########################################################################################################
##########################################################################################################
##########################################################################################################

def create_gdb_if_needed(gdb_to_create):
    '''
    This creates a geodatabase in the provided path
    
    @param gdb_to_create: the complete path where the GDB will be created
    @type gdb_to_create: string
    '''
    if not arcpy.Exists(gdb_to_create):
        folder, file = os.path.split(gdb_to_create)
        arcpy.CreateFileGDB_management(folder, file) 
##########################################################################################################
##########################################################################################################
##########################################################################################################

def create_feature_dataset_if_needed(fds_to_create):
    '''
    This creates a feature dataset in the provided path.  If the file geodatabase
    does not exist it will be created.
    
    @param fds_to_create: the complete path where the FDS will be created
    @type fds_to_create: string
    '''
    gdb, feature_dataset = os.path.split(fds_to_create)
    create_gdb_if_needed(gdb)
    if not arcpy.Exists(fds_to_create):
#          tempEnvironment0 = arcpy.env.XYResolution
#          arcpy.env.XYResolution = "0.1 Meters"
#          tempEnvironment1 = arcpy.env.XYTolerance
#          arcpy.env.XYTolerance = "0.1 Meters"
        arcpy.CreateFeatureDataset_management(gdb, feature_dataset, "PROJCS['NAD_1983_BC_Environment_Albers',GEOGCS['GCS_North_American_1983',DATUM['D_North_American_1983',SPHEROID['GRS_1980',6378137.0,298.2572221]],PRIMEM['Greenwich',0.0],UNIT['Degree',0.0174532925199433]],PROJECTION['Albers'],PARAMETER['False_Easting',1000000.0],PARAMETER['False_Northing',0.0],PARAMETER['Central_Meridian',-126.0],PARAMETER['Standard_Parallel_1',50.0],PARAMETER['Standard_Parallel_2',58.5],PARAMETER['Latitude_Of_Origin',45.0],UNIT['Meter',1.0]];-13239300 -8610100 10000;-100000 10000;-100000 10000;0.001;0.001;0.001;IsHighPrecision")
#          arcpy.env.XYResolution = tempEnvironment0
#          arcpy.env.XYTolerance = tempEnvironment1
##########################################################################################################
##########################################################################################################
##########################################################################################################

def get_fc_directory_name(the_featureclass_path):
    '''
    This returns a string which is just the directory path of the featureclass 
    that has been passed in.
    
    @param the_featureclass_path: the featureclass path, can be a .shp file too
    @type the_featureclass_path: string
    
    @return: The directory the FC is in.
    @rtype: string
    '''

    the_featureclass_path = the_featureclass_path.lower()
    # find what type of file the input data is stored in
    if the_featureclass_path.find(".mdb") > -1:
        ext_type = ".mdb"
    if the_featureclass_path.find(".shp") > -1:
        ext_type = ".shp"
    if the_featureclass_path.find(".gdb") > -1:
        ext_type = ".gdb"

    
    extension = os.path.splitext(the_featureclass_path)[1]
    directory_string = the_featureclass_path
    while extension != ext_type:
        extension = os.path.splitext(directory_string)[1]
        directory_string = os.path.split(directory_string)[0]
    
    if extension == ".shp" or extension == ".mdb" :# or extension == ".gdb":
        directory_string = os.path.split(directory_string)[0]
    
    return directory_string          



##########################################################################################################
##########################################################################################################
##########################################################################################################

def read_xls_into_list_of_lists(xls_to_read):
    '''
    :Module author: Mark McGirr
    :Date:         May 2013
    :Description:  This reads an XLS file, and returns a list of lists of all its cells. \n
                   Each row in the list contains all the cell values
                   for that row in the xls file.

    :param: *xls_to_read:* the complete path to the spreadsheet you want to read.
    :ptype: string

    :return: All the spreadsheet cell values
    :rtype: string  list or lists

    code_example::


        xls_to_read = os.path.join(project_dir,"site_series_lut.xlsx")
        site_series_list = CE_library.read_xls_into_list_of_lists(xls_to_read)
        del site_series_list[0] # remove the first list entry because its just the column names
        for x in site_series_list:
           print x[0],x[1]


    '''
    arcpy.AddMessage(xls_to_read)
    book = openpyxl.load_workbook(xls_to_read)
    sheet = book.active

    row_count = sheet.max_row
    column_count = sheet.max_column

    return_list = []
    for x in range(1,row_count + 1): #loop through the rows
        new_list = []
        is_spreadsheet_row_blank  = 'yes'

        for y in range(1,column_count + 1): #loop through the columns
            this_field = sheet.cell(x, y).value
            if  str(this_field) == 'None':
                this_field = ""
            else:
                is_spreadsheet_row_blank  = 'no'

            this_field = str(this_field)
            new_list.append(this_field)
#             excel_line_string = ";".join(new_list)

        if is_spreadsheet_row_blank == 'no':
            return_list.append(new_list)
#             excel_line_string = ""

    return (return_list)

##########################################################################################################
##########################################################################################################
##########################################################################################################

def set_border(ws, cell_range):
    #print ('cell_range' , cell_range)
    rows = ws[cell_range]
    for row in rows:
        if row == rows[0][0] or row == rows[0][-1] or row == rows[-1][0] or row == rows[-1][-1]:
            pass
        else:
            row[0].border = Border(left=Side(style='thick'))
            row[-1].border = Border(right=Side(style='thick'))
        for c in rows[0]:
            c.border = Border(top=Side(style='thick'))
        for c in rows[-1]:
            c.border = Border(bottom=Side(style='thick'))
    rows[0][0].border = Border(left=Side(style='thick'), top=Side(style='thick'))
    rows[0][-1].border = Border(right=Side(style='thick'), top=Side(style='thick'))
    rows[-1][0].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
    rows[-1][-1].border = Border(right=Side(style='thick'), bottom=Side(style='thick'))
##########################################################################################################
##########################################################################################################
##########################################################################################################




        
if __name__ == '__main__':
    arcpy.AddMessage("======================================================================")
    arcpy.AddMessage('Running universal overlap tool')


    #=======================================================================
    # Create some test variables so I don't need to run the tool interface.
    # When the interface is run these arguments will be passed in from it.
    #=======================================================================
    input_feature_class = r"W:\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\automated_status_ARCPRO\test_runs\mm_v1\UOT_TestPoly.shp"   # TEXT - analyize_this_featureclass 
    sub_reports_on_this_field = "MAP_BLOCK_"                   # TEXT - create_subreports_on_this_field
    xls_file1_overlap_to_run_from_dropdown_list = "cariboo_specific_debug.xlsx" # TEXT - what_type_of_overlap_to_run
    xls_file1_overlap_to_run_from_user_specified = ""# "r"W:\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\automated_status_ARCPRO\statusing_input_spreadsheets\one_status_common_datasets_debug_version.xlsx"    # TEXT - xls_file_for_analysis_input
    report_header_line1 = r"Marks test analysis"     # TEXT - text_header1
    report_header_line2 = r"Header line 2"           # TEXT - text_header2
    report_header_line3 = r"Header Line Number 3"    # TEXT - text_header3
    report_header_line4 = r"Header line 4"           # TEXT - text_header4
    sub_reports_on_seperate_sheets = r""             # TEXT - True - subreports_on_seperate_sheets
    directory_to_store_output = r""                  # TEXT - directory_to_store_output
    summary_fields_on_seperate_lines = r""           # TEXT - True - summary_fields_on_seperate_lines
    dont_overwrite_existing_data = r'true'           #('true') # TEXT - True - test_dont_overwrite_data_and_maps
    xls_file2_overlap_to_add_for_autostatus_region = ""# "r"W:\srm\wml\Workarea\arcproj\!Williams_Lake_Toolbox_Development\automated_status_ARCPRO\statusing_input_spreadsheets\one_status_" + "Cariboo" + "_specific_debug_version.xlsx"   # TEXT - xls_file_for_analysis_input2 reads a second input spreadsheet into the list
    red_report_header_disclaimer = "This product contains sensitive information.  INTERNAL GOVERNMENT USE ONLY"   # red disclaimer to print at top of report
    suppress_map_creation = r"true"                       # TEXT - Dont try to create the maps on tab 3
    keep_mxd_files = ""                              # TEXT - Dont delete mxd files
    aprx_path = 'not_used_in_this_mode'  
    tool_user = 'mamcgirr'  
    tool_password = 'testicle_password'  

    #------------------------------------------------------------------------------ 
    
    


    #===========================================================================
    # Populate the overlay criteria list.  This list will be passed in by the
    # tool interface if your are running it that way.
    #===========================================================================
    revolt_criteria_to_pass = []
    revolt_criteria_to_pass.append(input_feature_class)  
    revolt_criteria_to_pass.append(sub_reports_on_this_field)
    revolt_criteria_to_pass.append(xls_file1_overlap_to_run_from_dropdown_list) 
    revolt_criteria_to_pass.append(xls_file1_overlap_to_run_from_user_specified)   
    revolt_criteria_to_pass.append(report_header_line1)
    revolt_criteria_to_pass.append(report_header_line2)     
    revolt_criteria_to_pass.append(report_header_line3)     
    revolt_criteria_to_pass.append(report_header_line4)      
    revolt_criteria_to_pass.append(sub_reports_on_seperate_sheets)      
    revolt_criteria_to_pass.append(directory_to_store_output)        
    revolt_criteria_to_pass.append(summary_fields_on_seperate_lines)      
    revolt_criteria_to_pass.append(dont_overwrite_existing_data)        
    revolt_criteria_to_pass.append(xls_file2_overlap_to_add_for_autostatus_region)  
    revolt_criteria_to_pass.append(red_report_header_disclaimer) 
    revolt_criteria_to_pass.append(suppress_map_creation) 
    revolt_criteria_to_pass.append(keep_mxd_files)  
    revolt_criteria_to_pass.append(aprx_path)  
    revolt_criteria_to_pass.append(tool_user)  
    revolt_criteria_to_pass.append(tool_password)  

    #------------------------------------------------------------------------------ 



    universal_overlap_Obj = revolt_tool()
    universal_overlap_Obj.run_revolt_tool(revolt_criteria_to_pass)
