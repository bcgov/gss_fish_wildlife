# autoast is a script for batch processing the automated status tool
# author: wburt
# copyrite Governent of British Columbia
# Copyright 2019 Province of British Columbia

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at 

#    http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.






import os
import shutil
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import geopandas
import arcpy
import datetime
import logging
import traceback
import subprocess
import multiprocessing as mp
from tqdm import tqdm
import sys
import time

# Test Comment

## *** INPUT YOUR EXCEL FILE NAME HERE ***
excel_file = 'Cariboo_replacement_jobs.xlsx'

# Set the job timeout further down. Use CNTL + F to search for JOB_TIMEOUT

# Number of CPUS to use for multiprocessing ** Currently not used
NUM_CPUS = mp.cpu_count()
###############################################################################################################################################################################
# Set up logging

def setup_logging():
    ''' Set up logging for the script '''
    # Create the log folder filename
    log_folder = f'autoast_logs_{datetime.datetime.now().strftime("%Y%m%d")}'

    # Create the log folder in the current directory if it doesn't exits
    if not os.path.exists(log_folder):
        os.mkdir(log_folder)
    
    # Check if the log folder was created successfully
    assert os.path.exists(log_folder), "Error creating log folder, check permissions and path"

    # Create the log file path with the date and time appended
    log_file = os.path.join(log_folder, f'ast_log_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.log')



    # Set up logging config to DEBUG level
    logging.basicConfig(filename=log_file, 
                        level=logging.DEBUG, 
                        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    # Create the logger object and set to the current file name
    logger = logging.getLogger(__name__)

    print("Logging set up")
    logger.info("Logging set up")

    print("Starting Script")
    logger.info("Starting Script")
    
    return logger
###############################################################################################################################################################################



def import_ast(logger):
    # Get the toolbox path from environment variables
    ast_toolbox = os.getenv('TOOLBOX') # File path 

    if ast_toolbox is None:
        print("Unable to find the toolbox. Check the path in .env file")
        logger.error("Unable to find the toolbox. Check the path in .env file")
        exit() 

    # Import the toolbox
    try:
        arcpy.ImportToolbox(ast_toolbox)
        print(f"AST Toolbox imported successfully.")
        logger.info(f"AST Toolbox imported successfully.")
    except Exception as e:
        print(f"Error importing toolbox: {e}")
        logger.error(f"Error importing toolbox: {e}")
        exit()

    # Assign the shapefile template for FW Setup to a variable
    template = os.getenv('TEMPLATE') # File path in .env
    if template is None:
        print("Unable to find the template. Check the path in .env file")
        logger.error("Unable to find the template. Check the path in .env file")
        
    return template
    

###############################################################################################################################################################################
#
# Set up the database connection
#
###############################################################################################################################################################################
def setup_bcgw(logger):
    # Get the secret file containing the database credentials
    SECRET_FILE = os.getenv('SECRET_FILE')

    # If secret file found, load the secret file and display a print message, if not found display an error message
    if SECRET_FILE:
        load_dotenv(SECRET_FILE)
        print(f"Secret file {SECRET_FILE} found")
        logger.info(f"Secret file {SECRET_FILE} found")
    else:
        print("Secret file not found")
        logger.error("Secret file not found")

    # Assign secret file data to variables    
    DB_USER = os.getenv('BCGW_USER')
    DB_PASS = os.getenv('BCGW_PASS')
    

    # If DB_USER and DB_PASS found display a print message, if not found display an error message
    if DB_USER and DB_PASS:
        print(f"Database user {DB_USER} and password found")
        logger.info(f"Database user {DB_USER} and password found")
    else:
        print("Database user and password not found")
        logger.error("Database user and password not found")

    # Define current path of the executing script
    current_path = os.path.dirname(os.path.realpath(__file__))

    # Create the connection folder
    connection_folder = 'connection'
    connection_folder = os.path.join(current_path, connection_folder)

    # Check for the existance of the connection folder and if it doesn't exist, print an error and create a new connection folder
    if not os.path.exists(connection_folder):
        print("Connection folder not found, creating new connection folder")
        logger.info("Connection folder not found, creating new connection folder")
        os.mkdir(connection_folder)

    # Check for an existing bcgw connection, if there is one, remove it
    if os.path.exists(os.path.join(connection_folder, 'bcgw.sde')):
        os.remove(os.path.join(connection_folder, 'bcgw.sde'))

    # Create a bcgw connection
    bcgw_con = arcpy.management.CreateDatabaseConnection(connection_folder,
                                                        'bcgw.sde',
                                                        'ORACLE',
                                                        'bcgw.bcgov/idwprod1.bcgov',
                                                        'DATABASE_AUTH',
                                                        DB_USER,
                                                        DB_PASS,
                                                        'DO_NOT_SAVE_USERNAME')

    print("new db connection created")
    logger.info("new db connection created")


    arcpy.env.workspace = bcgw_con.getOutput(0)

    print("workspace set to bcgw connection")
    logger.info("workspace set to bcgw connection")
    
    secrets = [DB_USER, DB_PASS]
    
    return secrets
###############################################################################################################################################################################

class AST_FACTORY:
    ''' AST_FACTORY creates and manages status tool runs '''
    XLSX_SHEET_NAME = 'ast_config'
    AST_PARAMETERS = {
        0: 'region',
        1: 'feature_layer',
        2: 'crown_file_number',
        3: 'disposition_number',
        4: 'parcel_number',
        5: 'output_directory',
        6: 'output_directory_same_as_input',
        7: 'dont_overwrite_outputs',
        8: 'skip_conflicts_and_constraints',
        9: 'suppress_map_creation',
        10: 'add_maps_to_current',
        11: 'run_as_fcbc',

    }
    
    ADDITIONAL_PARAMETERS = {
        12: 'ast_condition',
        13: 'file_number'
    }
    
    AST_CONDITION_COLUMN = 'ast_condition'
    DONT_OVERWRITE_OUTPUTS = 'dont_overwrite_outputs'
    AST_SCRIPT = ''
    job_index = None  # Initialize job_index as a global variable
    
    def __init__(self, queuefile, db_user, db_pass, logger=None, current_path=None) -> None:
            self.user = db_user
            self.user_cred = db_pass
            self.queuefile = queuefile
            self.jobs = []
            self.logger = logger or logging.getLogger(__name__)
            self.current_path = current_path  
#LOAD JOBS
    def load_jobs(self):
        '''
        load jobs will check for the existence of the queuefile, if it exists it will load the jobs from the queuefile. Checking if they 
        are Complete and if not, it will add them to the jobs  as Queued
        '''
        # NOTE pass job index into load jobs function
        #global job_index
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Loading Jobs...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")

        # Initialize the jobs list to store jobs
        self.jobs = []

        # Check if the queue file exists
        assert os.path.exists(self.queuefile), "Queue file does not exist"
        if os.path.exists(self.queuefile):

            try:
                # Open the Excel workbook and select the correct sheet
                wb = load_workbook(filename=self.queuefile)
                ws = wb[self.XLSX_SHEET_NAME]

                # Get the header (column names) from the first row of the sheet
                header = list([row for row in ws.iter_rows(min_row=1, max_col=None, values_only=True)][0])
                
                # Read all the data rows (starting from the second row to skip the header)
                data = []
                
                for row in ws.iter_rows(min_row=2, max_col=None, values_only=True):
                    print(f'Row is {row}')
                    
                    data.append(row)
                


            
                
                # Iterate over each row of data; enumerate to keep track of the row number in Excel
                for job_index, row_data in enumerate(data):  

                    # Dictionary where key is index key is Job number dictionary is the dictionary of jobs
                    # Send job to processer and include status
                    self.logger.info(f"\n")
                    self.logger.info(f"-------------------------------------------------------------------------------")
                    self.logger.info(f"-                        Load Jobs: Start of Job {job_index}                               -")
                    self.logger.info(f"-------------------------------------------------------------------------------")
                    self.logger.info(f"\n")
                    
                    # Initialize a dictionary to store the job's parameters
                    job = {}
                    self.logger.info('Load Jobs - Creating empty dictionary')
                    ast_condition = None  # Initialize the ast_condition for the current row
                        
                    
                    # Skip any completely blank rows
                    if all((value is None or str(value).strip() == '') for value in row_data):
                        print(f"Load Jobs - Skipping blank row at job index ({job_index}) ")
                        self.logger.info(f"Load Jobs - Skipping blank row at index ({job_index}) ")
                        continue  # Skip this row entirely


                    # Loop through each column header and corresponding value in the current row
                    for key, value in zip(header, row_data):
                        # If the key isn't empty and equals "ast_condition", assign the value to ast_condition
                        if key is not None and key.lower() == self.AST_CONDITION_COLUMN.lower():
                            ast_condition = value if value is not None else ""

                        # Assign an empty string to any None values
                        value = "" if value is None else value

                        # Assign the value to the job dictionary if the key is not None
                        if key is not None:
                            job[key] = value

                    # Skip if marked as "COMPLETE"
                    if ast_condition.upper() == 'COMPLETE':
                        print(f"Skipping job {job_index} as it is marked COMPLETE.")
                        self.logger.info(f"Load Jobs - Skipping job {job_index} as it is marked COMPLETE.")
                        # continue  # Skip this job as it's already marked as COMPLETE

                    # Check if the ast_condition is None, empty, or not 'COMPLETE'
                    if ast_condition is None or ast_condition.strip() == '' or ast_condition.upper() != 'COMPLETE':
                        # Assign 'Queued' to the ast_condition and update the job dictionary
                        ast_condition = 'Queued'
                        
                        # Assign the updated ast_condition to the job dictionary (queued)
                        job[self.AST_CONDITION_COLUMN] = ast_condition
                        self.logger.info(f"Load Jobs - (Queued assigned to Job ({job_index}) is ({ast_condition})")

                        # Immediately update the Excel sheet with the new condition
                        #LOAD JOBS ADD_JOB_RESULT FUNCTION IS CALLED HERE
                        try:
                            self.add_job_result(job_index, ast_condition)
                            self.logger.info(f"Load Jobs - Added job condition '{ast_condition}' for job {job_index} to jobs list")
                        except Exception as e:
                            print(f"Error updating Excel sheet at row {job_index}: {e}")
                            self.logger.error(f"Load Jobs - Error updating Excel sheet at row {job_index}: {e}")
                            continue

                        # Classify the input type for the job
                        try:
                            self.logger.info(f"Classifying input type for job {job_index}")
                            self.classify_input_type(job)
 
                        except Exception as e:
                            print(f"Error classifying input type for job {job}: {e}")
                            self.logger.error(f"Error classifying input type for job {job}: {e}")
                            
                    # Add the job to the jobs list after all checks and processing
                    self.jobs.append(job)
                    print(f"Load Jobs - Job Condition for job ({job_index}) is not Complete: Writing ({ast_condition}) to ast_contion. Adding job: {job_index} to jobs list")
                    self.logger.info(f"Load Jobs - Job Condition is not Complete ({ast_condition}), adding job: {job_index} to jobs list")

                    # print(f"Load Jobs - Job dictionary is {job}")
                    # self.logger.info(f"Load Jobs - Job {job_index} dictionary is {job}")
                    self.logger.info(f"\n")
                    self.logger.info(f"-------------------------------------------------------------------------------")
                    self.logger.info(f"-                        End of Job {job_index}                                -")
                    self.logger.info(f"-------------------------------------------------------------------------------")
                    self.logger.info(f"\n")
                    
                    
            except FileNotFoundError as e:
                print(f"Error: Queue file not found - {e}")
                self.logger.error(f"Error: Queue file not found - {e}")
            except Exception as e:
                print(f"Unexpected error loading jobs: {e}")
                self.logger.error(f"Unexpected error loading jobs: {e}")

            return self.jobs


    def classify_input_type(self, job):
        '''Classify the input type and process accordingly.'''

        if job.get('feature_layer'):
            print(f'Feature layer found: {job["feature_layer"]}')
            self.logger.info(f'Classifying Input Type - Feature layer found: {job["feature_layer"]}')
            feature_layer_path = job['feature_layer']
            print(f"Processing feature layer: {feature_layer_path}")
            self.logger.info(f"Classifying Input Type - Processing feature layer: {feature_layer_path}")

            if feature_layer_path.lower().endswith('.kml'):
                print('KML found, building AOI from KML')
                self.logger.info('Classifying Input Type - KML found, building AOI from KML')
                job['feature_layer'] = self.build_aoi_from_kml(feature_layer_path)

            elif feature_layer_path.lower().endswith('.shp'):
                if job.get('file_number'):
                    print(f"File number found, running FW setup on shapefile: {feature_layer_path}")
                    self.logger.info(f"Classifying Input Type - File number found, running FW setup on shapefile: {feature_layer_path}")
                    new_feature_layer_path = self.build_aoi_from_shp(job, feature_layer_path)
                    job['feature_layer'] = new_feature_layer_path
                else:
                    print('No FW File Number provided for the shapefile, using original shapefile path')
                    self.logger.info('Classifying Input Type - No FW File Number provided, using original shapefile path')
            else:
                print(f"Unsupported feature layer format: {feature_layer_path}")
                self.logger.warning(f"Classifying Input Type - Unsupported feature layer format: {feature_layer_path} - Marking job as Failed")
                self.add_job_result(job, 'Failed')
        else:
            print('No feature layer provided in job')
            self.logger.warning('Classifying Input Type - No feature layer provided in job')

#ADD JOB RESULT                        
    def add_job_result(self, job_index, condition):
        ''' 
        Function adds result information to the Excel spreadsheet. If the job is successful, it will update the ast_condition column to "COMPLETE",
        if the job failed, it will update the ast_condition column to "Failed".
        '''

        print("Running Add Job Results...")
        self.logger.info("\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Running Add Job Results from Load Jobs Function")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info("\n")
     

        try:
            # Load the workbook
            wb = load_workbook(filename=self.queuefile)
            self.logger.info(f"Add Job Result - Workbook loaded")
            
            # Load the correct worksheet
            ws = wb[self.XLSX_SHEET_NAME]

            # Read the header index for the ast_condition column
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            
            # Check if 'AST CONDITION COLUMN' exists in the header. If it is not found, raise a ValueError
            # if ast condition column IS found, log a message
            if self.AST_CONDITION_COLUMN not in header:
                raise ValueError(f"'{self.AST_CONDITION_COLUMN}' column not found in the spreadsheet.")
            
            if self.AST_CONDITION_COLUMN in header:
                self.logger.info(f"Add Job Result - '{self.AST_CONDITION_COLUMN}' column found in the spreadsheet.")
                
            # # Check if 'DONT OVERWRITE OUTPUTS' exists in the header
            # if self.DONT_OVERWRITE_OUTPUTS not in header:
            #     raise ValueError(f"'{self.DONT_OVERWRITE_OUTPUTS}' column not found in the spreadsheet.")
            
            # if self.DONT_OVERWRITE_OUTPUTS in header:
            #     self.logger.info(f"Add Job Result - '{self.DONT_OVERWRITE_OUTPUTS}' column found in the spreadsheet.")
            
            # Find the ast condition column and assign it to the correct index
            ast_condition_index = header.index(self.AST_CONDITION_COLUMN) + 1  # +1 because Excel columns are 1-indexed

            # # Find the dont_overwrite_outputs column and assign it to the correct index
            # dont_overwrite_outputs_index = header.index(self.DONT_OVERWRITE_OUTPUTS) + 1  # +1 because Excel columns are 1-indexed
            
            # Calculate the actual row index in Excel, +2 to account for header and 0-index
            excel_row_index = job_index + 2  # NOTE I changed this to +1 and it changes the ast_condition header row to Failed. So it must stay at +2
            self.logger.info(f"Add Job Result - Calculated Excel row index as {excel_row_index} for job index {job_index}")
            
            # Check if the row is blank before updating,  If all cell values in row_values are either None or empty strings, then all() will return True, indicating that the row is blank.
            row_values = []
            for col in range(1, len(header) + 1):
                cell_value = ws.cell(row=excel_row_index, column=col).value
                row_values.append(cell_value)
            if all(value is None or str(value).strip() == '' for value in row_values):
                print(f"Row {excel_row_index} is blank, not updating.")
                self.logger.info(f"Add_Job_Result -Job {job_index} / Row {excel_row_index} ")
                return  # Do not update if the row is blank

            # Update the ast condition for the specific job to the new condition (failed, queued, complete)
            ws.cell(row=excel_row_index, column=ast_condition_index, value=condition)

            # # if the condition in AST_CONDITION_COLUMN is 'Requeued" then go to the dont overwrite output column and change false to true
            # if condition == 'Requeued':
            #     # print(f"Add Job Result - Job {job_index} failed, updating condition to 'Requeued'.  **CHANGED JOB INDEX +1 to JOB INDEX ***") #NOTE CHANGED JOB INDEX + 1 to JOB INDEX
            #     self.logger.info(f"Add Job Result - Job {job_index} (Row {excel_row_index})  updating condition to 'Requeued'.") 
            #     ws.cell(row=excel_row_index, column=dont_overwrite_outputs_index, value="True")
            #     self.logger.info(f"Add Job Result - Job {job_index} (Row {excel_row_index})  updating dont_overwrite_outputs to 'True'.")
            
            # Save the workbook with the updated condition
            self.logger.info(f"Add Job Result - Updated Job {job_index} with condition '{condition}'.")
            wb.save(self.queuefile)
            self.logger.info(f"Add Job Result - Saving Workbook with updated condition")
            print(f"Updated row {excel_row_index} with condition '{condition}'.")

        except FileNotFoundError as e:
            print(f"Error: Queue file not found - {e}")
            self.logger.error(f"Error: Queue file not found - {e}")

        except ValueError as e:
            print(f"Error: {e}")
            self.logger.error(f"Error: {e}")

        except IndexError as e:
            print(f"Error: Index out of range when accessing row {excel_row_index} - {e}")
            self.logger.error(f"Error: Index out of range when accessing row {excel_row_index} - {e}")

        except PermissionError as e:
            print(f"Error: Permission denied when trying to access the Excel file - {e}")
            self.logger.error(f"Error: Permission denied when trying to access the Excel file - {e}")

        except Exception as e:
            print(f"Unexpected error while adding job result: {e}")
            self.logger.error(f"Unexpected error while adding job result: {e}")

#BATCH AST
    def batch_ast(self):
        '''
        Uses multiprocessing to run the NUMBER_OF_JOBS in parallel.
        '''
        self.logger.info(f"\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Batch AST: Batching Jobs with Multiprocessing...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info(f"\n")
        
        import time
        
        # Set job timeout to 6 hours
        JOB_TIMEOUT = 21600  # 6 hours in seconds
        self.logger.info(f"Batch Ast: Job Timeout set to {JOB_TIMEOUT} seconds")
        print(f"Batch Ast: Job Timeout set to {JOB_TIMEOUT} seconds")

        processes = []
        manager = mp.Manager()
        return_dict = manager.dict()

        for job_index, job in enumerate(self.jobs):
            self.logger.info(f"Batch Ast: Starting job {job_index}")
            print(f"Batch Ast: Starting job {job_index} Job ({job})")

            # if ast condition is queued or requeued, run the job
            if job.get(self.AST_CONDITION_COLUMN) in ['Queued', 'Requeued']: #NOTE Add QUEUED AFTER TESTING***
                
                # Start each job in a separate process
                p = mp.Process(target=process_job_mp, args=(self, job, job_index, self.current_path, return_dict))
                
                # Append the process object to the processes list and job_index. This list keeps track of all the processes and their corresponding job indices.
                processes.append((p, job_index))
                
                # Start method is called on the process object p. This begins the execution of the job in a separate process.
                p.start()
                self.logger.info(f"Batch Ast: {job.get(self.AST_CONDITION_COLUMN)} Job {job_index}.....Multiproccessing started......")
                print(f"Batch Ast: Queued Job...Multiproccessing started......")
                


        # Monitor and enforce timeouts
        timeout_failed_counter = 0
        success_counter = 0
        worker_failed_counter = 0
        other_exception_failed_counter = 0
        for process, job_index in processes:
            
            # Join the process to timeout which waits for the process to complete within the timeout
            process.join(JOB_TIMEOUT)
            
            # If the process exceeds the timeout, terminate the process and mark the job as failed
            if process.is_alive():
                
                print(f"Batch Ast: Job {job_index} exceeded timeout. Terminating process.")
                self.logger.warning(f"Batch Ast: Job {job_index} exceeded timeout. Terminating process.")
                
                # End the hung up job
                process.terminate()
                
                # Call the join method again to ensure the process is terminated
                process.join()
                
                # Call add job result and update the job as failed
                self.add_job_result(job_index, 'Failed') 
                
                # Increase the job timeout counter
                timeout_failed_counter+= 1
                self.logger.error(f"Batch Ast: Job {job_index} exceeded timeout. Marking as Failed. Failed counter is {timeout_failed_counter}")
                
            else:
                # Get the result of the job from (return_dict). 
                # If the result is 'Success', increment the success_counter and call the add_job_result method to mark the job as 'COMPLETE'
                
                result = return_dict.get(job_index)
                if result == 'Success':
                    success_counter += 1
                    self.add_job_result(job_index, 'COMPLETE')
                    print(f"Batch Ast: Job {job_index} completed successfully.")
                    self.logger.info(f"Batch Ast: Job {job_index} completed successfully. Success counter is {success_counter}")
                
                elif result == 'Failed':
                    
                    # If the result is 'Failed', increment the other_failed_counter and mark the job as 'Failed' (Other failed counter means it failed due to something other than a timeout)
                    # Job failed due to an exception in the worker
                    self.add_job_result(job_index, 'Failed')
                    worker_failed_counter += 1
                    print(f"Batch Ast: Job {job_index} failed due to an exception.")
                    self.logger.error(f"Batch AST: Job {job_index} failed due to an exception in the Worker. Other exception failed counter is {worker_failed_counter}")
                
                else:
                    # Handle unexpected cases
                    self.add_job_result(job_index, 'Unknown Error')
                    other_exception_failed_counter += 1
                    print(f"Batch Ast: Job {job_index} failed with unknown status.")
                    self.logger.error(f"Batch AST: Job {job_index} failed with unknown status. Other Exception failed counter is {other_exception_failed_counter}")
         
        self.logger.info('\n')    
        self.logger.info("Batch Ast Complete - Check separate worker log file for more details")
    


# This is the newer version of the re_load_failed_jobs function from the autoastv2Script unedited
# NOTE ** Reload failed jobs may be able to be incorporated into load failed jobs to tighten up the script
#RELOAD JOBS
    def re_load_failed_jobs_V2(self):
        '''
        re load failed jobs will check for the existence of the queuefile, if it exists it will load the jobs from the queuefile. Checking if they 
        are Failed and if they are, will change Dont Overwrite Outputs to True and add them to the jobs list as Queued
        '''
        self.logger.info("\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Re loading Failed Jobs V2.....")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info("\n")


        # Initialize the jobs list to store jobs
        self.jobs = []

        # Check if the queue file exists
        assert os.path.exists(self.queuefile), "Queue file does not exist"
        if os.path.exists(self.queuefile):

            try:
                # Open the Excel workbook and select the correct sheet
                wb = load_workbook(filename=self.queuefile)
                ws = wb[self.XLSX_SHEET_NAME]
                
                
                print(f'Workbook loaded is {wb}')   
                self.logger.info(f'Re load Failed Jobs: Workbook loaded is {wb}') 
                
                # Get the header (column names) from the first row of the sheet
                header = list([row for row in ws.iter_rows(min_row=1, max_col=None, values_only=True)][0])

                # Read all the data rows (starting from the second row to skip the header)
                data = []
                self.logger.info(f'Re load Failed Jobs: Reading all data rows and building data list')
                for row in ws.iter_rows(min_row=2, max_col=None, values_only=True):
                    print(f'Row is {row}')
                    
                    data.append(row)

                # Iterate over each row of data; enumerate to keep track of the row number in Excel
                self.logger.info(f'Re load Failed Jobs: Iterating over each row of data')
                for job_index, row_data in enumerate(data):  # Start from 2 to account for Excel header
                    
                    self.logger.info(f"\n")
                    self.logger.info(f"------------------------------------------------------------------------------------")
                    self.logger.info(f"-                        Re Load Failed Jobs: Start of Job {job_index}                               -")
                    self.logger.info(f"------------------------------------------------------------------------------------")
                    self.logger.info(f"\n")
                    
                                    
                    # Initialize a dictionary to store the job's parameters
                    job = {}
                    self.logger.info('Re load Jobs - Creating empty dictionary')
                    ast_condition = ''  # Initialize the ast_condition for the current row
                        
                    # Skip any completely blank rows
                    if all((value is None or str(value).strip() == '') for value in row_data):
                        print(f"Re Load Failed Jobs: Skipping blank row at index {job_index}")
                        self.logger.info(f"Re Load Failed Jobs: Skipping blank row at index {job_index}")
                        continue  # Skip this row entirely


                    # Loop through each column header and corresponding value in the current row
                    for key, value in zip(header, row_data):
                        # Check if the key corresponds to the ast_condition column
                        if key is not None and key.lower() == self.AST_CONDITION_COLUMN.lower():
                            ast_condition = value if value is not None else ""

                        # Assign an empty string to any None values
                        value = "" if value is None else value

                        # Assign the value to the job dictionary if the key is not None
                        if key is not None:
                            # logger.info(f"Re Load Failed Jobs: Assigning values to job dictionary")
                            print(f"Re Load Failed Jobs: Assigning values to job dictionary")
                            job[key] = value

                    # Skip if marked as "COMPLETE"
                    if ast_condition.upper() == 'COMPLETE':
                        print(f"Re Load Failed Jobs: Skipping job {job_index} as it is marked COMPLETE.")
                        self.logger.info(f"Re Load Failed Jobs: Adding Complete to dictionary Skipping job {job_index} as it is marked COMPLETE.")
                        # continue  
                        # ast_condition = 'COMPLETE'    
                    
                    # **Only requeue jobs that are marked as 'FAILED'**
                    elif ast_condition.upper() == 'FAILED':
                        self.logger.info(f"Re Load Failed Jobs: Requeuing {job_index} as it is marked Failed.")
                        ast_condition = 'Requeued'
                    
                    else:
                        self.logger.warning(f"Re Load Failed Jobs: Job {job_index} is not marked as Complete or Failed. Please check the workbook. Skipping this job.")
                        # continue
                        ast_condition = 'ERROR'
                    
                    # Assign updated condition to the job dictionary
                    job[self.AST_CONDITION_COLUMN] = ast_condition
                        
                    print(f"Re Load Failed Jobs: Job {job_index} is marked as Failed, re-assigning ast condition to {ast_condition}")
                    self.logger.info(f"Re Load Failed Jobs: Job {job_index}'s ast condition has been updated as '{ast_condition}'")


                    # Immediately update the Excel sheet with the new condition
                    try:
                        self.add_job_result(job_index, ast_condition)
                        self.logger.info(f"Re load Jobs - Added job condition '{ast_condition}' for job {job_index} to jobs list")
                    except Exception as e:
                        print(f"Error updating Excel sheet at row {job_index}: {e}")
                        self.logger.error(f"Re load Jobs - Error updating Excel sheet at row {job_index}: {e}")
                        self.logger.error(traceback.format_exc())
                        continue
                        
                        
                    # Check the condition of DONT_OVERWRITE_OUTPUTS
                    current_value = job.get(self.DONT_OVERWRITE_OUTPUTS, '')

                    # If DONT_OVERWRITE_OUTPUTS is anything but True, change it to 'True'
                    if current_value != 'True':
                        # Log the current state before changing
                        if current_value == 'False':
                            print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is False, changing to True")
                            self.logger.info(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is False, changing to True")
                        elif current_value == '':
                            print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is empty, changing to True")
                            self.logger.error(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is empty, changing to True")
                        else:
                            print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is '{current_value}', changing to True")
                            self.logger.warning(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is '{current_value}', changing to True")
                        
                        # Set the value to 'True'
                        job[self.DONT_OVERWRITE_OUTPUTS] = "True"
                    
                    # If DONT_OVERWRITE_OUTPUTS is already 'True, don't change it. 
                    else:
                        # If it's already 'True', log that no change is needed
                        print(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is already True, no change needed")
                        self.logger.info(f"Re Load Failed Jobs: Job {job_index} DONT_OVERWRITE_OUTPUTS is already True, no change needed")

                    # Add the job to the jobs list after all checks and processing
                    self.jobs.append(job)
                    print(f"Re load Jobs - Job Condition is not Complete ({ast_condition}), adding job: {job_index} to jobs list")
                    self.logger.info(f"Re load Jobs - Job Condition is not Complete ({ast_condition}), adding job: {job_index} to jobs list")
                    self.logger.info(f"\n")
                    self.logger.info(f"------------------------------------------------------------------------------------")   
                    self.logger.info(f" Job list is {job}")
                    self.logger.info(f"------------------------------------------------------------------------------------")
                    self.logger.info(f"\n")
                    
                    
                    
                    print(f"Re Load Jobs - Job dictionary is {job}")
                    self.logger.info(f"Re load Jobs - Job {job_index} dictionary is {job}")
                                
            except FileNotFoundError as e:
                print(f"Error: Queue file not found - {e}")
                self.logger.error(f"Re Load Failed Jobs Error: Queue file not found - {e}")
                self.logger.error(traceback.format_exc())
            except Exception as e:
                print(f"Unexpected error re loading jobs: {e}")
                self.logger.error(f"Re Load Failed Jobs Unexpected error loading jobs: {e}")
                self.logger.error(traceback.format_exc())

        return self.jobs   

    def create_new_queuefile(self):
        '''write a new queuefile with preset header'''

        
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Creating New Queuefile...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")

        
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.XLSX_SHEET_NAME
        headers = list(self.AST_PARAMETERS.values())
        headers.append(self.AST_CONDITION_COLUMN)
        for h in headers:
            c = headers.index(h) + 1
            ws.cell(row=1, column=c).value = h
        wb.save(self.queuefile)

    def build_aoi_from_kml(self, aoi):
        "Write shp file for temporary use"

        # Ensure the KML file exists
        if not os.path.exists(aoi):
            raise FileNotFoundError(f"The KML file '{aoi}' does not exist.")

        print("Building AOI from KML")
        self.logger.info("Building AOI from KML")
        from fiona.drvsupport import supported_drivers
        supported_drivers['LIBKML'] = 'rw'
        tmp = os.getenv('TEMP')
        if not tmp:
            raise EnvironmentError("TEMP environment variable is not set.")
        bname = os.path.basename(aoi).split('.')[0]
        fc = bname.replace(' ', '_')
        out_name = os.path.join(tmp, bname + '.gdb')
        if os.path.exists(out_name):
            shutil.rmtree(out_name, ignore_errors=True)
        df = geopandas.read_file(aoi)
        df.to_file(out_name, layer=fc, driver='OpenFileGDB')

        #DELETE
        print(f' kml ouput is {out_name} / {fc}')
        self.logger.info(f' kml ouput is {out_name} / {fc}')
        return out_name + '/' + fc

    def build_aoi_from_shp(self, job, feature_layer_path):
        """This is snippets of Mike Eastwoods FW Setup Script, if run FW Setup is set to true **Not sure if we need this
        as an option or just make it standard.
        This function will take the raw un-appended shapefile and run it through the FW Setup Script"""

        # Mike Eastwoods FW Setup Script
        print("Processing shapefile using FW Setup Script")
        self.logger.info("Processing shapefile using FW Setup Script")
        
        fsj_workspace = os.getenv('FSJ_WORKSPACE')
        arcpy.env.workspace = fsj_workspace
        arcpy.env.overwriteOutput = False

        # Check if there is a file path in Feature Layer
        if feature_layer_path:
            print(f"Processing feature layer: {feature_layer_path}")
            self.logger.info(f"Processing feature layer: {feature_layer_path}")

        # Check to see if a file number was entered in the excel sheet, if so, use it to name the output directory and authorize the build_aoi_from_shp function to run
        file_number = job.get('file_number')

        if not file_number:
            raise ValueError("Error: File Number is required if you are putting in a shapefile that has not been processed in the FW Setup Tool.")
        else:
            print(f"Running FW Setup on File Number: {file_number}")
            self.logger.info(f"Running FW Setup on File Number: {file_number}")

        # Convert file_number to string and make it uppercase
        file_number_str = str(file_number).upper()

        # Calculate date variables
        date = datetime.date.today()
        year = str(date.year)

        # Set variables
        base = arcpy.env.workspace
        baseYear = os.path.join(base, year)
        outName = file_number_str
        geometry = "POLYGON"
    
        m = "SAME_AS_TEMPLATE"
        z = "SAME_AS_TEMPLATE"
        spatialReference = arcpy.Describe(template).spatialReference

        # ===========================================================================
        # Create Folders
        # ===========================================================================

        print("Creating FW Setup folders . . .")
        self.logger.info("Creating FW Setup folders . . .")
        outName = file_number_str

        # Create path to folder location
        fileFolder = os.path.join(baseYear, outName)
        shapeFolder = fileFolder
        outPath = shapeFolder
        if os.path.exists(fileFolder):
            print(outName + " folder already exists.")
            self.logger.info(outName + " folder already exists.")
        else:
            os.mkdir(fileFolder)

        # ===========================================================================
        # Create Shapefile(s) and add them to the current map
        # ===========================================================================

        print("Creating Shapefiles using FW Setup . . .")
        self.logger.info("Creating Shapefiles using FW Setup . . .")
        if os.path.isfile(os.path.join(outPath, outName + ".shp")):
            print(os.path.join(outPath, outName + ".shp") + " already exists")
            self.logger.info(os.path.join(outPath, outName + ".shp") + " already exists")
            print("Exiting without creating files")
            self.logger.info("Exiting without creating files")
            return os.path.join(outPath, outName + ".shp")
        else:
            # Creating template shapefile
            create_shp = arcpy.management.CreateFeatureclass(outPath, outName, geometry, template, m, z, spatialReference)
            # Append the newly created shapefile with area of interest
            append_shp = arcpy.management.Append(feature_layer_path, create_shp, "NO_TEST")
            print("Append Successful")
            self.logger.info("Append Successful")
            # Making filename for kml
            create_kml = os.path.join(outPath, outName + ".kml")
            # Make layer for kml to be converted from 
            layer_shp = arcpy.management.MakeFeatureLayer(append_shp, outName)
            # Populate the shapefile                          
            arcpy.conversion.LayerToKML(layer_shp, create_kml)
            # Send message to user that kml has been created
            print("kml created: " + create_kml)
            self.logger.info("kml created: " + create_kml)

            print(f"FW Setup complete, returned shapefile is {os.path.join(outPath, outName + '.shp')}")
            self.logger.info(f"FW Setup complete, returned shapefile is {os.path.join(outPath, outName + '.shp')}")

            return os.path.join(outPath, outName + ".shp")
        

    def capture_arcpy_messages(self):
        ''' Re assigns the arcpy messages  (0 for all messages, 1 for warnings, and 2 for errors) to variables and passes them to the logger'''
        
        arcpy_messages = arcpy.GetMessages(0) # Gets all messages
        arcpy_warnings = arcpy.GetMessages(1) # Gets all warnings only
        arcpy_errors = arcpy.GetMessages(2) # Gets all errors only
        
        if arcpy_messages:
            self.logger.info(f'ast_toobox arcpy messages: {arcpy_messages}')
        if arcpy_warnings:
            self.logger.warning(f'ast_toobox arcpy warnings: {arcpy_warnings}')
        if arcpy_errors:
            self.logger.error(f'ast_toobox arcpy errors: {arcpy_errors}')   


###############################################################################################################################################################################

def process_job_mp(ast_instance, job, job_index, current_path, return_dict):
    import os
    import arcpy
    import datetime
    import logging
    import multiprocessing as mp
    import traceback

    logger = logging.getLogger(f"Process Job Mp: worker_{job_index}")

    logger.info("##########################################################################################################################")
    logger.info("#")
    logger.info("Running Multiprocessing Worker Function.....")
    logger.info("#")
    logger.info("##########################################################################################################################")

    print(f"Process Job Mp: Processing job {job_index}: {job}")

    # Set up logging folder in the worker process
    logger.info(f"Process Job Mp: Worker process {mp.current_process().pid} started for job {job_index}")
    log_folder = os.path.join(current_path, f'autoast_logs_{datetime.datetime.now().strftime("%Y%m%d")}')
    if not os.path.exists(log_folder):
        os.mkdir(log_folder)
        logger.info(f"Process Job Mp: Created log folder {log_folder}")

    # Generate a unique log file name per process
    log_file = os.path.join(
        log_folder,
        f'ast_worker_log_{datetime.datetime.now().strftime("%Y_%m_%d_%H%M%S")}_{mp.current_process().pid}_job_{job_index}.log'
    )
    logger.info(f"Process Job Mp: Log file for worker process is: {log_file}")
    
    # Set up logging config in the worker process
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,  # Set level to DEBUG to capture all messages
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    try:
        # Re-import the toolbox in each process
        ast_toolbox = os.getenv('TOOLBOX')  # Get the toolbox path from environment variables
        if ast_toolbox:
            arcpy.ImportToolbox(ast_toolbox)
            print(f"Process Job Mp: AST Toolbox imported successfully in worker.")
            logger.info(f"Process Job Mp: AST Toolbox imported successfully in worker.")
        else:
            raise ImportError("Process Job Mp: AST Toolbox path not found. Ensure TOOLBOX path is set correctly in environment variables.")

        # Prepare parameters
        params = []

        # Convert 'true'/'false' strings to booleans
        for param in AST_FACTORY.AST_PARAMETERS.values():
            value = job.get(param)
            if isinstance(value, str) and value.lower() in ['true', 'false']:
                value = True if value.lower() == 'true' else False
            params.append(value)
        
        #NOTE: This is where the output directory is set
        # Get the output directory from the job
        output_directory = job.get('output_directory')

        # # If output_directory is not provided
        # if not output_directory:
        #     # Check if 'output directory is same as input directory' is set to True
        #     output_same_as_input = job.get('output_directory_is_same_as_input_directory')
        #     if output_same_as_input == True or str(output_same_as_input).lower() == 'true':
        #         # Use the input_directory as output_directory
        #         #NOTE This handling is already present in the AST Tool
        #         output_directory = job.get('input_directory')
        #         if not output_directory:
        #             raise ValueError(f"Process Job Mp: 'Input Directory' is required when 'Output Directory is same as Input Directory' is True for job {job_index}.")
        #         job['output_directory'] = output_directory
        #         logger.info(f"Process Job Mp: Output directory is same as input directory for job {job_index}. Using: {output_directory}")
        #     else:
        #         # If there was no output directory provided and 'output directory is same as input directory' is False
        #         # Set the default output directory to a default location (This can be changed later) This will prevent the job from failing due to a user error
                
        #         #DELETE This was put in for testing so that it's easy to delete all outputs from one place at once. 
        #         DEFAULT_DIR = os.getenv('DIR')
        #         output_directory = os.path.join("T:", f'job{job_index}')
        #         job['output_directory'] = output_directory
        #         logger.warning(f"Process Job Mp: Output directory not provided for job {job_index}. Using default path: {output_directory}")
        # else:
        #     # Output directory is provided
        #     job['output_directory'] = output_directory

        # Create the output directory if the user put in a path but failed to create the output directory in Windows explorer
        if output_directory and not os.path.exists(output_directory):
            try:
                os.makedirs(output_directory)
                print(f"Output directory '{output_directory}' created.")
                logger.warning(f"Process Job Mp: Output directory doesn't exist for job ({job_index}).")
                logger.warning(f"\n")
                logger.warning(f"'{output_directory}' created.")
            except OSError as e:
                raise RuntimeError(f"Failed to create the output directory '{output_directory}'. Check your permissions: {e}")


        # Ensure that region has been entered otherwise job will fail
        if not job.get('region'):
            raise ValueError("Process Job Mp: Region is required and was not provided. Job Failed")

        # Log the parameters being used
        logger.debug(f"Process Job Mp: Job Parameters: {params}")

        # Run the ast tool
        logger.info("Process Job Mp: Running MakeAutomatedStatusSpreadsheet_ast...")
        arcpy.MakeAutomatedStatusSpreadsheet_ast(*params)
        logger.info("Process Job Mp: MakeAutomatedStatusSpreadsheet_ast completed successfully.")
        ast_instance.add_job_result(job_index, 'COMPLETE')

        # Capture and log arcpy messages
        logger.info("Process Job Mp: Capturing arcpy messages...")
        arcpy_messages = arcpy.GetMessages(0)
        arcpy_warnings = arcpy.GetMessages(1)
        arcpy_errors = arcpy.GetMessages(2)

        if arcpy_messages:
            logger.info(f'arcpy messages: {arcpy_messages}')
        if arcpy_warnings:
            logger.warning(f'arcpy warnings: {arcpy_warnings}')
        if arcpy_errors:
            logger.error(f'arcpy errors: {arcpy_errors}')
        
        # Indicate success
        return_dict[job_index] = 'Success'  

    except Exception as e:
        # Indicate failure
        return_dict[job_index] = 'Failed'
        logger.error(f"Process Job Mp: Job {job_index} failed with error: {e}")
        logger.debug(traceback.format_exc())
        exc_type, exc_value, exc_traceback = sys.exc_info()
        traceback_str = ''.join(traceback.format_exception(exc_type, exc_value, exc_traceback))
        logger.error(f"Process Job Mp: Job {job_index} failed with error: {e}")
        logger.error(f"Process Job Mp: Traceback:\n{traceback_str}")


#################################################################################################################################################################################
if __name__ == '__main__':
    current_path = os.path.dirname(os.path.realpath(__file__))

    # Call the setup_logging function to log the messages
    logger = setup_logging()

    # Load the default environment
    load_dotenv()

    # Call the import_ast function to import the AST toolbox
    template = import_ast(logger)

    # Call the setup_bcgw function to set up the database connection
    secrets = setup_bcgw(logger)

    # Create the path for the queuefile
    qf = os.path.join(current_path, excel_file)

    # Create an instance of the Ast Factory class, assign the queuefile path and the bcgw username and passwords to the instance
    ast = AST_FACTORY(qf, secrets[0], secrets[1], logger, current_path)

    if not os.path.exists(qf):
        print("Main: Queuefile not found, creating new queuefile")
        logger.info("Main: Queuefile not found, creating new queuefile")
        ast.create_new_queuefile()
        
    # Load the jobs using the load_jobs method. This will scan the excel sheet and assign to "jobs"    
    jobs = ast.load_jobs()
    
    ast.batch_ast()
    
    ast.re_load_failed_jobs_V2()
    
    ast.batch_ast()
    
    print("Main: AST Factory COMPLETE")
    logger.info("Main: AST Factory COMPLETE")

